"""
LNG update PREFLIGHT - validate the master's STRUCTURE before extracting/publishing.

Run before every refresh. Locates the Monthly-Exports blocks by anchor (catching
layout changes) and compares the project roster against the last known-good
manifest (INPUT/lng_manifest.json):

  exit 0  -> SAFE to proceed. Either no structural change, or new project(s) in an
             EXISTING region/country (their names are printed so colour/region
             mapping can be checked after the build).
  exit 2  -> STOP, needs review: a missing anchor / layout change, a removed or
             renamed project, or a NEW region/country. Call Claude before updating.

Roster changes are found by NAME (a set diff), NOT by row position, so a project
inserted anywhere in the list is reported precisely.

  py -3 src/preflight_lng.py            # check against the manifest
  py -3 src/preflight_lng.py --write    # (re)initialise the manifest from the master
                                        #   (done automatically after an approved update)
"""
import json
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

from extract_lng_input import SRC, REGION_OF, COUNTRIES, _clean, locate_exports

ROOT = Path(__file__).resolve().parent.parent
MANIFEST = ROOT / "WORKING" / "lng_manifest.json"   # pipeline state (not in INPUT)

# Assumptions-tab columns we fingerprint per project so a change to any of them is
# REPORTED on the next update (the values already flow through automatically; this
# is the visible confirmation / "did I paste the right file?" check). Status is
# tracked separately (it can be a structural-ish change). Keyed col -> label.
ASSUMP_COLS = [
    (7, "CoS"), (8, "util forecast"), (9, "util decline"),
    (6, "start"), (10, "decline start"), (4, "unrisked (mmt)"),
]
# Ownership lives across several columns; compared as one group (operator col 23,
# partners col 24, primary stake col 25, partner stakes S1..S7 cols 26-32).
OWNERSHIP_COLS = [23, 24, 25] + list(range(26, 33))


def _av(v):
    """Normalise an Assumptions cell to a stable, JSON-serialisable value."""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, float):
        return round(v, 6)
    if isinstance(v, str):
        return v.strip()
    return v


def _fmt(v):
    return "(blank)" if v in (None, "") else str(v)


def snapshot():
    """Structural fingerprint of the master: project roster (name -> country/region/
    status) + the country/region sets + the anchor block positions."""
    wb = openpyxl.load_workbook(SRC, data_only=True)
    try:
        me = wb["Monthly Exports"]
        asm = wb["Assumptions"]
        loc = locate_exports(me)                       # raises if an anchor is missing
        # Read the Assumptions tab once: status + the fingerprinted assumption fields
        # + an ownership signature, keyed by project name.
        status_of, assum_of = {}, {}
        for r in range(3, asm.max_row + 1):
            nm = _clean(asm.cell(r, 1).value)
            if not nm or nm == "Total":
                continue
            status_of[nm] = _clean(asm.cell(r, 3).value) or ""
            a = {label: _av(asm.cell(r, c).value) for c, label in ASSUMP_COLS}
            a["ownership"] = [_av(asm.cell(r, c).value) for c in OWNERSHIP_COLS]
            assum_of[nm] = a
        # Walk the Reported section exactly as the extractor does: mains (M) only,
        # country assigned on the country aggregate row.
        projects, pending = {}, []
        for r in range(loc["rep"][0], loc["rep"][1] + 1):
            nm = _clean(me.cell(r, 1).value)
            if not nm:
                continue
            mk = _clean(me.cell(r, 3).value)
            if mk in ("M", "T"):
                pending.append((nm, mk))
            elif nm in COUNTRIES:
                for pn, pm in pending:
                    if pm == "M":
                        projects[pn] = {"country": nm, "region": REGION_OF.get(nm, nm),
                                        "status": status_of.get(pn, ""),
                                        "assumptions": assum_of.get(pn, {})}
                pending = []
            elif nm == "Grand total":
                pending = []
        return {
            "projects": projects,
            "countries": sorted({p["country"] for p in projects.values()}),
            "regions": sorted({p["region"] for p in projects.values()}),
            "anchors": {k: (list(v) if isinstance(v, tuple) else v) for k, v in loc.items()},
        }
    finally:
        wb.close()


def main():
    write = "--write" in sys.argv
    try:
        snap = snapshot()
    except Exception as e:
        print(f"[PREFLIGHT] LAYOUT ERROR: {e}")
        print("STOP - the master's layout could not be read by the anchors. Call Claude.")
        return 2

    if write or not MANIFEST.exists():
        first = not MANIFEST.exists()
        MANIFEST.parent.mkdir(exist_ok=True)
        MANIFEST.write_text(json.dumps(snap, indent=1))
        print(f"[PREFLIGHT] manifest {'initialised' if first else 'rewritten'}: "
              f"{len(snap['projects'])} projects, {len(snap['countries'])} countries, "
              f"{len(snap['regions'])} regions")
        return 0

    old = json.loads(MANIFEST.read_text())
    on, nn = set(old["projects"]), set(snap["projects"])
    added = sorted(nn - on)
    removed = sorted(on - nn)
    statchg = [(p, old["projects"][p]["status"], snap["projects"][p]["status"])
               for p in sorted(on & nn)
               if old["projects"][p]["status"] != snap["projects"][p]["status"]]
    new_regions = sorted(set(snap["regions"]) - set(old["regions"]))
    new_countries = sorted(set(snap["countries"]) - set(old["countries"]))

    # --- Assumptions-tab changes on surviving projects (CoS / utilisation / start /
    #     decline / unrisked / ownership). Reported, never blocking — these flow into
    #     the build automatically; this is the visible confirmation of what changed. ---
    assumpchg = []
    for p in sorted(on & nn):
        oa = old["projects"][p].get("assumptions") or {}
        na = snap["projects"][p].get("assumptions") or {}
        if not oa or not na:
            continue                       # manifest predates this check -> nothing to compare
        changes = []
        for _c, label in ASSUMP_COLS:
            if oa.get(label) != na.get(label):
                changes.append(f"{label} {_fmt(oa.get(label))}->{_fmt(na.get(label))}")
        if oa.get("ownership") != na.get("ownership"):
            changes.append("ownership/stakes")
        if changes:
            assumpchg.append((p, changes))

    print(f"[PREFLIGHT] master: {len(snap['projects'])} projects | "
          f"added {len(added)}, removed {len(removed)}, status-changes {len(statchg)}, "
          f"assumption-changes {len(assumpchg)}")
    for p in added:
        print(f"   + NEW PROJECT: {p}  ({snap['projects'][p]['country']}, {snap['projects'][p]['status']})")
    for p in removed:
        print(f"   - REMOVED/RENAMED: {p}")
    for p, o, n in statchg:
        print(f"   ~ STATUS CHANGE: {p}: {o or '(blank)'} -> {n or '(blank)'}")
    for p, ch in assumpchg:
        print(f"   ~ ASSUMPTIONS: {p}: {', '.join(ch)}")
    if new_regions:
        print(f"   ! NEW REGION(S): {new_regions}")
    if new_countries:
        print(f"   ! NEW COUNTRY(IES): {new_countries}")
    if old.get("projects") and not any("assumptions" in old["projects"][p] for p in on & nn):
        print("   (note: previous manifest had no assumptions fingerprint — assumption "
              "changes will be reported from the NEXT update onward.)")

    if removed or new_regions or new_countries:
        print("STOP - structural change needs review (removed/renamed project or new "
              "region/country). Call Claude before updating.")
        return 2
    if added or assumpchg or statchg:
        print("PROCEED - changes detected above (new projects / status / assumptions); all "
              "flow into the build automatically. Check new projects' colour/region mapping.")
        return 0
    print("OK - no structural or assumption changes detected vs the last update. "
          "(If you expected changes, check you pasted the latest master.)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
