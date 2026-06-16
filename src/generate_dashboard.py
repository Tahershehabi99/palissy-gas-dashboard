"""
Palissy Multi-Dataset Dashboard Generator

Reads multiple dataset tabs from INPUT/gas_model_input.xlsx and produces a
single self-contained HTML page with a Gas/Power toggle.

Datasets are declared in DATASETS below. Each dataset is one config block —
adding a new page (e.g. LNG, Storage) is a config addition, not a code change.

Gas page preserves all prior behavior; Power page is new.
"""

import openpyxl
import json
import os
import base64
from datetime import datetime
from calendar import monthrange  # noqa: F401  (kept for parity)

# ============================================================
# ADMIN CONFIGURATION
# ============================================================
DISPLAY_START_YEAR = 2020
DISPLAY_END_YEAR = 2030
# ============================================================

# Paths
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(SCRIPT_DIR)
INPUT_FILE = os.path.join(PROJECT_DIR, "INPUT", "gas_model_input.xlsx")
LNG_INPUT_FILE = os.path.join(PROJECT_DIR, "INPUT", "lng_model_input.xlsx")
OUTPUT_DIR = os.path.join(PROJECT_DIR, "output")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "index.html")
LOGO_FILE = os.path.join(PROJECT_DIR, "Context", "Palissy Logo.png")
FONT_FILE = os.path.join(PROJECT_DIR, "Context", "Gotham-Book.otf")

# Palissy brand colors
COLORS = {
    "dark_blue": "#272962",
    "light_green": "#539648",
    "dark_green": "#0C5B19",
    "red": "#C00000",
    "grey": "#9395A2",
    "light_blue": "#0B5AAB",
    "blue": "#258EEB",
    "card_bg": "#f8f8fb",
    "border": "rgba(39, 41, 98, 0.15)",
    "grid": "#E0E0E8",
}

# ============================================================
# DATASETS
# Each entry is a self-contained spec for one page.
# ============================================================
GAS_UNIT_CONFIG = {
    "bcf":    {"volLabel": "bcf",  "rateLabel": "bcf/d",  "volFactor": 1.0,         "rateFactor": 1.0,           "isRate": False},
    "bcf/d":  {"volLabel": "bcf",  "rateLabel": "bcf/d",  "volFactor": 1.0,         "rateFactor": 1.0,           "isRate": True},
    "bcm":    {"volLabel": "bcm",  "rateLabel": "mmcm/d", "volFactor": 1.0/35.3,    "rateFactor": 1000.0/35.3,   "isRate": False},
    "mmcm/d": {"volLabel": "bcm",  "rateLabel": "mmcm/d", "volFactor": 1.0/35.3,    "rateFactor": 1000.0/35.3,   "isRate": True},
    "TWh":    {"volLabel": "TWh",  "rateLabel": "GWh/d",  "volFactor": 1.0/3.41,    "rateFactor": 1000.0/3.41,   "isRate": False},
    "GWh/d":  {"volLabel": "TWh",  "rateLabel": "GWh/d",  "volFactor": 1.0/3.41,    "rateFactor": 1000.0/3.41,   "isRate": True},
    "mmt":    {"volLabel": "mmt",  "rateLabel": "kt/d",   "volFactor": 1.0/48.0,    "rateFactor": 1000.0/48.0,   "isRate": False},
    "kt/d":   {"volLabel": "mmt",  "rateLabel": "kt/d",   "volFactor": 1.0/48.0,    "rateFactor": 1000.0/48.0,   "isRate": True},
}

# Gas-to-power efficiency assumption for BCFE conversions.
# BCFE = (TWh) * 3.41 / POWER_EFFICIENCY
# Industry-standard "thermal efficiency" placeholder; change here if the
# assumption updates.
POWER_EFFICIENCY = 0.39
BCFE_FACTOR = 3.41 / POWER_EFFICIENCY  # ~= 8.7436

POWER_UNIT_CONFIG = {
    "TWh":    {"volLabel": "TWh",  "rateLabel": "TWh",    "volFactor": 1.0,         "rateFactor": 1.0,           "isRate": False},
    "BCFE":   {"volLabel": "BCFE", "rateLabel": "BCFE/d", "volFactor": BCFE_FACTOR, "rateFactor": BCFE_FACTOR,   "isRate": False},
    "BCFE/d": {"volLabel": "BCFE", "rateLabel": "BCFE/d", "volFactor": BCFE_FACTOR, "rateFactor": BCFE_FACTOR,   "isRate": True},
}

# Palissy-aligned source palette for power generation.
# Tuned for visual differentiation between adjacent stack layers and
# industry-standard fuel-source associations (gas=red, coal=charcoal,
# nuclear=gold, solar=orange, wind=sky, hydro=deep blue, etc.).
# Tweakable - user said "play around with the colors".
POWER_SOURCE_COLORS = {
    # Bottom five sources match the seasonality chart's line colors so that
    # users can pattern-match between the two charts. Top four picked for
    # contrast and Palissy-palette consistency.
    # Stack order bottom -> top: Nuclear / Wind / Gas / Solar / Hydro / Coal
    # / Bioenergy / Other Fossil / Other Renewables.
    "Nuclear":                              "#272962",  # Palissy dark blue (= seasonality avg)
    "Wind":                                 "#539648",  # Palissy light green (= seasonality forecast)
    "Gas":                                  "#C00000",  # Palissy red (= seasonality current)
    "Solar":                                "#E5B83A",  # warm gold-yellow
    "Hydro":                                "#0C5B19",  # Palissy dark green (= seasonality prev)
    "Coal":                                 "#2C2C2E",  # charcoal
    "Bioenergy":                            "#708B5A",  # muted olive
    "Other Fossil":                         "#92591C",  # earth brown
    "Other Renewables":                     "#B8D67A",  # pale mint
    "Total Europe Electricity Generation":  "#272962",  # Palissy dark blue
}

# Gas balance series colors for the gas build-up chart (stack layers / lines).
# Keyed by the exact row labels in the Monthly Data tab. Tuned for contrast
# between adjacent stack layers; supply = greens/blues, consumption = warm tones.
GAS_SERIES_COLORS = {
    "+ Domestic Production":       "#0C5B19",  # Palissy dark green
    "Russia":                      "#C00000",  # Palissy red
    "Norway":                      "#272962",  # Palissy dark blue
    "Algeria":                     "#258EEB",  # bright blue
    "Azerbaijan":                  "#2A9D8F",  # teal
    "Libya":                       "#E5B83A",  # gold
    "Reverse flow into Ukraine":   "#9395A2",  # Palissy grey
    "+ LNG Send Out":              "#539648",  # Palissy light green
    "Power":                       "#8E44AD",  # purple
    "Industrial":                  "#2C2C2E",  # charcoal
    "Residential and Commercial":  "#708B5A",  # olive
    "Adjustment":                  "#B8860B",  # dark goldenrod
    "- Unreported Consumption":    "#92591C",  # earth brown
    "- Exports":                   "#0B5AAB",  # Palissy light blue
    "Storage percentage":          "#272962",  # Palissy dark blue (line)
}

# ============================================================
# LNG (Global LNG tab) — base unit is MMT.
# Conversions reuse the verified gas factors, re-based to MMT:
#   1 mmt = 48 bcf ; 1 bcm = 35.3 bcf ; 1 TWh = 3.41 bcf.
# So from a base of mmt: bcf = x48, bcm = x(48/35.3), TWh = x(48/3.41),
# kt/d = mmt/day x1000 (1 mmt = 1000 kt).
# ============================================================
LNG_UNIT_CONFIG = {
    "mmt":    {"volLabel": "mmt", "rateLabel": "kt/d",   "volFactor": 1.0,          "rateFactor": 1000.0,             "isRate": False},
    "kt/d":   {"volLabel": "mmt", "rateLabel": "kt/d",   "volFactor": 1.0,          "rateFactor": 1000.0,             "isRate": True},
    "bcf":    {"volLabel": "bcf", "rateLabel": "bcf/d",  "volFactor": 48.0,         "rateFactor": 48.0,               "isRate": False},
    "bcf/d":  {"volLabel": "bcf", "rateLabel": "bcf/d",  "volFactor": 48.0,         "rateFactor": 48.0,               "isRate": True},
    "bcm":    {"volLabel": "bcm", "rateLabel": "mmcm/d", "volFactor": 48.0/35.3,    "rateFactor": 48.0/35.3*1000.0,   "isRate": False},
    "mmcm/d": {"volLabel": "bcm", "rateLabel": "mmcm/d", "volFactor": 48.0/35.3,    "rateFactor": 48.0/35.3*1000.0,   "isRate": True},
    "TWh":    {"volLabel": "TWh", "rateLabel": "GWh/d",  "volFactor": 48.0/3.41,    "rateFactor": 48.0/3.41*1000.0,   "isRate": False},
    "GWh/d":  {"volLabel": "TWh", "rateLabel": "GWh/d",  "volFactor": 48.0/3.41,    "rateFactor": 48.0/3.41*1000.0,   "isRate": True},
}
LNG_UNITS = ["mmt", "kt/d", "bcf", "bcf/d", "bcm", "mmcm/d", "TWh", "GWh/d"]

# Import-country colors (stack layers / range lines). Placeholder palette —
# tweak with user. Keyed by exact row labels in the Monthly Imports tab.
LNG_IMPORT_COLORS = {
    "EU+UK":        "#272962",  # Palissy dark blue
    "China":        "#C00000",  # Palissy red
    "Japan":        "#0B5AAB",  # Palissy light blue
    "South Korea":  "#539648",  # Palissy light green
    "Taiwan":       "#E5B83A",  # gold
    "India":        "#8E44AD",  # purple
    "Other Asia":   "#258EEB",  # bright blue
    "LatAm":        "#0C5B19",  # dark green
    "Middle East":  "#92591C",  # earth brown
    "Egypt":        "#2A9D8F",  # teal
    "Turkey":       "#708B5A",  # olive
    "RoW":          "#9395A2",  # Palissy grey
    "Unaccounted demand": "#B8B9C2",
    "Total":        "#272962",
}

# Export-region colors. Keyed by the regional subtotal / standalone labels.
LNG_EXPORT_COLORS = {
    "Asia":                "#272962",  # Palissy dark blue
    "Australia":           "#258EEB",  # bright blue
    "LatAm":               "#0C5B19",  # dark green
    "MENA and Europe":     "#C00000",  # Palissy red
    "North America":       "#E5B83A",  # gold
    "Russia":              "#9395A2",  # Palissy grey
    "Sub-Saharan Africa":  "#539648",  # Palissy light green
    "Grand total":         "#272962",
}

# --- LNG sub-datasets (the Imports + Exports halves of the Global LNG tab) ---
# Each is read like any other dataset, but they share one composite tab.
LNG_IMPORTS_CFG = {
    "key": "lng_imports",
    "tab_label": "Imports",
    "title": "LNG Imports",
    "input_file": "lng",                 # use INPUT/lng_model_input.xlsx
    "sheet": "Monthly Imports",
    "date_row": 4,
    "days_row": None,                    # no days row -> derive from month dates
    "data_start_row": 7,                 # EU+UK (row 5 'Regional', row 6 'MMT' are headers)
    "data_end_row": 20,                  # ... Total
    "data_start_col": 3,                 # data begins at column C
    "base_unit": "mmt",
    "units": LNG_UNITS,
    "default_unit": "mmt",
    "unit_config": LNG_UNIT_CONFIG,
    "stock_rows": [],
    "pct_rows": [],
    "use_hierarchy": False,              # flat list of countries + Total
    "skip_label_rows": ["Regional", "MMT"],
    "total_row": "Total",
    "source_colors": LNG_IMPORT_COLORS,
    # Charts: flat country list. Range = multiselect (countries + Total, mutex).
    # Stacked series = the 12 countries (Unaccounted/Total excluded).
    "range_kind": "flat",
    "chart_series": ["EU+UK", "China", "Japan", "South Korea", "Taiwan", "India",
                     "Other Asia", "LatAm", "Middle East", "Egypt", "Turkey", "RoW"],
    # 'unaccounted_row' marks the actual-vs-forecast boundary: the last month
    # where this row is EMPTY is the latest month with real data.
    "unaccounted_row": "Unaccounted demand",
}
LNG_EXPORTS_CFG = {
    "key": "lng_exports",
    "tab_label": "Exports",
    "title": "LNG Exports",
    "input_file": "lng",
    "sheet": "Monthly Exports",
    "date_row": 4,
    "days_row": None,
    "data_start_row": 8,                 # 'Production (mmt)' header skipped
    "data_end_row": 44,                  # ... Grand total
    "data_start_col": 6,                 # data begins at column F (B-E are 'Working' cols)
    "base_unit": "mmt",
    "units": LNG_UNITS,
    "default_unit": "mmt",
    "unit_config": LNG_UNIT_CONFIG,
    "stock_rows": [],
    "pct_rows": [],
    "use_hierarchy": False,
    "skip_label_rows": ["Production (mmt)"],
    "total_row": "Grand total",
    "source_colors": LNG_EXPORT_COLORS,
    # Exports table is hierarchical: regional subtotal rows with member countries
    # above them in the sheet. We declare the structure explicitly (the sheet has
    # no +/- prefixes). Parent rows reuse the sheet's subtotal values.
    "explicit_hierarchy": [
        {"type": "group", "label": "Asia",
         "members": ["Brunei", "Indonesia", "Malaysia", "Papua New Guinea", "Timor Leste"]},
        {"type": "standalone", "label": "Australia"},
        {"type": "group", "label": "LatAm",
         "members": ["Argentina", "Mexico", "Peru", "Suriname", "Trinidad", "Venezuela"]},
        {"type": "group", "label": "MENA and Europe",
         "members": ["Algeria", "Egypt", "Israel", "Mauritania", "Norway", "Oman", "Qatar", "UAE"]},
        {"type": "group", "label": "North America",
         "members": ["Canada", "United States"]},
        {"type": "standalone", "label": "Russia"},
        {"type": "group", "label": "Sub-Saharan Africa",
         "members": ["Angola", "Cameroon", "Congo (Rep.)", "Equatorial Guinea",
                     "Mozambique", "Nigeria", "Tanzania"]},
        {"type": "standalone", "label": "Grand total"},
    ],
    "range_kind": "hierarchical",
    # Stacked series = top-level regions (subtotals + standalones), not Grand total.
    "chart_series": ["Asia", "Australia", "LatAm", "MENA and Europe",
                     "North America", "Russia", "Sub-Saharan Africa"],
}

DATASETS = [
    {
        "key": "gas",
        "tab_label": "Gas Balance",
        "title": "European Gas Balance",
        "sheet": "Monthly Data",
        "date_row": 1,
        "days_row": 2,
        "data_start_row": 5,
        "base_unit": "bcf",
        "units": ["bcf", "bcf/d", "bcm", "mmcm/d", "TWh", "GWh/d", "mmt", "kt/d"],
        "default_unit": "bcf",
        "unit_config": GAS_UNIT_CONFIG,
        "stock_rows": ["Opening Storage", "Closing Storage", "Storage percentage"],
        "pct_rows": ["Storage percentage"],
        "use_hierarchy": True,
        "skip_label_rows": [],
        "charts_enabled": False,
        # Gas uses the 2-tables-on-top / 2-charts-below layout.
        "chart_area": True,
        # LEFT chart = range/seasonality with a hierarchical source selector.
        # Each entry is one selectable item in the dropdown:
        #   kind 'pct'   -> percentage series (Storage %), mutually exclusive,
        #                   default selection, no unit conversion.
        #   kind 'leaf'  -> a single base row, combinable with other totals.
        #   kind 'group' -> expands to member rows; the group total is the sum
        #                   of all members; selecting an individual member drills
        #                   in (exclusive to that group).
        # abs=True means the stored values are negative (outflows) and must be
        # shown as positive magnitudes on the chart.
        "chart_groups": [
            {"key": "storage", "label": "Storage percentage", "kind": "pct",
             "rows": ["Storage percentage"], "abs": False},
            {"key": "domprod", "label": "Domestic Production", "kind": "leaf",
             "rows": ["+ Domestic Production"], "abs": False},
            {"key": "imports", "label": "Imports", "kind": "group",
             "rows": ["Russia", "Norway", "Algeria", "Azerbaijan", "Libya",
                      "Reverse flow into Ukraine"], "abs": False},
            {"key": "lng", "label": "LNG Send Out", "kind": "leaf",
             "rows": ["+ LNG Send Out"], "abs": False},
            {"key": "consumption", "label": "Consumption", "kind": "group",
             "rows": ["Power", "Industrial", "Residential and Commercial",
                      "Adjustment", "- Unreported Consumption"], "abs": True},
            {"key": "exports", "label": "Exports", "kind": "leaf",
             "rows": ["- Exports"], "abs": True},
        ],
        "total_row": None,
        "source_colors": GAS_SERIES_COLORS,
    },
    {
        "key": "power",
        "tab_label": "Power",
        "title": "European Power Generation",
        "sheet": "Power Data",
        "date_row": 6,
        "days_row": 4,
        "data_start_row": 8,
        "base_unit": "TWh",
        "units": ["TWh", "BCFE", "BCFE/d"],
        "default_unit": "TWh",
        "unit_config": POWER_UNIT_CONFIG,
        "stock_rows": [],
        "pct_rows": [],
        "use_hierarchy": False,
        # Skip the "Generation by Source - EU + UK" section header which has no numeric data.
        "skip_label_rows": ["Generation by Source - EU + UK"],
        "charts_enabled": True,
        # Row label that represents the total - used by charts to handle Total
        # as a special selection (mutex with individual sources).
        "total_row": "Total Europe Electricity Generation",
        "source_colors": POWER_SOURCE_COLORS,
        # Sort non-total rows descending by total generation in this calendar year.
        # Used to order the table rows and the buildup-chart stack so the largest
        # contributor sits at the top of the table / bottom of the stack.
        "sort_by_year": 2025,
    },
    {
        # Global LNG: a COMPOSITE tab. Unlike gas/power (one table + 2 charts) it
        # stacks four full-width tables (Imports, Imports change, Exports, Exports
        # change) then two rows of side-by-side charts (Import/Export range, then
        # Import/Export stacked). Built from two sub-datasets sharing the top
        # Period/Unit/From/To bar. See build_composite_blob + the lng* JS block.
        "key": "global_lng",
        "tab_label": "Global LNG",
        "title": "Global LNG",
        "composite": True,
        "subs": [LNG_IMPORTS_CFG, LNG_EXPORTS_CFG],
        # Top-bar unit machinery reads these off the composite blob directly.
        "base_unit": "mmt",
        "units": LNG_UNITS,
        "default_unit": "mmt",
        "unit_config": LNG_UNIT_CONFIG,
    },
    {
        # LNG Projects: a project database (NOT the table/charts engine). Its own
        # sub-tabs (Projects assumptions view + Supply Outlook time-series), a
        # progressive filter bar, and a Region->Country->Project tree. Reads the
        # 'LNG Projects' sheet built by extract_lng_input.py (mains only).
        "key": "lng_projects",
        "tab_label": "LNG Projects",
        "title": "Global LNG Projects",
        "projects_tab": True,
        "input_file": "lng",
    },
]

# ============================================================
# EMBED TABS
# External dashboards surfaced as tabs via iframe. These are the live,
# daily-updated pages from the separate "Storage and LNG" project
# (eu-gas-dashboard GitHub Pages). We embed them rather than re-implementing
# so their independent daily pipeline keeps updating them — this dashboard
# just surfaces them as extra tabs. "?embedded=1" hides their own logo.
# ============================================================
EMBED_TABS = [
    {"key": "storage", "label": "Storage",
     "url": "https://tahershehabi99.github.io/eu-gas-dashboard/storage.html?embedded=1"},
    {"key": "lng", "label": "LNG Sendout",
     "url": "https://tahershehabi99.github.io/eu-gas-dashboard/lng.html?embedded=1"},
]


# ============================================================
# DATA LOADING
# ============================================================
def read_dataset(wb, config):
    """Read one dataset tab according to its config. Returns (dates, days, rows)."""
    sheet_name = config["sheet"]
    print(f"\n[{config['key']}] Reading sheet '{sheet_name}'...")
    ws = wb[sheet_name]

    date_row = config["date_row"]
    days_row = config.get("days_row")
    data_start_row = config["data_start_row"]
    data_start_col = config.get("data_start_col", 2)   # first data column
    data_end_row = config.get("data_end_row")          # optional hard stop

    # Detect last column with a date in the date_row
    last_col = 1
    for c in range(data_start_col, ws.max_column + 1):
        if ws.cell(row=date_row, column=c).value is not None:
            last_col = c

    # Dates
    dates = []
    for c in range(data_start_col, last_col + 1):
        v = ws.cell(row=date_row, column=c).value
        if isinstance(v, datetime):
            dates.append(v)
        elif isinstance(v, str):
            dates.append(datetime.strptime(v, "%Y-%m-%d"))
        else:
            dates.append(None)

    # Days: read from days_row if configured, otherwise derive from the month
    # dates (calendar days in that month). LNG inputs have no days row.
    days_per_month = []
    for ci, c in enumerate(range(data_start_col, last_col + 1)):
        if days_row is not None:
            v = ws.cell(row=days_row, column=c).value
            days_per_month.append(int(v) if v is not None else 30)
        else:
            d = dates[ci]
            days_per_month.append(monthrange(d.year, d.month)[1] if d is not None else 30)

    # Data rows
    skip = set(config.get("skip_label_rows", []))
    rows = []
    r = data_start_row
    blank_streak = 0
    while r <= ws.max_row and blank_streak < 3:
        if data_end_row is not None and r > data_end_row:
            break
        label = ws.cell(row=r, column=1).value
        if label is None:
            blank_streak += 1
            r += 1
            continue
        blank_streak = 0
        label_str = str(label).strip()
        if label_str in skip:
            r += 1
            continue
        values = []
        for c in range(data_start_col, last_col + 1):
            v = ws.cell(row=r, column=c).value
            values.append(float(v) if v is not None else 0.0)
        # If no numeric values at all (e.g. orphan label), skip
        if all(v == 0.0 for v in values) and label_str not in config.get("stock_rows", []):
            # Could be a real all-zero forecast row; only skip if also has no
            # children-style behavior. For safety, keep all-zero rows so we
            # don't accidentally drop a real (empty) source.
            pass
        rows.append({"label": label_str, "values": values})
        r += 1

    print(f"  Loaded {len(rows)} rows x {len(dates)} months ({dates[0].strftime('%b %Y')} - {dates[-1].strftime('%b %Y')})")
    return dates, days_per_month, rows


# ============================================================
# HIERARCHY DETECTION
# ============================================================
def detect_hierarchy(rows, stock_rows, use_hierarchy):
    """Detect parent/child structure.

    With use_hierarchy=True (gas pattern):
      Rows prefixed + or - = parent totals
      Un-prefixed rows between parents = children of the NEXT parent
      stock_rows = always standalone (no children)

    With use_hierarchy=False (flat pattern, e.g. power):
      All rows are standalone, in source order.
    """
    if not use_hierarchy:
        return [
            {
                "label": row["label"],
                "row_index": i,
                "children": [],
                "type": "standalone",
            }
            for i, row in enumerate(rows)
        ]

    standalone = set(stock_rows)
    classified = []
    for i, row in enumerate(rows):
        label = row["label"]
        classified.append({
            "index": i,
            "label": label,
            "is_parent": label.startswith("+") or label.startswith("-"),
            "is_standalone": label in standalone,
        })

    hierarchy = []
    pending = []
    for item in classified:
        if item["is_standalone"]:
            for child in pending:
                hierarchy.append({"label": child["label"], "row_index": child["index"], "children": [], "type": "standalone"})
            pending = []
            hierarchy.append({"label": item["label"], "row_index": item["index"], "children": [], "type": "standalone"})
        elif item["is_parent"]:
            if pending:
                children = [{"label": c["label"], "row_index": c["index"]} for c in pending]
                hierarchy.append({"label": item["label"], "row_index": item["index"], "children": children, "type": "parent"})
                pending = []
            else:
                hierarchy.append({"label": item["label"], "row_index": item["index"], "children": [], "type": "standalone"})
        else:
            pending.append(item)

    for child in pending:
        hierarchy.append({"label": child["label"], "row_index": child["index"], "children": [], "type": "standalone"})
    return hierarchy


# ============================================================
# AGGREGATION (period-level)
# ============================================================
def aggregate_monthly_to_periods(dates, days_per_month):
    """Group monthly indices into period columns. Generic across datasets."""
    results = {}

    # Monthly
    monthly_cols = [{
        "label": d.strftime("%b %Y"),
        "short": d.strftime("%b-%y"),
        "year": d.year,
        "month": d.month,
        "gas_year": d.year if d.month >= 10 else d.year - 1,
        "indices": [i],
        "days": days_per_month[i],
    } for i, d in enumerate(dates)]
    results["Monthly"] = monthly_cols

    # Quarterly
    quarters = {}
    for i, d in enumerate(dates):
        q = (d.month - 1) // 3 + 1
        key = (d.year, q)
        quarters.setdefault(key, {"indices": [], "days": 0})
        quarters[key]["indices"].append(i)
        quarters[key]["days"] += days_per_month[i]
    results["Quarterly"] = [{
        "label": f"Q{q} {y}", "short": f"Q{q}-{str(y)[2:]}",
        "year": y, "quarter": q,
        "gas_year": y if q == 4 else y - 1,
        "indices": info["indices"], "days": info["days"],
    } for (y, q), info in sorted(quarters.items()) if len(info["indices"]) == 3]

    # Annual Calendar Year
    years = {}
    for i, d in enumerate(dates):
        years.setdefault(d.year, {"indices": [], "days": 0})
        years[d.year]["indices"].append(i)
        years[d.year]["days"] += days_per_month[i]
    results["Annual CY"] = [{
        "label": str(y), "short": str(y), "year": y,
        "indices": info["indices"], "days": info["days"],
    } for y, info in sorted(years.items()) if len(info["indices"]) == 12]

    # Annual Gas Year
    gas_years = {}
    for i, d in enumerate(dates):
        gy = d.year if d.month >= 10 else d.year - 1
        gas_years.setdefault(gy, {"indices": [], "days": 0})
        gas_years[gy]["indices"].append(i)
        gas_years[gy]["days"] += days_per_month[i]
    results["Gas Year"] = [{
        "label": f"GY {str(gy)[2:]}/{str(gy+1)[2:]}",
        "short": f"{str(gy)[2:]}/{str(gy+1)[2:]}",
        "year": gy, "indices": info["indices"], "days": info["days"],
    } for gy, info in sorted(gas_years.items()) if len(info["indices"]) == 12]

    # Winters (Oct-Mar)
    winters = {}
    for i, d in enumerate(dates):
        if d.month >= 10:
            gy = d.year
        elif d.month <= 3:
            gy = d.year - 1
        else:
            continue
        winters.setdefault(gy, {"indices": [], "days": 0})
        winters[gy]["indices"].append(i)
        winters[gy]["days"] += days_per_month[i]
    results["Winter"] = [{
        "label": f"Win {str(gy)[2:]}/{str(gy+1)[2:]}",
        "short": f"Win {str(gy)[2:]}/{str(gy+1)[2:]}",
        "year": gy, "indices": info["indices"], "days": info["days"],
    } for gy, info in sorted(winters.items()) if len(info["indices"]) == 6]

    # Summers (Apr-Sep)
    summers = {}
    for i, d in enumerate(dates):
        if 4 <= d.month <= 9:
            summers.setdefault(d.year, {"indices": [], "days": 0})
            summers[d.year]["indices"].append(i)
            summers[d.year]["days"] += days_per_month[i]
    results["Summer"] = [{
        "label": f"Sum {y}", "short": f"Sum {y}",
        "year": y, "indices": info["indices"], "days": info["days"],
    } for y, info in sorted(summers.items()) if len(info["indices"]) == 6]

    return results


def compute_period_values(rows, period_cols, stock_rows, pct_rows):
    """For each period column, compute aggregated value per row.
    Stock rows: opening = first index value; closing = last index value;
    Storage percentage row: last index value.
    All other rows: sum of the period.
    """
    out = []
    for row in rows:
        label = row["label"]
        vals = row["values"]
        period_values = []
        for col in period_cols:
            idx = col["indices"]
            if not idx:
                period_values.append(0)
                continue
            if label == "Opening Storage":
                period_values.append(vals[idx[0]])
            elif label == "Closing Storage":
                period_values.append(vals[idx[-1]])
            elif label in pct_rows:
                period_values.append(vals[idx[-1]])
            elif label in stock_rows:
                # Generic stock fallback (last value)
                period_values.append(vals[idx[-1]])
            else:
                period_values.append(sum(vals[i] for i in idx))
        out.append({"label": label, "base_values": period_values})
    return out


# ============================================================
# DATASET BUILDER
# ============================================================
def build_explicit_hierarchy(rows, spec):
    """Build a hierarchy from an explicit config spec (used by LNG exports,
    whose sheet has regional subtotal rows with no +/- prefixes).

    spec entries:
      {"type":"group","label":"Asia","members":[...]} -> parent row + children
      {"type":"standalone","label":"Australia"}        -> standalone row
    Labels are matched to the loaded rows by label; the parent row reuses the
    sheet's subtotal values.
    """
    idx_by_label = {}
    for i, r in enumerate(rows):
        idx_by_label.setdefault(r["label"], i)
    hierarchy = []
    for entry in spec:
        label = entry["label"]
        if label not in idx_by_label:
            print(f"  WARNING: explicit hierarchy label not found in data: {label!r}")
            continue
        if entry["type"] == "group":
            children = []
            for m in entry.get("members", []):
                if m in idx_by_label:
                    children.append({"label": m, "row_index": idx_by_label[m]})
                else:
                    print(f"  WARNING: hierarchy member not found: {m!r}")
            hierarchy.append({"label": label, "row_index": idx_by_label[label],
                              "children": children, "type": "parent"})
        else:
            hierarchy.append({"label": label, "row_index": idx_by_label[label],
                              "children": [], "type": "standalone"})
    return hierarchy


def build_dataset_blob(config):
    """Load + aggregate one dataset and produce its JSON-ready blob."""
    src_file = LNG_INPUT_FILE if config.get("input_file") == "lng" else INPUT_FILE
    wb = openpyxl.load_workbook(src_file, data_only=True)
    try:
        dates, days, rows = read_dataset(wb, config)
    finally:
        wb.close()

    # Optional: sort non-total rows descending by total in a target calendar year.
    # Keeps the largest contributors at the top of the table and at the bottom
    # of the stacked buildup chart.
    sort_year = config.get("sort_by_year")
    if sort_year:
        total_label = config.get("total_row")
        target_indices = [i for i, d in enumerate(dates) if d.year == sort_year]
        def row_year_sum(r):
            return sum(r["values"][i] for i in target_indices) if target_indices else 0
        sortable = [r for r in rows if r["label"] != total_label]
        total_rows = [r for r in rows if r["label"] == total_label]
        sortable.sort(key=row_year_sum, reverse=True)
        rows = sortable + total_rows
        print(f"  Sorted rows by CY {sort_year} generation (descending):")
        for r in sortable:
            print(f"    {r['label']}: {row_year_sum(r):,.1f}")

    if config.get("explicit_hierarchy"):
        hierarchy = build_explicit_hierarchy(rows, config["explicit_hierarchy"])
    else:
        hierarchy = detect_hierarchy(rows, config["stock_rows"], config["use_hierarchy"])
    print(f"  Hierarchy: {len(hierarchy)} items, {sum(1 for h in hierarchy if h['children']) } with children")

    period_results = aggregate_monthly_to_periods(dates, days)
    print(f"  Periods: " + ", ".join(f"{k}={len(v)}" for k, v in period_results.items()))

    views = {}
    for view_name, period_cols in period_results.items():
        aggregated = compute_period_values(rows, period_cols, config["stock_rows"], config["pct_rows"])
        days_array = [c["days"] for c in period_cols]
        col_meta = []
        for c in period_cols:
            meta = {"label": c["label"], "short": c["short"], "year": c.get("year", 0), "days": c["days"]}
            if "month" in c: meta["month"] = c["month"]
            if "quarter" in c: meta["quarter"] = c["quarter"]
            if "gas_year" in c: meta["gas_year"] = c["gas_year"]
            col_meta.append(meta)
        views[view_name] = {
            "columns": [c["label"] for c in period_cols],
            "short_columns": [c["short"] for c in period_cols],
            "col_meta": col_meta,
            "days": days_array,
            "rows": [{"label": a["label"], "base": a["base_values"]} for a in aggregated],
        }

    ui_hierarchy = []
    for item in hierarchy:
        entry = {
            "label": item["label"],
            "index": item["row_index"],
            "type": item["type"],
            "is_stock": item["label"] in config["stock_rows"],
            "is_pct": item["label"] in config["pct_rows"],
            "children": [],
        }
        for child in item["children"]:
            entry["children"].append({
                "label": child["label"],
                "index": child["row_index"],
                "is_stock": False,
                "is_pct": False,
            })
        ui_hierarchy.append(entry)

    return {
        "key": config["key"],
        "tab_label": config["tab_label"],
        "title": config["title"],
        "base_unit": config["base_unit"],
        "units": config["units"],
        "default_unit": config["default_unit"],
        "unit_config": config["unit_config"],
        "use_hierarchy": config["use_hierarchy"],
        "charts_enabled": config.get("charts_enabled", False),
        # chart_area = does this dataset use the charts layout (tables on top,
        # chart boxes below). Defaults to charts_enabled so power gets it for free.
        "chart_area": config.get("chart_area", config.get("charts_enabled", False)),
        # Per-slot chart availability. Power has both; gas (chart_groups) has the
        # left range chart only (right stays a placeholder until specified).
        "has_range": config.get("charts_enabled", False) or bool(config.get("chart_groups")),
        "has_buildup": config.get("charts_enabled", False) or bool(config.get("chart_groups")),
        "chart_groups": config.get("chart_groups"),
        "total_row": config.get("total_row"),
        "source_colors": config.get("source_colors", {}),
        "views": views,
        "hierarchy": ui_hierarchy,
        "selectable_start": DISPLAY_START_YEAR,
        "selectable_end": DISPLAY_END_YEAR,
        # LNG composite metadata (ignored by gas/power).
        "range_kind": config.get("range_kind"),
        "chart_series": config.get("chart_series"),
    }


def build_projects_blob(config):
    """Build the LNG Projects blob from the clean 'LNG Projects' sheet (mains only,
    already joined to Assumptions by extract_lng_input.py). Parses owner/partner
    companies for the company filter and pairs partners with their stakes."""
    wb = openpyxl.load_workbook(LNG_INPUT_FILE, data_only=True)
    try:
        production = build_projects_production(wb)
        ws = wb["LNG Projects"]
        projects = []
        for r in range(2, ws.max_row + 1):
            name = ws.cell(r, 1).value
            if not name:
                continue
            cell = lambda c: ws.cell(r, c).value
            operator = (cell(12) or "").strip()
            partners_str = (cell(13) or "").strip()
            primary_stake = cell(14)
            stakes = [cell(c) for c in range(15, 22)]  # S1..S7
            partners = [p.strip() for p in partners_str.split(",") if p.strip()] if partners_str else []
            # Companies for the filter = operator name(s) (split on "/") + partners.
            op_names = [o.strip() for o in operator.split("/") if o.strip()] if operator else []
            companies = []
            for c in op_names + partners:
                if c not in companies:
                    companies.append(c)
            # Owner/stake rows for the project-detail panel.
            owners = []
            if operator:
                owners.append({"name": operator, "stake": primary_stake})
            for i, p in enumerate(partners):
                owners.append({"name": p, "stake": stakes[i] if i < len(stakes) else None})
            projects.append({
                "name": str(name).strip(), "country": cell(2), "region": cell(3),
                "status": cell(4) or "", "unrisked": cell(5), "unrisked_bcfd": cell(6),
                "cos": cell(7), "risked": cell(8), "start": cell(9) or "",
                "util_forecast": cell(10), "util_decline": cell(11),
                "operator": operator, "owners": owners, "companies": companies,
            })
    finally:
        wb.close()

    region_order = ["Asia", "Australia", "LatAm", "MENA and Europe",
                    "North America", "Russia", "Sub-Saharan Africa"]
    status_order = ["Producing", "Under construction", "Pre-FID", "Shut-down"]
    countries = sorted(set(p["country"] for p in projects if p["country"]))
    companies = sorted(set(c for p in projects for c in p["companies"]))
    print(f"  Projects blob: {len(projects)} mains, {len(countries)} countries, {len(companies)} companies")
    return {
        "key": config["key"], "tab_label": config["tab_label"], "title": config["title"],
        "projects_tab": True,
        "projects": projects,
        "region_order": region_order, "status_order": status_order,
        "countries": countries, "companies": companies,
        # Units for capacity (Projects) + production (Supply Outlook). Base = mmt.
        "base_unit": "mmt", "units": LNG_UNITS, "default_unit": "mmt",
        "unit_config": LNG_UNIT_CONFIG,
        # Project-level production period views (rows = projects; JS groups them).
        "production": production,
        # latest actual month (shared with Global LNG) caps the range chart's
        # current-year line; region colors keep the stacked area consistent.
        "latest_actual": detect_latest_actual(LNG_IMPORTS_CFG),
        "region_colors": LNG_EXPORT_COLORS,
        "selectable_start": DISPLAY_START_YEAR, "selectable_end": DISPLAY_END_YEAR,
    }


def build_projects_production(wb):
    """Build project-level production period views from the 'LNG Proj Production'
    sheet (row 1 = month dates from col B; col A = project name). Rows are
    projects; the JS groups them into region/country totals on the fly."""
    ws = wb["LNG Proj Production"]
    dates = []
    last_col = 1
    for c in range(2, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if isinstance(v, datetime):
            dates.append(v); last_col = c
        elif v is not None:
            dates.append(v); last_col = c
    days = [monthrange(d.year, d.month)[1] if isinstance(d, datetime) else 30 for d in dates]

    rows = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        if not name:
            continue
        vals = []
        for c in range(2, last_col + 1):
            v = ws.cell(row=r, column=c).value
            vals.append(float(v) if isinstance(v, (int, float)) else 0.0)
        rows.append({"label": str(name).strip(), "values": vals})

    period_results = aggregate_monthly_to_periods(dates, days)
    views = {}
    for view_name, period_cols in period_results.items():
        aggregated = compute_period_values(rows, period_cols, [], [])   # production = pure flow (sum)
        views[view_name] = {
            "columns": [c["label"] for c in period_cols],
            "short_columns": [c["short"] for c in period_cols],
            "col_meta": [{"label": c["label"], "short": c["short"], "year": c.get("year", 0),
                          "month": c.get("month"), "days": c["days"]} for c in period_cols],
            "days": [c["days"] for c in period_cols],
            "rows": [{"label": a["label"], "base": a["base_values"]} for a in aggregated],
        }
    print(f"  Production views built: {len(rows)} projects x {len(dates)} months")
    return {"views": views}


def detect_latest_actual(config):
    """Find the latest month with real data.

    The Monthly Imports tab leaves the 'Unaccounted demand' row EMPTY for actual
    months and populates it for forecast months. So the latest actual month is
    the month just before the first non-empty Unaccounted demand cell.
    Returns {"year","month","index"} where index is the 0-based monthly index,
    or None if the marker row can't be found.
    """
    marker = config.get("unaccounted_row")
    if not marker:
        return None
    wb = openpyxl.load_workbook(LNG_INPUT_FILE, data_only=True)
    try:
        ws = wb[config["sheet"]]
        date_row = config["date_row"]
        start_col = config.get("data_start_col", 2)
        # last data column = last date in date_row
        last_col = start_col
        for c in range(start_col, ws.max_column + 1):
            if ws.cell(row=date_row, column=c).value is not None:
                last_col = c
        # find marker row
        marker_row = None
        for r in range(config["data_start_row"], (config.get("data_end_row") or ws.max_row) + 1):
            if str(ws.cell(row=r, column=1).value or "").strip() == marker:
                marker_row = r
                break
        if marker_row is None:
            return None
        first_forecast = None
        for ci, c in enumerate(range(start_col, last_col + 1)):
            if ws.cell(row=marker_row, column=c).value is not None:
                first_forecast = ci
                break
        if first_forecast is None or first_forecast == 0:
            return None
        idx = first_forecast - 1
        d = ws.cell(row=date_row, column=start_col + idx).value
        return {"year": d.year, "month": d.month, "index": idx}
    finally:
        wb.close()


def build_composite_blob(config):
    """Build a composite tab (Global LNG): multiple sub-datasets sharing one tab
    and the top Period/Unit/From/To bar."""
    subs = {}
    sub_order = []
    for sub_cfg in config["subs"]:
        print(f"\n--- Building LNG sub-dataset: {sub_cfg['key']} ({sub_cfg['title']}) ---")
        # Short panel key the JS uses: lng_imports -> imports, lng_exports -> exports.
        panel = sub_cfg["key"].replace("lng_", "")
        subs[panel] = build_dataset_blob(sub_cfg)
        sub_order.append(panel)

    # Latest actual month (from the imports Unaccounted demand row) drives chart
    # defaults: range chart current line caps here; stacked default window is
    # Jan 2025 -> this month + 8.
    latest_actual = None
    for sub_cfg in config["subs"]:
        if sub_cfg.get("unaccounted_row"):
            latest_actual = detect_latest_actual(sub_cfg)
            break
    if latest_actual:
        print(f"  Latest actual month: {latest_actual['year']}-{latest_actual['month']:02d} (index {latest_actual['index']})")
    else:
        print("  WARNING: latest actual month not detected; charts will fall back to display range")

    shared_views = subs[sub_order[0]]["views"]   # imports + exports share the date axis
    return {
        "key": config["key"],
        "tab_label": config["tab_label"],
        "title": config["title"],
        "composite": True,
        "base_unit": config["base_unit"],
        "units": config["units"],
        "default_unit": config["default_unit"],
        "unit_config": config["unit_config"],
        "selectable_start": DISPLAY_START_YEAR,
        "selectable_end": DISPLAY_END_YEAR,
        "views": shared_views,            # for the shared top From/To + period note
        "latest_actual": latest_actual,
        "sub_order": sub_order,
        "sub": subs,
    }


# ============================================================
# ASSET LOADING
# ============================================================
def load_assets():
    assets = {}
    if os.path.exists(LOGO_FILE):
        with open(LOGO_FILE, "rb") as f:
            assets["logo_b64"] = base64.b64encode(f.read()).decode("utf-8")
    else:
        assets["logo_b64"] = ""
    if os.path.exists(FONT_FILE):
        with open(FONT_FILE, "rb") as f:
            assets["font_b64"] = base64.b64encode(f.read()).decode("utf-8")
    else:
        assets["font_b64"] = ""
    return assets


# ============================================================
# HTML GENERATION
# ============================================================
def generate_html(datasets_by_key, ordered_keys, assets):
    master = {
        "datasets": datasets_by_key,
        "order": ordered_keys,
        "generated": datetime.now().strftime("%Y-%m-%d %H:%M"),
    }
    data_json = json.dumps(master, separators=(",", ":"))
    logo_b64 = assets.get("logo_b64", "")
    font_b64 = assets.get("font_b64", "")
    generated = master["generated"]
    db = COLORS["dark_blue"]
    grey = COLORS["grey"]
    red = COLORS["red"]
    card = COLORS["card_bg"]
    border = COLORS["border"]
    grid = COLORS["grid"]

    css = """
@font-face {
    font-family: 'Gotham Book';
    src: url('data:font/opentype;base64,""" + font_b64 + """') format('opentype');
    font-weight: normal; font-style: normal;
}
* { margin:0; padding:0; box-sizing:border-box; }
body {
    font-family: 'Gotham Book', 'Segoe UI', Calibri, sans-serif;
    background: #ffffff; color: """ + db + """; font-size: 13px; line-height: 1.4;
}
.header {
    text-align: center; padding: 20px 20px 10px;
    border-bottom: 2px solid """ + db + """; margin-bottom: 12px;
}
.header img { height: 70px; object-fit: contain; margin-bottom: 8px; }
.header h1 {
    font-size: 18px; font-weight: bold; color: """ + db + """;
    letter-spacing: 1px; text-transform: uppercase;
}

.dataset-toggle-bar {
    display: flex; justify-content: center; align-items: center;
    padding: 8px 20px 14px; gap: 0;
}
.dataset-toggle-btn {
    font-family: 'Gotham Book', 'Segoe UI', Calibri, sans-serif;
    font-size: 14px; font-weight: bold; padding: 9px 30px;
    border: 2px solid """ + db + """; background: #ffffff; color: """ + db + """;
    cursor: pointer; transition: all 0.2s ease;
}
.dataset-toggle-btn:first-child { border-radius: 6px 0 0 6px; border-right: 1px solid """ + db + """; }
.dataset-toggle-btn:last-child  { border-radius: 0 6px 6px 0; border-left: 1px solid """ + db + """; }
.dataset-toggle-btn.active { background: """ + db + """; color: #ffffff; }
.dataset-toggle-btn:hover:not(.active) { background: """ + card + """; }

.controls {
    display: flex; justify-content: center; align-items: center; gap: 20px;
    padding: 14px 24px; background: """ + card + """;
    border: 1px solid """ + border + """; border-radius: 12px;
    margin: 0 20px 8px; flex-wrap: wrap;
    box-shadow: 0 1px 4px rgba(39, 41, 98, 0.06);
}
.control-group { display: flex; align-items: center; gap: 8px; }
.control-group label {
    font-size: 11px; font-weight: bold; text-transform: uppercase;
    color: """ + grey + """; letter-spacing: 0.5px;
}
.control-group select {
    font-family: 'Gotham Book', 'Segoe UI', Calibri, sans-serif;
    font-size: 12px; padding: 8px 32px 8px 14px;
    border: 1.5px solid rgba(39, 41, 98, 0.25); border-radius: 8px;
    color: """ + db + """; background: #ffffff; cursor: pointer;
    appearance: none; -webkit-appearance: none;
    background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='10' height='6'%3E%3Cpath d='M0 0l5 6 5-6z' fill='%23272962'/%3E%3C/svg%3E");
    background-repeat: no-repeat; background-position: right 10px center;
    transition: border-color 0.2s, box-shadow 0.2s;
}
.control-group select:hover { border-color: """ + db + """; }
.control-group select:focus {
    outline: none; border-color: """ + db + """;
    box-shadow: 0 0 0 3px rgba(39, 41, 98, 0.12);
}
.reset-btn {
    font-family: 'Gotham Book', 'Segoe UI', Calibri, sans-serif;
    font-size: 11px; padding: 7px 16px; border-radius: 8px;
    border: 1.5px solid rgba(39, 41, 98, 0.25); background: #fff;
    color: """ + db + """; cursor: pointer; transition: all 0.2s;
}
.reset-btn:hover { background: """ + card + """; border-color: """ + db + """; }
.period-note {
    text-align: center; font-size: 10px; color: """ + grey + """;
    margin: 0 20px 8px; font-style: italic;
}

.table-container {
    margin: 0 20px 10px; overflow-x: auto;
    border: 1px solid """ + border + """; border-radius: 10px;
    max-height: calc(100vh - 380px); overflow-y: auto;
    box-shadow: 0 1px 4px rgba(39, 41, 98, 0.06);
}
table { border-collapse: collapse; width: max-content; min-width: 100%; }
thead th {
    position: sticky; top: 0; z-index: 10;
    background: """ + db + """; color: #ffffff;
    font-size: 11px; font-weight: normal; padding: 10px 12px;
    text-align: right; white-space: nowrap;
    border-bottom: 2px solid """ + db + """; letter-spacing: 0.3px;
}
thead th:first-child {
    text-align: left; position: sticky; left: 0; z-index: 20;
    min-width: 250px; background: """ + db + """;
    border-top-left-radius: 9px;
}
thead th:last-child { border-top-right-radius: 9px; }
tbody td {
    padding: 6px 12px; text-align: right; font-size: 12px;
    border-bottom: 1px solid """ + grid + """;
    white-space: nowrap; font-variant-numeric: tabular-nums;
}
tbody td:first-child {
    text-align: left; position: sticky; left: 0; z-index: 5;
    background: #ffffff; border-right: 1px solid """ + grid + """;
    font-size: 12px; min-width: 250px;
}
tr.parent-row td { font-weight: bold; background: #fbfbfd; }
tr.parent-row td:first-child {
    background: #fbfbfd; cursor: pointer; user-select: none;
}
tr.parent-row td:first-child:hover { color: """ + red + """; }
tr.child-row td { font-weight: normal; font-size: 11.5px; color: #444466; }
tr.child-row td:first-child { padding-left: 32px; }
tr.child-row.hidden { display: none; }
tr.standalone-row td { font-weight: bold; background: #fbfbfd; }
tr.standalone-row td:first-child { background: #fbfbfd; }
.toggle-arrow {
    display: inline-block; width: 14px; font-size: 10px;
    color: """ + grey + """; transition: transform 0.15s ease;
}
.toggle-arrow.expanded { transform: rotate(90deg); }
tr.pct-row td { font-style: italic; color: """ + grey + """; }
tr.pct-row td:first-child { background: #fbfbfd; }
td.col-highlight { background-color: rgba(39, 41, 98, 0.07) !important; }
td.row-highlight { background-color: rgba(39, 41, 98, 0.07) !important; }
td.cell-highlight { background-color: rgba(39, 41, 98, 0.13) !important; }

/* Growth table */
.growth-section { margin: 0 20px; }
.growth-controls {
    display: flex; justify-content: center; align-items: center; gap: 16px;
    padding: 10px 20px; background: """ + card + """;
    border: 1px solid """ + border + """; border-radius: 10px 10px 0 0;
    border-bottom: none; flex-wrap: wrap;
}
.growth-controls .control-group label { font-size: 10px; }
.growth-toggle { display: flex; gap: 0; }
.growth-toggle button {
    font-family: 'Gotham Book', 'Segoe UI', Calibri, sans-serif;
    font-size: 11px; padding: 6px 14px; border: 1.5px solid rgba(39, 41, 98, 0.25);
    background: #fff; color: """ + db + """; cursor: pointer; transition: all 0.2s;
}
.growth-toggle button:first-child { border-radius: 7px 0 0 7px; }
.growth-toggle button:last-child { border-radius: 0 7px 7px 0; border-left: none; }
.growth-toggle button.active {
    background: """ + db + """; color: #fff; border-color: """ + db + """;
}
.growth-table-container {
    overflow-x: auto; border: 1px solid """ + border + """;
    border-radius: 0 0 10px 10px;
    max-height: 400px; overflow-y: auto;
    box-shadow: 0 1px 4px rgba(39, 41, 98, 0.06);
    margin-bottom: 15px;
}
.growth-table-container table { border-collapse: collapse; width: max-content; min-width: 100%; }
.growth-table-container thead th {
    position: sticky; top: 0; z-index: 10;
    background: """ + db + """; color: #ffffff;
    font-size: 11px; font-weight: normal; padding: 8px 12px;
    text-align: right; white-space: nowrap;
    border-bottom: 2px solid """ + db + """; letter-spacing: 0.3px;
}
.growth-table-container thead th:first-child {
    text-align: left; position: sticky; left: 0; z-index: 20;
    min-width: 250px; background: #3a3d70;
    border-top-left-radius: 0;
}
.growth-table-container tbody td {
    padding: 5px 12px; text-align: right; font-size: 11.5px;
    border-bottom: 1px solid """ + grid + """;
    white-space: nowrap; font-variant-numeric: tabular-nums;
}
.growth-table-container tbody td:first-child {
    text-align: left; position: sticky; left: 0; z-index: 5;
    background: #ffffff; border-right: 1px solid """ + grid + """;
    font-size: 11.5px; min-width: 250px; font-weight: bold;
}
.growth-table-container tr.child-row td:first-child {
    font-weight: normal; padding-left: 32px; color: #444466;
}
.growth-table-container tr.child-row.hidden { display: none; }
td.g-pos { color: #0C7A1E; }
td.g-neg { color: #C00000; }

.table-container::-webkit-scrollbar, .growth-table-container::-webkit-scrollbar { height: 8px; width: 8px; }
.table-container::-webkit-scrollbar-track, .growth-table-container::-webkit-scrollbar-track { background: #f0f0f4; border-radius: 4px; }
.table-container::-webkit-scrollbar-thumb, .growth-table-container::-webkit-scrollbar-thumb { background: """ + grey + """; border-radius: 4px; }
.footer { text-align: center; padding: 12px; font-size: 10px; color: """ + grey + """; }
.unit-label { font-size: 10px; color: rgba(255,255,255,0.6); font-style: italic; margin-left: 6px; }
@media (max-width: 1024px) {
    .header { padding: 16px 16px 8px; }
    .header img { height: 55px; }
    .header h1 { font-size: 16px; }
    .dataset-toggle-btn { font-size: 13px; padding: 8px 22px; }
    .controls { margin: 0 12px 8px; padding: 12px 18px; gap: 14px; }
    .table-container, .growth-section { margin-left: 12px; margin-right: 12px; }
    thead th { font-size: 10.5px; padding: 8px 10px; }
    tbody td { font-size: 11.5px; padding: 5px 10px; }
    thead th:first-child, tbody td:first-child { min-width: 200px; }
    tr.child-row td:first-child { padding-left: 26px; }
}
@media (max-width: 768px) {
    .header { padding: 14px 12px 8px; margin-bottom: 8px; }
    .header img { height: 45px; margin-bottom: 6px; }
    .header h1 { font-size: 14px; letter-spacing: 0.5px; }
    .dataset-toggle-bar { padding: 6px 12px 10px; }
    .dataset-toggle-btn { font-size: 12px; padding: 7px 18px; }
    .controls {
        flex-direction: column; gap: 10px; padding: 10px 14px;
        margin: 0 8px 8px; border-radius: 10px;
    }
    .control-group { width: 100%; justify-content: space-between; }
    .control-group select { flex: 1; font-size: 13px; padding: 10px 32px 10px 12px; }
    .control-group label { font-size: 10px; min-width: 50px; }
    .table-container, .growth-section { margin-left: 8px; margin-right: 8px; }
    .table-container { max-height: calc(100vh - 340px); -webkit-overflow-scrolling: touch; }
    table { font-size: 11px; }
    thead th { font-size: 10px; padding: 8px 8px; }
    tbody td { font-size: 11px; padding: 5px 8px; }
    thead th:first-child, tbody td:first-child { min-width: 150px; font-size: 10.5px; }
    tr.child-row td:first-child { padding-left: 22px; }
    .growth-controls { flex-direction: column; gap: 8px; padding: 8px 12px; }
    .footer { font-size: 9px; padding: 10px 8px; }
    .unit-label { display: none; }
}
@media (max-width: 480px) {
    .header { padding: 10px 8px 6px; margin-bottom: 6px; }
    .header img { height: 36px; margin-bottom: 4px; }
    .header h1 { font-size: 12px; }
    .dataset-toggle-btn { font-size: 11px; padding: 6px 14px; }
    .controls { margin: 0 6px 6px; padding: 8px 10px; gap: 8px; border-radius: 8px; }
    .control-group select { font-size: 12px; padding: 8px 28px 8px 10px; }
    .table-container, .growth-section { margin-left: 6px; margin-right: 6px; }
    thead th { font-size: 9px; padding: 6px 6px; }
    tbody td { font-size: 10px; padding: 4px 6px; }
    thead th:first-child, tbody td:first-child { min-width: 120px; font-size: 9.5px; }
    tr.child-row td:first-child { padding-left: 18px; }
    .toggle-arrow { width: 10px; font-size: 8px; }
}

/* ============================================================
   Multi-pane layout (table + charts)
   ============================================================ */
.main-grid {
    display: grid;
    grid-template-columns: 1fr;
    grid-template-areas:
        "table"
        "growth";
    gap: 16px;
    margin: 0 20px 12px;
}
/* Datasets that declare chart_area use a 2-tables-on-top / 2-charts-below grid.
   Keyed off the .charts-area body class (set when DATA.chart_area), so any
   dataset (power now, gas/LNG/storage later) opts in via config, not CSS. */
body.charts-area .main-grid {
    grid-template-columns: 1fr 1fr;
    grid-template-rows: auto auto auto;
    grid-template-areas:
        "table        table"
        "growth       growth"
        "seasonality  buildup";
}
.quad-table       { grid-area: table; }
.quad-growth      { grid-area: growth; }
.quad-seasonality { grid-area: seasonality; }
.quad-buildup     { grid-area: buildup; }
.grid-quad { display: flex; flex-direction: column; min-width: 0; }
.grid-quad .table-container,
.grid-quad .growth-table-container {
    margin: 0; flex: 1;
}
.grid-quad .growth-section { margin: 0; }

/* Real charts show per-slot, based on which charts the dataset provides.
   Left slot = range/seasonality (has-range); right slot = buildup (has-buildup). */
.chart-quadrant { display: none; }
body.has-range  .quad-seasonality.chart-quadrant { display: flex; }
body.has-buildup .quad-buildup.chart-quadrant    { display: flex; }

/* Placeholder chart boxes: shown per-slot when that slot has no real chart yet. */
.chart-placeholder-quad { display: none; flex-direction: column; min-width: 0; }
.chart-placeholder-quad.left  { grid-area: seasonality; }
.chart-placeholder-quad.right { grid-area: buildup; }
body.charts-area:not(.has-range)   .chart-placeholder-quad.left  { display: flex; }
body.charts-area:not(.has-buildup) .chart-placeholder-quad.right { display: flex; }

/* Embed tabs (Storage / LNG Sendout): an external live page in an iframe.
   When active, the data dashboard (controls + grid) is hidden and the iframe
   fills the area below the toggle bar. */
.embed-container { display: none; margin: 0 20px 12px; }
body.embed-active .embed-container { display: block; }
body.embed-active .controls,
body.embed-active .period-note,
body.embed-active .main-grid { display: none; }
.embed-container iframe {
    width: 100%;
    height: calc(100vh - 150px);
    min-height: 620px;
    border: none;
    display: block;
    background: #ffffff;
}
@media (max-width: 768px) {
    .embed-container { margin: 0 8px 8px; }
    .embed-container iframe { height: calc(100vh - 120px); }
}
.chart-placeholder-box {
    flex: 1;
    display: flex;
    align-items: center;
    justify-content: center;
    min-height: 320px;
    border: 1px dashed """ + border + """;
    border-radius: 8px;
    background: """ + card + """;
    color: """ + grey + """;
    font-size: 13px;
    letter-spacing: 0.3px;
}

/* Value-mode toggle above the generation table (Power only) */
.table-value-toggle-bar { display: none; padding: 0 0 8px; }
body.dataset-power .table-value-toggle-bar { display: flex; justify-content: flex-start; }
.value-toggle { display: flex; gap: 0; }
.value-toggle button {
    font-family: 'Gotham Book', 'Segoe UI', Calibri, sans-serif;
    font-size: 11px; padding: 6px 14px;
    border: 1.5px solid rgba(39, 41, 98, 0.25);
    background: #fff; color: """ + db + """;
    cursor: pointer; transition: all 0.15s;
}
.value-toggle button:first-child { border-radius: 7px 0 0 7px; }
.value-toggle button:last-child  { border-radius: 0 7px 7px 0; }
.value-toggle button:not(:first-child) { border-left: none; }  /* clean joins for 3-button groups */
.value-toggle button.active { background: """ + db + """; color: #fff; border-color: """ + db + """; }

.unit-efficiency-note {
    font-size: 10px; color: """ + grey + """; font-style: italic;
    margin-left: 2px;
}

.chart-controls {
    display: flex; align-items: center; gap: 14px;
    padding: 10px 14px; background: """ + card + """;
    border: 1px solid """ + border + """; border-radius: 10px 10px 0 0;
    border-bottom: none; flex-wrap: wrap;
    box-shadow: 0 1px 4px rgba(39, 41, 98, 0.06);
}
.chart-controls .control-group label { font-size: 10px; }
.chart-controls .control-group select {
    font-size: 11.5px; padding: 6px 28px 6px 12px;
}

/* The buildup (right) slot carries different controls per dataset: power keeps
   Sources/Period/From/To; gas uses a chart-type toggle + Series picker. */
.buildup-controls-power, .buildup-controls-gas { display: none; }
body.dataset-power .buildup-controls-power { display: flex; }
body.dataset-gas   .buildup-controls-gas   { display: flex; }

.chart-title {
    text-align: center; font-size: 11px; font-weight: bold;
    color: """ + db + """; letter-spacing: 0.5px; text-transform: uppercase;
    padding: 8px 12px 4px;
    background: #ffffff;
    border-left: 1px solid """ + border + """;
    border-right: 1px solid """ + border + """;
}

.chart-canvas-wrap {
    position: relative;
    flex: 1;
    min-height: 360px;
    padding: 10px 12px 6px;
    background: #ffffff;
    border: 1px solid """ + border + """;
    border-top: none;
    border-bottom: none;
    box-shadow: 0 1px 4px rgba(39, 41, 98, 0.06);
}
.chart-canvas-wrap canvas { width: 100% !important; height: 100% !important; }

/* Custom HTML legend (replaces Chart.js native to control fade-on-hidden) */
.custom-legend {
    display: flex; flex-wrap: wrap; justify-content: center;
    gap: 6px 16px; padding: 10px 14px 12px;
    background: #ffffff;
    border: 1px solid """ + border + """;
    border-top: none;
    border-radius: 0 0 10px 10px;
    box-shadow: 0 1px 4px rgba(39, 41, 98, 0.06);
    font-family: 'Gotham Book', 'Segoe UI', Calibri, sans-serif;
    font-size: clamp(10px, 0.85vw, 13px);
    color: """ + db + """;
}
.cl-item {
    display: inline-flex; align-items: center; gap: 6px;
    cursor: pointer; user-select: none;
    opacity: 1;
    transition: opacity 0.18s;
    padding: 2px 4px;
}
.cl-item:hover { opacity: 0.85; }
.cl-item.cl-hidden { opacity: 0.28; }
.cl-marker {
    display: inline-block; flex-shrink: 0;
    width: 18px; height: 12px;
}
.cl-marker-box { border-radius: 2px; border: 1px solid rgba(39,41,98,0.18); }
.cl-marker-line {
    height: 3px; border-radius: 2px;
    align-self: center;
    background-color: var(--ml, currentColor);
}
.cl-marker-line.cl-marker-dashed {
    background-color: transparent;
    background-image: repeating-linear-gradient(
        90deg,
        var(--ml, currentColor) 0 4px,
        transparent 4px 7px
    );
}
.cl-text { white-space: nowrap; }

/* ============================================================
   Multi-select dropdown widget (Sources picker)
   ============================================================ */
.multi-select {
    position: relative; display: inline-block;
    font-family: 'Gotham Book', 'Segoe UI', Calibri, sans-serif;
}
.ms-button {
    font-family: inherit;
    font-size: 11.5px; padding: 6px 30px 6px 12px;
    border: 1.5px solid rgba(39, 41, 98, 0.25); border-radius: 8px;
    background: #ffffff; color: """ + db + """;
    cursor: pointer; min-width: 180px; text-align: left;
    appearance: none; -webkit-appearance: none;
    background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='10' height='6'%3E%3Cpath d='M0 0l5 6 5-6z' fill='%23272962'/%3E%3C/svg%3E");
    background-repeat: no-repeat; background-position: right 10px center;
    transition: border-color 0.15s, box-shadow 0.15s;
    white-space: nowrap; overflow: hidden; text-overflow: ellipsis;
}
.ms-button:hover { border-color: """ + db + """; }
.ms-button:focus { outline: none; border-color: """ + db + """; box-shadow: 0 0 0 3px rgba(39, 41, 98, 0.12); }

.ms-panel {
    display: none; position: absolute; top: calc(100% + 4px); left: 0;
    min-width: 220px; background: #ffffff;
    border: 1.5px solid rgba(39, 41, 98, 0.25); border-radius: 8px;
    box-shadow: 0 4px 14px rgba(39, 41, 98, 0.15);
    z-index: 100; padding: 6px;
    max-height: 320px; overflow-y: auto;
}
.multi-select.open .ms-panel { display: block; }
.ms-item {
    display: flex; align-items: center; gap: 8px;
    padding: 6px 10px; cursor: pointer;
    font-size: 12px; color: """ + db + """;
    border-radius: 6px; user-select: none;
}
.ms-item:hover { background: """ + card + """; }
.ms-item input[type="checkbox"] {
    width: 14px; height: 14px; cursor: pointer; accent-color: """ + db + """;
}
.ms-item.ms-total {
    border-top: 1px solid """ + border + """;
    margin-top: 4px; padding-top: 8px;
    font-weight: bold;
}
.ms-item.ms-disabled { opacity: 0.4; cursor: not-allowed; }
.ms-item.ms-disabled input { cursor: not-allowed; }
/* Hierarchical (gas range chart) selector: group headers + indented members. */
.ms-item.ms-group-header { font-weight: bold; }
.ms-item.ms-child { padding-left: 28px; font-size: 11.5px; color: """ + db + """; }
.ms-item.ms-child span { opacity: 0.92; }
.ms-divider { border-top: 1px solid """ + border + """; margin: 4px 6px; }

/* Responsive: stack 2x2 grid on narrow screens */
@media (max-width: 1100px) {
    body.charts-area .main-grid {
        grid-template-columns: 1fr;
        grid-template-rows: auto auto auto auto;
        grid-template-areas:
            "table"
            "growth"
            "seasonality"
            "buildup";
    }
    .chart-canvas-wrap { min-height: 320px; }
}
@media (max-width: 768px) {
    .main-grid { margin: 0 8px 8px; gap: 10px; }
    .chart-controls { padding: 8px 10px; gap: 10px; }
    .ms-button { font-size: 12px; padding: 8px 28px 8px 10px; min-width: 150px; }
    .chart-canvas-wrap { min-height: 280px; padding: 8px; }
    .chart-title { font-size: 10px; padding: 6px 10px 3px; }
}

/* ============================================================
   Global LNG composite tab: 4 full-width tables stacked, then two
   rows of side-by-side charts (range, then build-up). Shown only when
   body.lng-composite; hides the single-table main-grid.
   ============================================================ */
.lng-grid { display: none; margin: 0 20px 12px; }
body.lng-composite .lng-grid { display: flex; flex-direction: column; gap: 16px; }
body.lng-composite .main-grid { display: none; }
.lng-block { display: flex; flex-direction: column; min-width: 0; }
.lng-block-title {
    font-size: 12px; font-weight: bold; color: """ + db + """;
    text-transform: uppercase; letter-spacing: 0.5px; padding: 2px 2px 6px;
}
.lng-grid .growth-controls .lng-block-title { padding: 0; margin-right: auto; }
.lng-grid .table-container { max-height: 340px; }
.lng-grid .growth-table-container { max-height: 320px; margin-bottom: 0; }
.lng-charts-row { display: grid; grid-template-columns: 1fr 1fr; gap: 16px; }
.lng-charts-row .grid-quad { min-width: 0; }
@media (max-width: 1100px) { .lng-charts-row { grid-template-columns: 1fr; } }
@media (max-width: 768px) { .lng-grid { margin: 0 8px 8px; } }

/* Exports hierarchical selector: region expand arrow + collapsible members. */
.ms-exp-arrow {
    display: inline-block; width: 14px; flex-shrink: 0;
    cursor: pointer; color: """ + grey + """; font-size: 10px;
    transition: transform 0.15s ease; text-align: center;
}
.ms-exp-arrow.expanded { transform: rotate(90deg); }
.ms-item.ms-exp-member { padding-left: 30px; }

/* ============================================================
   LNG Projects tab: sub-tabs + filter bar + Region/Country/Project tree.
   Shown only when body.projects-tab; hides the period/unit bar and the
   table/chart grids used by the other tabs.
   ============================================================ */
.projects-tab-wrap { display: none; margin: 0 20px 12px; }
body.projects-tab .projects-tab-wrap { display: block; }
body.projects-tab .controls,
body.projects-tab .period-note,
body.projects-tab .main-grid,
body.projects-tab .lng-grid,
body.projects-tab .embed-container { display: none; }

/* sub-tab switcher (Projects / Supply Outlook) */
.prj-subtabs { display: flex; justify-content: center; gap: 0; margin: 0 0 12px; }
.prj-subtab-btn {
    font-family: 'Gotham Book','Segoe UI',Calibri,sans-serif;
    font-size: 12.5px; font-weight: bold; padding: 8px 24px;
    border: 1.5px solid """ + db + """; background: #fff; color: """ + db + """;
    cursor: pointer; transition: all 0.15s;
}
.prj-subtab-btn:first-child { border-radius: 7px 0 0 7px; border-right: none; }
.prj-subtab-btn:last-child  { border-radius: 0 7px 7px 0; }
.prj-subtab-btn.active { background: """ + db + """; color: #fff; }
.prj-subtab-btn:hover:not(.active) { background: """ + card + """; }

.prj-pane { display: none; }
.prj-pane.active { display: block; }

/* filter bar */
.prj-filter-bar {
    display: flex; flex-wrap: wrap; align-items: flex-end; gap: 12px;
    padding: 12px 16px; background: """ + card + """;
    border: 1px solid """ + border + """; border-radius: 12px; margin-bottom: 12px;
}
.prj-filter-group { display: flex; flex-direction: column; gap: 4px; }
.prj-filter-group label {
    font-size: 10px; font-weight: bold; text-transform: uppercase;
    color: """ + grey + """; letter-spacing: 0.5px;
}
.prj-filter-reset {
    font-family: 'Gotham Book','Segoe UI',Calibri,sans-serif;
    font-size: 11px; padding: 7px 16px; border-radius: 8px;
    border: 1.5px solid rgba(39,41,98,0.25); background: #fff; color: """ + db + """;
    cursor: pointer; transition: all 0.15s; align-self: flex-end;
}
.prj-filter-reset:hover { background: #fff; border-color: """ + db + """; }
.ms-search {
    width: calc(100% - 12px); margin: 2px 6px 6px; padding: 5px 8px;
    border: 1px solid rgba(39,41,98,0.25); border-radius: 6px;
    font-family: inherit; font-size: 12px; color: """ + db + """;
}
.ms-item.ms-unavail { display: none; }

/* summary as rounded pill cards */
.prj-summary { display: flex; flex-wrap: wrap; gap: 14px; margin-bottom: 14px; }
.prj-pill {
    display: flex; align-items: baseline; gap: 10px;
    padding: 11px 26px; border-radius: 999px;
    background: #fff; border: 1.5px solid """ + border + """;
    box-shadow: 0 1px 4px rgba(39,41,98,0.07);
}
.prj-pill .v { font-size: 19px; font-weight: bold; color: """ + db + """; font-variant-numeric: tabular-nums; }
.prj-pill .k { font-size: 10px; text-transform: uppercase; letter-spacing: 0.5px; color: """ + grey + """; }
.prj-pill { transition: background 0.15s, border-color 0.15s; }
.prj-pill:hover { background: """ + db + """; border-color: """ + db + """; }
.prj-pill:hover .v { color: #fff; }
.prj-pill:hover .k { color: rgba(255,255,255,0.7); }

/* unit dropdown in the filter bar + Supply Outlook control row */
.prj-filter-group select, .prj-outlook-controls select {
    font-family: 'Gotham Book','Segoe UI',Calibri,sans-serif; font-size: 12px;
    padding: 7px 30px 7px 12px; border: 1.5px solid rgba(39,41,98,0.25); border-radius: 8px;
    color: """ + db + """; background: #fff; cursor: pointer; appearance: none; -webkit-appearance: none;
    background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='10' height='6'%3E%3Cpath d='M0 0l5 6 5-6z' fill='%23272962'/%3E%3C/svg%3E");
    background-repeat: no-repeat; background-position: right 10px center;
}
.prj-outlook-controls {
    display: flex; flex-wrap: wrap; align-items: flex-end; gap: 14px;
    padding: 12px 16px; margin-bottom: 12px;
    background: """ + card + """; border: 1px solid """ + border + """; border-radius: 12px;
}
.prj-outlook-controls .control-group { display: flex; flex-direction: column; gap: 4px; }
.prj-outlook-controls .control-group label {
    font-size: 10px; font-weight: bold; text-transform: uppercase;
    color: """ + grey + """; letter-spacing: 0.5px;
}
.prj-outlook-table tr.prj-row-region td,
.prj-outlook-table tr.prj-row-country td { font-variant-numeric: tabular-nums; }
.prj-table tr.prj-row-total td {
    background: #eef0f6; font-weight: bold; border-top: 2px solid """ + db + """;
    position: sticky; bottom: 0;
}
/* Supply Outlook charts row (range left, stacked area right) */
.prj-outlook-charts {
    display: grid; grid-template-columns: 1fr 1fr; gap: 16px; margin-top: 14px;
}
.prj-outlook-charts .grid-quad { min-width: 0; display: flex; flex-direction: column; }
@media (max-width: 1100px) { .prj-outlook-charts { grid-template-columns: 1fr; } }
/* Chart header: grey bar carrying a centered title (left chart also has the
   Average-range dropdown on the left; right chart has just the title). */
.prj-chart-head { position: relative; min-height: 42px; }
.prj-chart-head .control-group { position: relative; z-index: 1; }
.prj-chart-head .chart-title-inline {
    position: absolute; left: 14px; right: 14px; top: 0; bottom: 0;
    display: flex; align-items: center; justify-content: center;
    font-size: 11px; font-weight: bold; color: """ + db + """;
    letter-spacing: 0.5px; text-transform: uppercase; text-align: center;
    pointer-events: none;
}

/* Change-in-production controls (mirrors the gas growth controls) */
.prj-change-controls { display: flex; align-items: flex-end; gap: 14px; flex-wrap: wrap; padding: 14px 2px 8px; }
.prj-change-controls .control-group { display: flex; flex-direction: column; gap: 4px; }
.prj-change-controls .control-group label { font-size: 10px; font-weight: bold; text-transform: uppercase; color: """ + grey + """; letter-spacing: 0.5px; }
.prj-change-controls select, .prj-chart-filter-controls select {
    font-family: 'Gotham Book','Segoe UI',Calibri,sans-serif; font-size: 12px;
    padding: 7px 30px 7px 12px; border: 1.5px solid rgba(39,41,98,0.25); border-radius: 8px;
    color: """ + db + """; background: #fff; cursor: pointer; appearance: none; -webkit-appearance: none;
    background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='10' height='6'%3E%3Cpath d='M0 0l5 6 5-6z' fill='%23272962'/%3E%3C/svg%3E");
    background-repeat: no-repeat; background-position: right 10px center;
}

/* Chart filters: link-to-table toggle + (when unlinked) an independent filter set */
.prj-chart-filters { margin-top: 16px; border-top: 1px solid """ + border + """; padding-top: 14px; }
.prj-chart-filters-head { display: flex; align-items: center; gap: 14px; margin-bottom: 10px; }
.prj-link-btn {
    font-family: 'Gotham Book','Segoe UI',Calibri,sans-serif; font-size: 11px; font-weight: bold;
    padding: 7px 16px; border-radius: 999px; border: 1.5px solid """ + db + """;
    background: #fff; color: """ + db + """; cursor: pointer; transition: all 0.15s;
}
.prj-link-btn.active { background: """ + db + """; color: #fff; }
.prj-chart-filter-controls {
    display: flex; flex-wrap: wrap; align-items: flex-end; gap: 12px;
    padding: 12px 16px; background: """ + card + """; border: 1px solid """ + border + """; border-radius: 12px;
}

/* projects tree table */
.prj-table-container {
    border: 1px solid """ + border + """; border-radius: 10px; overflow: auto;
    max-height: calc(100vh - 360px);
    box-shadow: 0 1px 4px rgba(39,41,98,0.06);
}
/* table-layout:fixed -> column widths come from the header only, so expanding
   rows or switching periods never shifts the columns. */
.prj-table { border-collapse: collapse; width: 100%; min-width: 760px; table-layout: fixed; }
.prj-table thead th {
    position: sticky; top: 0; z-index: 10; background: """ + db + """; color: #fff;
    font-size: 11px; font-weight: normal; padding: 9px 12px; text-align: center; white-space: nowrap;
    overflow: hidden; text-overflow: ellipsis;
}
.prj-table thead th:first-child { text-align: left; }
.prj-table td {
    padding: 6px 12px; text-align: center; font-size: 12px; white-space: nowrap;
    border-bottom: 1px solid """ + grid + """; font-variant-numeric: tabular-nums;
    overflow: hidden; text-overflow: ellipsis;
}
.prj-table td:first-child { text-align: left; }
/* Projects assumptions table: fixed column widths (7 cols). */
#prjTableHead th:nth-child(1) { width: 24%; }
#prjTableHead th:nth-child(2) { width: 14%; }
#prjTableHead th:nth-child(3), #prjTableHead th:nth-child(4) { width: 12%; }
#prjTableHead th:nth-child(5) { width: 8%; }
#prjTableHead th:nth-child(6) { width: 12%; }
#prjTableHead th:nth-child(7) { width: 18%; }
/* Supply Outlook table: fixed first column, uniform value columns, scrolls wide. */
.prj-outlook-table { width: auto; min-width: 100%; }
.prj-outlook-table th, .prj-outlook-table td { width: 68px; }
.prj-outlook-table th:first-child, .prj-outlook-table td:first-child { width: 240px; }
.prj-row-region td { background: #eef0f6; font-weight: bold; cursor: pointer; }
.prj-row-country td { background: #f6f7fb; font-weight: bold; cursor: pointer; }
.prj-row-country td:first-child { padding-left: 28px; }
.prj-row-project td:first-child { padding-left: 48px; cursor: pointer; }
.prj-row-project:hover td { background: #fafbff; }
.prj-arrow { display: inline-block; width: 14px; font-size: 10px; color: """ + grey + """; transition: transform 0.15s; }
.prj-arrow.expanded { transform: rotate(90deg); }
.prj-detail td { background: #fbfbfe; padding: 0; border-bottom: 2px solid """ + border + """; }
.prj-detail-inner { padding: 16px 20px 18px 48px; display: flex; flex-wrap: wrap; gap: 14px 40px; align-items: flex-start; }
.prj-detail-col { min-width: 250px; flex: 1; }
.prj-detail-col.owners { flex: 1.1; }
.prj-detail-col h4 {
    font-size: 9.5px; text-transform: uppercase; letter-spacing: 0.7px;
    color: """ + grey + """; font-weight: bold; margin-bottom: 10px;
    border-bottom: 1px solid """ + border + """; padding-bottom: 5px;
}
.prj-stakebar { display: flex; height: 9px; border-radius: 5px; overflow: hidden; margin-bottom: 12px; background: #e9ebf2; }
.prj-stakebar span { display: block; height: 100%; }
.prj-owners { display: flex; flex-direction: column; gap: 7px; }
.prj-owner { display: flex; align-items: center; gap: 9px; font-size: 12px; }
.prj-owner .sw { width: 11px; height: 11px; border-radius: 3px; flex-shrink: 0; }
.prj-owner .nm { color: """ + db + """; flex: 1; }
.prj-owner .tag {
    font-size: 8.5px; text-transform: uppercase; letter-spacing: 0.4px;
    color: #fff; background: """ + db + """; padding: 1px 6px; border-radius: 9px; margin-left: 7px;
}
.prj-owner .pct { font-variant-numeric: tabular-nums; color: """ + db + """; font-weight: bold; }
.prj-stats { display: grid; grid-template-columns: repeat(auto-fit, minmax(128px, 1fr)); gap: 12px 20px; }
.prj-stat { display: flex; flex-direction: column; gap: 2px; }
.prj-stat .k { font-size: 9px; text-transform: uppercase; letter-spacing: 0.4px; color: """ + grey + """; }
.prj-stat .v { font-size: 15px; font-weight: bold; color: """ + db + """; font-variant-numeric: tabular-nums; }

/* status badge */
.prj-badge {
    display: inline-block; font-size: 10px; padding: 2px 8px; border-radius: 10px;
    font-weight: bold; letter-spacing: 0.2px;
}
.prj-badge.s-producing          { background: #E2F0DE; color: #0C5B19; }
.prj-badge.s-under-construction { background: #FDF1D6; color: #92591C; }
.prj-badge.s-pre-fid            { background: #E5EAF5; color: #272962; }
.prj-badge.s-shut-down          { background: #F2DADA; color: #C00000; }

.prj-outlook-stub {
    display: flex; align-items: center; justify-content: center; min-height: 300px;
    border: 1px dashed """ + border + """; border-radius: 10px; background: """ + card + """;
    color: """ + grey + """; font-size: 13px; text-align: center; padding: 20px;
}
@media (max-width: 768px) {
    .projects-tab-wrap { margin: 0 8px 8px; }
    .prj-filter-bar { gap: 8px; padding: 10px; }
    .prj-summary { gap: 14px; padding: 10px 12px; }
    .prj-summary .metric .v { font-size: 16px; }
}
"""

    js = r"""
var ROOT = __DATA_PLACEHOLDER__;
var ALL_DATASETS = ROOT.datasets;
var DATASET_ORDER = ROOT.order;
var currentKey = DATASET_ORDER[0];
var DATA = ALL_DATASETS[currentKey];

// Embed tabs (Storage / LNG Sendout) — external live pages shown via iframe.
var EMBED_TABS = __EMBED_TABS__;
var EMBED_MAP = {};
for (var _e=0; _e<EMBED_TABS.length; _e++) EMBED_MAP[EMBED_TABS[_e].key] = EMBED_TABS[_e];
// Every tab button, in bar order: data datasets first, then embed tabs.
var ALL_TAB_KEYS = DATASET_ORDER.concat(EMBED_TABS.map(function(t){ return t.key; }));
var embedActive = false;  // true while an embed (Storage/LNG) tab is showing

function setActiveTabButton(key) {
    for (var i=0;i<ALL_TAB_KEYS.length;i++) {
        var btn = document.getElementById('btnDataset-'+ALL_TAB_KEYS[i]);
        if (btn) btn.classList.toggle('active', ALL_TAB_KEYS[i] === key);
    }
}

// Show an embed tab: hide the data dashboard sections, show the iframe.
// The iframe src is set lazily and only when switching to a different embed
// page, so toggling data<->embed for the same page doesn't reload it.
function showEmbedTab(key) {
    var tab = EMBED_MAP[key];
    if (!tab) return;
    embedActive = true;
    setActiveTabButton(key);
    document.body.className = 'embed-active embed-' + key;
    document.getElementById('pageTitle').textContent = tab.label;
    document.title = 'Palissy Advisors - ' + tab.label;
    var frame = document.getElementById('embedFrame');
    if (frame && frame.getAttribute('data-key') !== key) {
        frame.src = tab.url;
        frame.setAttribute('data-key', key);
    }
}

var expandedMain = {};
var expandedGrowth = {};
var highlightedCol = -1, highlightedRow = -1;
var growthMode = 'pct';
var growthType = 'yoy';

var GROWTH_OPTS = {
    'Monthly':   [{k:'yoy',l:'Year-on-Year'},{k:'mom',l:'Month-on-Month'},{k:'ytd',l:'YTD Year-on-Year'},{k:'ltm',l:'Last 12 Months'},{k:'ytd_gy',l:'YTD Gas Year Year-on-Year'}],
    'Quarterly': [{k:'yoy',l:'Year-on-Year'},{k:'qoq',l:'Quarter-on-Quarter'},{k:'ytd',l:'YTD Year-on-Year'},{k:'ytd_gy',l:'YTD Gas Year Year-on-Year'}],
    'Annual CY': [{k:'yoy',l:'Year-on-Year'}],
    'Gas Year':  [{k:'yoy',l:'Year-on-Year'}],
    'Winter':    [{k:'yoy',l:'Year-on-Year'}],
    'Summer':    [{k:'yoy',l:'Year-on-Year'}]
};

function unitCfg(unitKey) { return DATA.unit_config[unitKey] || DATA.unit_config[DATA.default_unit]; }
function isVolUnit() { return !unitCfg(document.getElementById('unitSelector').value).isRate; }

function formatNum(val, isPct) {
    if (isPct) return (val*100).toFixed(1)+'%';
    var a = Math.abs(val);
    if (a < 0.005) return '0';
    if (a >= 100) return Math.round(val).toString().replace(/\B(?=(\d{3})+(?!\d))/g,',');
    if (a >= 10) return val.toFixed(1);
    return val.toFixed(2);
}

function formatGrowth(val, isPctRow) {
    if (val===null||val===undefined||!isFinite(val)) return '';
    if (growthMode==='pct') {
        var pct = val*100;
        var a = Math.abs(pct);
        if (a < 0.05) return '0.0%';
        if (a >= 100) return Math.round(pct)+'%';
        return pct.toFixed(1)+'%';
    }
    if (isPctRow) {
        var pp = val*100;
        var a = Math.abs(pp);
        if (a < 0.05) return '0.0 pp';
        return pp.toFixed(1)+' pp';
    }
    var cfg = unitCfg(document.getElementById('unitSelector').value);
    var converted = cfg.isRate ? (val * cfg.rateFactor) : (val * cfg.volFactor);
    return formatNum(converted, false);
}

function computeDisplayValue(baseVal, unitKey, isStock, isPct, days) {
    if (isPct) return baseVal;
    var cfg = unitCfg(unitKey);
    if (isStock) return baseVal * cfg.volFactor;
    if (cfg.isRate) return (baseVal / days) * cfg.rateFactor;
    return baseVal * cfg.volFactor;
}

function getStockLabel(label, unitKey) {
    var cfg = unitCfg(unitKey);
    if (cfg.isRate) return label + ' (' + cfg.volLabel + ')';
    return label;
}

function getHeaderUnitLabel(unitKey) {
    var cfg = unitCfg(unitKey);
    return cfg.isRate ? cfg.rateLabel : cfg.volLabel;
}

/* === DATE RANGE === */
function getSelectableIndices(view) {
    var m = view.col_meta, out = [];
    for (var i=0;i<m.length;i++) {
        if (m[i].year >= DATA.selectable_start && m[i].year <= DATA.selectable_end) out.push(i);
    }
    return out;
}

function getDefaultRange(viewName, selectable, view) {
    if (viewName==='Monthly') {
        var now = new Date(), cy = now.getFullYear();
        var sy = cy-1, ey = cy+1, m = view.col_meta;
        var s=null, e=null;
        for (var i=0;i<selectable.length;i++) {
            var idx=selectable[i];
            if (m[idx].year >= sy && s===null) s=i;
            if (m[idx].year <= ey) e=i;
        }
        return [s!==null?s:0, e!==null?e:selectable.length-1];
    }
    return [0, selectable.length-1];
}

function updateRangeSelectors() {
    var period = document.getElementById('periodSelector').value;
    var view = DATA.views[period];
    var sel = getSelectableIndices(view);
    var fs = document.getElementById('rangeFrom');
    var ts = document.getElementById('rangeTo');
    fs.innerHTML=''; ts.innerHTML='';
    for (var i=0;i<sel.length;i++) {
        var lbl = view.col_meta[sel[i]].label;
        fs.innerHTML += '<option value="'+i+'">'+lbl+'</option>';
        ts.innerHTML += '<option value="'+i+'">'+lbl+'</option>';
    }
    var def = getDefaultRange(period, sel, view);
    fs.value = def[0]; ts.value = def[1];
}

function getVisibleIndices() {
    var period = document.getElementById('periodSelector').value;
    var view = DATA.views[period];
    var sel = getSelectableIndices(view);
    var fi = parseInt(document.getElementById('rangeFrom').value);
    var ti = parseInt(document.getElementById('rangeTo').value);
    if (isNaN(fi)||isNaN(ti)) { var d=getDefaultRange(period,sel,view); fi=d[0]; ti=d[1]; }
    if (fi>ti) { var tmp=fi; fi=ti; ti=tmp; }
    var out = [];
    for (var i=fi;i<=ti;i++) out.push(sel[i]);
    return out;
}

function resetRange() {
    updateRangeSelectors();
    if (DATA.charts_enabled) {
        rebuildChartControls();
    }
    updateAll();
}

function updatePeriodNote() {
    var p = document.getElementById('periodSelector').value;
    var el = document.getElementById('periodNote');
    if (p==='Summer') el.textContent = 'Summer 2025 = April 2025 to September 2025';
    else if (p==='Winter') el.textContent = 'Winter 24/25 = October 2024 to March 2025';
    else el.textContent = '';
}

/* === GROWTH COMPUTATION === */
function getRate(base, days) { return base/days; }
function avgRate(rowBase, days, indices) {
    var tb=0, td=0;
    for (var i=0;i<indices.length;i++) { tb+=rowBase[indices[i]]; td+=days[indices[i]]; }
    return td===0?null:tb/td;
}
function sumVol(rowBase, indices) {
    var s=0;
    for (var i=0;i<indices.length;i++) s+=rowBase[indices[i]];
    return s;
}

function findYoY(meta, idx, vn) {
    var c=meta[idx];
    for (var i=0;i<meta.length;i++) {
        if (i===idx) continue;
        var m=meta[i];
        if (vn==='Monthly' && m.year===c.year-1 && m.month===c.month) return i;
        if (vn==='Quarterly' && m.year===c.year-1 && m.quarter===c.quarter) return i;
        if ((vn==='Annual CY'||vn==='Gas Year'||vn==='Winter'||vn==='Summer') && m.year===c.year-1) return i;
    }
    return -1;
}
function getYTDIndices(meta, idx) {
    var c=meta[idx], out=[];
    for (var i=0;i<meta.length;i++) { if (meta[i].year===c.year && meta[i].month<=c.month) out.push(i); }
    return out;
}
function getYTDQIndices(meta, idx) {
    var c=meta[idx], out=[];
    for (var i=0;i<meta.length;i++) { if (meta[i].year===c.year && meta[i].quarter<=c.quarter) out.push(i); }
    return out;
}
function getYTDGYIndices(meta, idx) {
    var c=meta[idx], gy=c.gas_year, out=[];
    for (var i=0;i<meta.length;i++) { if (meta[i].gas_year===gy && i<=idx) out.push(i); }
    return out;
}
function getLTMIndices(idx) {
    if (idx<11) return [];
    var out=[];
    for (var i=idx-11;i<=idx;i++) out.push(i);
    return out;
}

function computeGrowthCell(rowBase, days, meta, idx, vn, gt, isStock) {
    var useVol = (growthMode==='abs' && isVolUnit() && !isStock);
    function getVal(i) {
        if (isStock) return rowBase[i];
        if (growthMode==='pct') return getRate(rowBase[i],days[i]);
        return useVol ? rowBase[i] : getRate(rowBase[i],days[i]);
    }
    function getAgg(indices) {
        if (isStock) return rowBase[indices[indices.length-1]];
        if (growthMode==='pct') return avgRate(rowBase,days,indices);
        return useVol ? sumVol(rowBase,indices) : avgRate(rowBase,days,indices);
    }
    function result(cur,prev) {
        if (prev===null||cur===null) return null;       // no comparison period -> blank
        if (growthMode==='pct') {
            // prev==0: a 0 -> 0 period is genuinely 0% change (show it); 0 -> non-zero
            // is a start-from-nothing with no finite % (leave blank).
            if (prev===0) return cur===0 ? 0 : null;
            return cur/prev-1;
        }
        return cur-prev;
    }

    if (gt==='yoy') {
        var pi=findYoY(meta,idx,vn);
        return pi<0?null:result(getVal(idx),getVal(pi));
    }
    if (gt==='mom' || gt==='qoq') {
        return idx<=0?null:result(getVal(idx),getVal(idx-1));
    }
    if (gt==='ytd') {
        var curI,prevI;
        if (vn==='Monthly') {
            curI=getYTDIndices(meta,idx); var c=meta[idx]; prevI=[];
            for(var i=0;i<meta.length;i++){if(meta[i].year===c.year-1&&meta[i].month<=c.month)prevI.push(i);}
        } else {
            curI=getYTDQIndices(meta,idx); var c=meta[idx]; prevI=[];
            for(var i=0;i<meta.length;i++){if(meta[i].year===c.year-1&&meta[i].quarter<=c.quarter)prevI.push(i);}
        }
        if(curI.length===0||prevI.length<curI.length) return null;
        return result(getAgg(curI),getAgg(prevI));
    }
    if (gt==='ltm') {
        var curI=getLTMIndices(idx),prevI=getLTMIndices(idx-12);
        if(curI.length<12||prevI.length<12) return null;
        return result(getAgg(curI),getAgg(prevI));
    }
    if (gt==='ytd_gy') {
        var curI=getYTDGYIndices(meta,idx);
        if(curI.length===0) return null;
        var prevGY=meta[idx].gas_year-1,prevAll=[];
        for(var i=0;i<meta.length;i++){if(meta[i].gas_year===prevGY)prevAll.push(i);}
        var prevI=prevAll.slice(0,curI.length);
        if(prevI.length<curI.length) return null;
        return result(getAgg(curI),getAgg(prevI));
    }
    return null;
}

function updateGrowthTypeSelector() {
    var p = document.getElementById('periodSelector').value;
    var sel = document.getElementById('growthTypeSelector');
    var opts = GROWTH_OPTS[p] || [{k:'yoy',l:'Year-on-Year'}];
    sel.innerHTML = '';
    for (var i=0;i<opts.length;i++) {
        sel.innerHTML += '<option value="'+opts[i].k+'"'+(i===0?' selected':'')+'>'+opts[i].l+'</option>';
    }
    growthType = opts[0].k;
}

function setGrowthMode(mode) {
    growthMode = mode;
    document.getElementById('btnPct').className = mode==='pct'?'active':'';
    document.getElementById('btnAbs').className = mode==='abs'?'active':'';
    updateGrowthTable();
}

function rebuildUnitSelector() {
    var sel = document.getElementById('unitSelector');
    sel.innerHTML = '';
    for (var i=0;i<DATA.units.length;i++) {
        var u = DATA.units[i];
        var selected = (u === DATA.default_unit) ? ' selected' : '';
        sel.innerHTML += '<option value="'+u+'"'+selected+'>'+u+'</option>';
    }
    updateTableValueModeLabels();
}

// Value-mode toggle for the table (Generation vs % of Total Fuel)
var tableValueMode = 'gen'; // 'gen' or 'pct'

function setTableValueMode(mode) {
    tableValueMode = mode;
    var bGen = document.getElementById('btnTblGen');
    var bPct = document.getElementById('btnTblPct');
    if (bGen) bGen.className = mode === 'gen' ? 'active' : '';
    if (bPct) bPct.className = mode === 'pct' ? 'active' : '';
    updateTable();
}

function updateTableValueModeLabels() {
    var btn = document.getElementById('btnTblGen');
    if (!btn) return;
    var unitSel = document.getElementById('unitSelector');
    if (!unitSel || !unitSel.value) return;
    var unitLabel = getHeaderUnitLabel(unitSel.value);
    btn.textContent = 'Generation (' + unitLabel + ')';
}

// Returns ", at 39% efficiency" when a BCFE-family unit is active, else "".
function efficiencySuffix() {
    var unitSel = document.getElementById('unitSelector');
    if (!unitSel || !unitSel.value) return '';
    var u = unitSel.value;
    return (u === 'BCFE' || u === 'BCFE/d') ? ', at 39% efficiency' : '';
}

function updateEfficiencyNote() {
    var noteEl = document.getElementById('unitEfficiencyNote');
    if (!noteEl) return;
    var unitSel = document.getElementById('unitSelector');
    if (!unitSel) { noteEl.textContent = ''; return; }
    var u = unitSel.value;
    if (u === 'BCFE' || u === 'BCFE/d') {
        noteEl.textContent = '(at 39% gas-to-power efficiency)';
    } else {
        noteEl.textContent = '';
    }
}

function updateDatasetUI() {
    document.getElementById('pageTitle').textContent = DATA.title;
    document.title = 'Palissy Advisors - ' + DATA.title;
    var bodyCls = 'dataset-' + currentKey;
    if (DATA.projects_tab) bodyCls += ' projects-tab';
    if (DATA.composite) bodyCls += ' lng-composite';
    if (DATA.chart_area) bodyCls += ' charts-area';
    if (DATA.has_range) bodyCls += ' has-range';
    if (DATA.has_buildup) bodyCls += ' has-buildup';
    document.body.className = bodyCls;
    setActiveTabButton(currentKey);
}

function switchDataset(key) {
    if (key === currentKey && !embedActive) return;  // leaving an embed tab must re-render
    embedActive = false;
    currentKey = key;
    DATA = ALL_DATASETS[key];
    expandedMain = {};
    expandedGrowth = {};
    // The two chart slots share canvases across datasets — tear down any live
    // Chart.js instances so the incoming dataset builds fresh ones.
    if (seasonalityChart) { seasonalityChart.destroy(); seasonalityChart = null; }
    if (buildupChart)     { buildupChart.destroy();     buildupChart = null; }
    updateDatasetUI();
    if (DATA.projects_tab) { prjInit(); return; }   // own filter bar; no unit/period machinery
    rebuildUnitSelector();
    updateRangeSelectors();
    updateGrowthTypeSelector();
    updatePeriodNote();
    if (DATA.composite) {
        lngInit();   // builds selectors + renders all 4 tables and 4 charts
    } else if (DATA.charts_enabled) {
        rebuildChartControls();
        updateAll();
        updateCharts();
    } else if (DATA.chart_groups) {
        rebuildGasChartControls();
        updateAll();
    } else {
        updateAll();
    }
}

/* === MAIN TABLE === */
function updateTable() {
    var period = document.getElementById('periodSelector').value;
    var unitKey = document.getElementById('unitSelector').value;
    var view = DATA.views[period];
    if (!view) return;

    var vis = getVisibleIndices();
    var days = view.days;
    var rows = view.rows;
    var hierarchy = DATA.hierarchy;
    var unitLabel = getHeaderUnitLabel(unitKey);

    // % of Total mode: precompute totals by column index from the total row.
    var showAsPct = (tableValueMode === 'pct') && DATA.total_row;
    var totalRowData = null;
    if (showAsPct) {
        for (var r = 0; r < rows.length; r++) {
            if (rows[r].label === DATA.total_row) { totalRowData = rows[r]; break; }
        }
        if (!totalRowData) showAsPct = false;  // safety
    }

    var thead = document.getElementById('tableHead');
    var headerLabel = showAsPct ? '%' : unitLabel;
    var hHtml = '<tr><th>'+headerLabel+'<span class="unit-label">('+period+')</span></th>';
    for (var i=0;i<vis.length;i++) {
        hHtml += '<th>'+(view.short_columns[vis[i]]||view.columns[vis[i]])+'</th>';
    }
    thead.innerHTML = hHtml+'</tr>';

    var tbody = document.getElementById('tableBody');
    var bHtml = '';
    for (var h=0;h<hierarchy.length;h++) {
        var item=hierarchy[h], rowData=rows[item.index];
        var isExp = expandedMain[item.label]||false;
        var hasCh = item.children&&item.children.length>0;
        var isPct=item.is_pct, isStock=item.is_stock;
        var isTotalRow = (item.label === DATA.total_row);
        var rc = item.type==='standalone'?'standalone-row':'parent-row';
        if (isPct) rc+=' pct-row';
        bHtml+='<tr class="'+rc+'">';
        var lbl='';
        if (hasCh) lbl+='<span class="toggle-arrow'+(isExp?' expanded':'')+'">&#9654;</span> ';
        var dl = isStock?getStockLabel(item.label,unitKey):item.label;
        lbl+=dl.replace(/&/g,'&amp;').replace(/</g,'&lt;');
        bHtml+= hasCh?'<td data-toggle="'+h+'">'+lbl+'</td>':'<td>'+lbl+'</td>';
        for (var i=0;i<vis.length;i++) {
            var ci=vis[i], baseVal=rowData.base[ci];
            var dv;
            if (showAsPct && !isPct && !isStock) {
                if (isTotalRow) {
                    dv = '100%';
                } else {
                    var tot = totalRowData.base[ci];
                    if (!tot || tot === 0) dv = '';
                    else dv = (baseVal / tot * 100).toFixed(1) + '%';
                }
            } else if (isPct) {
                dv = formatNum(baseVal, true);
            } else {
                dv = formatNum(computeDisplayValue(baseVal, unitKey, isStock, false, days[ci]), false);
            }
            bHtml+='<td>'+dv+'</td>';
        }
        bHtml+='</tr>';
        if (hasCh) {
            for (var c=0;c<item.children.length;c++) {
                var ch=item.children[c], cd=rows[ch.index];
                var hid=!isExp?' hidden':'';
                bHtml+='<tr class="child-row'+hid+'">';
                bHtml+='<td>'+ch.label.replace(/&/g,'&amp;').replace(/</g,'&lt;')+'</td>';
                for (var i=0;i<vis.length;i++) {
                    var ci=vis[i];
                    var cdv;
                    if (showAsPct) {
                        var tot = totalRowData.base[ci];
                        if (!tot || tot === 0) cdv = '';
                        else cdv = (cd.base[ci] / tot * 100).toFixed(1) + '%';
                    } else {
                        cdv = formatNum(computeDisplayValue(cd.base[ci], unitKey, false, false, days[ci]), false);
                    }
                    bHtml+='<td>'+cdv+'</td>';
                }
                bHtml+='</tr>';
            }
        }
    }
    tbody.innerHTML = bHtml;
}

/* === GROWTH TABLE === */
function updateGrowthTable() {
    var period = document.getElementById('periodSelector').value;
    var view = DATA.views[period];
    if (!view) return;

    var vis = getVisibleIndices();
    var days = view.days;
    var rows = view.rows;
    var meta = view.col_meta;
    var hierarchy = DATA.hierarchy;
    var gt = document.getElementById('growthTypeSelector').value;
    growthType = gt;

    var thead = document.getElementById('growthHead');
    var hHtml = '<tr><th>Change</th>';
    for (var i=0;i<vis.length;i++) {
        hHtml+='<th>'+(view.short_columns[vis[i]]||view.columns[vis[i]])+'</th>';
    }
    thead.innerHTML = hHtml+'</tr>';

    var tbody = document.getElementById('growthBody');
    var bHtml = '';
    for (var h=0;h<hierarchy.length;h++) {
        var item=hierarchy[h], rowData=rows[item.index];
        var isExp=expandedGrowth[item.label]||false;
        var hasCh=item.children&&item.children.length>0;
        var isPct=item.is_pct, isStock=item.is_stock;
        var rc=isPct?'pct-row':(item.type==='standalone'?'standalone-row':'parent-row');
        bHtml+='<tr class="'+rc+'">';
        var dl='';
        if (hasCh) dl+='<span class="toggle-arrow'+(isExp?' expanded':'')+'">&#9654;</span> ';
        dl+=item.label.replace(/&/g,'&amp;').replace(/</g,'&lt;');
        bHtml+=hasCh?'<td data-toggle-g="'+h+'">'+dl+'</td>':'<td>'+dl+'</td>';
        for (var i=0;i<vis.length;i++) {
            var ci=vis[i];
            var gv = computeGrowthCell(rowData.base,days,meta,ci,period,gt,isStock||isPct);
            if (gv===null) { bHtml+='<td></td>'; continue; }
            var cls = gv>0.0001?'g-pos':(gv<-0.0001?'g-neg':'');
            bHtml+='<td'+(cls?' class="'+cls+'"':'')+'>'+formatGrowth(gv,isPct)+'</td>';
        }
        bHtml+='</tr>';
        if (hasCh) {
            for (var c=0;c<item.children.length;c++) {
                var ch=item.children[c], cd=rows[ch.index];
                var hid=!isExp?' hidden':'';
                bHtml+='<tr class="child-row'+hid+'">';
                bHtml+='<td>'+ch.label.replace(/&/g,'&amp;').replace(/</g,'&lt;')+'</td>';
                for (var i=0;i<vis.length;i++) {
                    var ci=vis[i];
                    var gv=computeGrowthCell(cd.base,days,meta,ci,period,gt,false);
                    if (gv===null) { bHtml+='<td></td>'; continue; }
                    var cls=gv>0.0001?'g-pos':(gv<-0.0001?'g-neg':'');
                    bHtml+='<td'+(cls?' class="'+cls+'"':'')+'>'+formatGrowth(gv,false)+'</td>';
                }
                bHtml+='</tr>';
            }
        }
    }
    tbody.innerHTML = bHtml;
}

/* ============================================================
   CHARTS (seasonality + buildup) - power dataset only
   ============================================================ */
var seasonalityChart = null;
var buildupChart = null;
var multiSelectState = {};

/* ---- Helpers ---- */
function currentGY() {
    var now = new Date();
    var m = now.getMonth() + 1;
    return m >= 10 ? now.getFullYear() : now.getFullYear() - 1;
}

function gyMonthLabels() { return ['Oct','Nov','Dec','Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep']; }

function gyMonthToActual(gy, gyMonthIndex) {
    if (gyMonthIndex < 3) return { year: gy, month: 10 + gyMonthIndex };
    return { year: gy + 1, month: gyMonthIndex - 2 };
}

/* Range/seasonality cycle basis: gas-year (Oct->Sep) by default, calendar
   (Jan->Dec) when the period selector is 'Annual CY'. Shared by every range
   chart (gas, power, Global LNG, Supply Outlook). */
function isCalCycle(periodVal) { return periodVal === 'Annual CY'; }
function cycleLabels(cal) { return cal ? ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'] : gyMonthLabels(); }
function cycleToActual(cal, base, i) { return cal ? { year: base, month: i + 1 } : gyMonthToActual(base, i); }
function cycleBase(cal) { return cal ? (new Date()).getFullYear() : currentGY(); }
function cycleSeriesLabel(cal, base, suffix) {
    return cal ? (String(base) + (suffix || ''))
               : ('GY ' + String(base).slice(-2) + '/' + String(base + 1).slice(-2) + (suffix || ''));
}

function findMonthlyIdx(year, month) {
    var meta = DATA.views.Monthly.col_meta;
    for (var i = 0; i < meta.length; i++) {
        if (meta[i].year === year && meta[i].month === month) return i;
    }
    return -1;
}

function getRowByLabel(label) {
    var rows = DATA.views.Monthly.rows;
    for (var i = 0; i < rows.length; i++) {
        if (rows[i].label === label) return rows[i];
    }
    return null;
}

function getSourceColor(label) { return DATA.source_colors[label] || '#888'; }

function getIndividualSourceLabels() {
    var labels = [];
    for (var i = 0; i < DATA.hierarchy.length; i++) {
        var item = DATA.hierarchy[i];
        if (item.label !== DATA.total_row) labels.push(item.label);
    }
    return labels;
}

function sumSourcesAtIndex(selectedLabels, monthlyIdx) {
    var sum = 0;
    for (var i = 0; i < selectedLabels.length; i++) {
        var row = getRowByLabel(selectedLabels[i]);
        if (row && monthlyIdx >= 0 && monthlyIdx < row.base.length) sum += row.base[monthlyIdx];
    }
    return sum;
}

function applyUnitConversion(baseValue, unitKey, days) {
    var cfg = DATA.unit_config[unitKey];
    if (!cfg) return baseValue;
    if (cfg.isRate) return (baseValue / days) * cfg.rateFactor;
    return baseValue * cfg.volFactor;
}

/* Color helper: fade a hex/rgba color to a given alpha (0-1). */
function fadeColor(color, alpha) {
    if (!color || typeof color !== 'string') return color;
    if (color.charAt(0) === '#') {
        var hex = color.slice(1);
        if (hex.length === 8) hex = hex.slice(0, 6);
        if (hex.length === 6) {
            var r = parseInt(hex.slice(0,2), 16);
            var g = parseInt(hex.slice(2,4), 16);
            var b = parseInt(hex.slice(4,6), 16);
            return 'rgba(' + r + ',' + g + ',' + b + ',' + alpha + ')';
        }
    }
    if (color.indexOf('rgba(') === 0) {
        return color.replace(/[\d.]+\)$/, alpha + ')');
    }
    if (color.indexOf('rgb(') === 0) {
        return color.replace('rgb(', 'rgba(').replace(')', ',' + alpha + ')');
    }
    return color;
}

/* Responsive font scale: charts get bigger labels when the canvas is wider. */
function chartFontScale(width) {
    if (!width || isNaN(width)) return 1.0;
    if (width < 500) return 1.0;
    if (width < 700) return 1.12;
    if (width < 1000) return 1.28;
    if (width < 1400) return 1.5;
    return 1.7;
}

function scaledFont(baseSize, fam) {
    return function(ctx) {
        var w = (ctx && ctx.chart && ctx.chart.width) ? ctx.chart.width : 600;
        return { size: Math.round(baseSize * chartFontScale(w)), family: fam };
    };
}

/* ---- Multi-select widget ---- */
function buildMultiSelect(id, items, totalLabel, defaultSelected, onChange) {
    var container = document.getElementById(id);
    if (!container) return;
    container.innerHTML = '';
    container.classList.add('multi-select');
    multiSelectState[id] = defaultSelected.slice();

    var button = document.createElement('button');
    button.type = 'button';
    button.className = 'ms-button';
    container.appendChild(button);

    var panel = document.createElement('div');
    panel.className = 'ms-panel';
    container.appendChild(panel);

    function addItem(value, isTotal) {
        var lbl = document.createElement('label');
        lbl.className = 'ms-item' + (isTotal ? ' ms-total' : '');
        var cb = document.createElement('input');
        cb.type = 'checkbox';
        cb.value = value;
        var sp = document.createElement('span');
        sp.textContent = value;
        lbl.appendChild(cb);
        lbl.appendChild(sp);
        panel.appendChild(lbl);
        cb.addEventListener('change', function(e) { handleMSChange(id, totalLabel, onChange, e.target.value); });
    }

    for (var i = 0; i < items.length; i++) addItem(items[i], false);
    if (totalLabel) addItem(totalLabel, true);

    button.addEventListener('click', function(e) {
        e.stopPropagation();
        container.classList.toggle('open');
    });

    applyMSState(id, totalLabel);
    updateMSButtonLabel(id, totalLabel);
}

function handleMSChange(id, totalLabel, onChange, changedValue) {
    var container = document.getElementById(id);
    var inputs = container.querySelectorAll('input[type="checkbox"]');

    // Mutex: if the just-clicked checkbox is Total and was just turned on, clear individuals;
    // if it's an individual and was just turned on, clear Total.
    var changedIsTotal = (changedValue === totalLabel);
    var changedNowOn = false;
    for (var i = 0; i < inputs.length; i++) {
        if (inputs[i].value === changedValue) { changedNowOn = inputs[i].checked; break; }
    }
    if (changedNowOn) {
        if (changedIsTotal) {
            for (var i = 0; i < inputs.length; i++) if (inputs[i].value !== totalLabel) inputs[i].checked = false;
        } else {
            for (var i = 0; i < inputs.length; i++) if (inputs[i].value === totalLabel) inputs[i].checked = false;
        }
    }

    // Collect resulting state
    var state = [];
    for (var i = 0; i < inputs.length; i++) {
        if (inputs[i].checked) state.push(inputs[i].value);
    }
    // Don't allow zero selection — fall back to Total
    if (state.length === 0 && totalLabel) {
        for (var i = 0; i < inputs.length; i++) {
            if (inputs[i].value === totalLabel) { inputs[i].checked = true; state = [totalLabel]; break; }
        }
    }
    multiSelectState[id] = state;

    updateMSButtonLabel(id, totalLabel);
    if (onChange) onChange(state);
}

function applyMSState(id, totalLabel) {
    var container = document.getElementById(id);
    var inputs = container.querySelectorAll('input[type="checkbox"]');
    var state = multiSelectState[id] || [];
    for (var i = 0; i < inputs.length; i++) {
        inputs[i].checked = state.indexOf(inputs[i].value) >= 0;
    }
}

function updateMSButtonLabel(id, totalLabel) {
    var container = document.getElementById(id);
    var btn = container.querySelector('.ms-button');
    var state = multiSelectState[id] || [];
    var txt = 'None';
    if (state.length === 1) txt = state[0];
    else if (state.length === 2) txt = state.join(' + ');
    else if (state.length > 2) txt = state[0] + ' + ' + (state.length - 1) + ' more';
    btn.textContent = txt;
    btn.title = state.join(', ');
}

function getMultiSelectState(id) { return multiSelectState[id] || []; }

// Outside-click closes any open multi-select panel
document.addEventListener('click', function(e) {
    var open = document.querySelectorAll('.multi-select.open');
    for (var i = 0; i < open.length; i++) {
        if (!open[i].contains(e.target)) open[i].classList.remove('open');
    }
});

/* ---- Seasonality chart ---- */
function updateSeasonalityChart() {
    if (!DATA || !DATA.charts_enabled) return;
    var selected = getMultiSelectState('msSeasonality');
    if (selected.length === 0) selected = [DATA.total_row];

    var unitKey = document.getElementById('unitSelector').value;
    var lookback = parseInt(document.getElementById('seasonalityLookback').value) || 5;
    var cal = isCalCycle(document.getElementById('periodSelector').value);
    var cgy = cycleBase(cal);
    var prevGY = cgy - 1, nextGY = cgy + 1;
    var labels = cycleLabels(cal);

    function gyValues(gy) {
        var out = [];
        for (var i = 0; i < 12; i++) {
            var amt = cycleToActual(cal, gy, i);
            var idx = findMonthlyIdx(amt.year, amt.month);
            if (idx < 0) { out.push(null); continue; }
            var b = sumSourcesAtIndex(selected, idx);
            var d = DATA.views.Monthly.days[idx];
            out.push(applyUnitConversion(b, unitKey, d));
        }
        return out;
    }
    var prevVals = gyValues(prevGY);
    var currVals = gyValues(cgy);
    var nextVals = gyValues(nextGY);

    // Lookback window: `lookback` years prior to current
    var lbStart = cgy - lookback, lbEnd = cgy - 1;
    var avg = [], mn = [], mx = [];
    for (var mo = 0; mo < 12; mo++) {
        var samples = [];
        for (var gy = lbStart; gy <= lbEnd; gy++) {
            var amt = cycleToActual(cal, gy, mo);
            var idx = findMonthlyIdx(amt.year, amt.month);
            if (idx < 0) continue;
            var b = sumSourcesAtIndex(selected, idx);
            var d = DATA.views.Monthly.days[idx];
            samples.push(applyUnitConversion(b, unitKey, d));
        }
        if (samples.length === 0) { avg.push(null); mn.push(null); mx.push(null); continue; }
        var s = 0; for (var k = 0; k < samples.length; k++) s += samples[k];
        avg.push(s / samples.length);
        mn.push(Math.min.apply(null, samples));
        mx.push(Math.max.apply(null, samples));
    }

    var titleSrc = selected.join(' + ');
    document.getElementById('seasonalityTitle').textContent =
        'Seasonality (' + (cal?'calendar':'gas-year') + ' cycle) — ' + titleSrc + ' (' + getHeaderUnitLabel(unitKey) + efficiencySuffix() + ')';

    // Order in array drives legend order (after the max filter).
    // We want: range (grey block), average, prev, current, next.
    // The max line must come BEFORE the range line in the dataset array so the
    // range's `fill: '-1'` correctly fills back to it - but we filter it out
    // of the legend.
    var prevLabel = cycleSeriesLabel(cal, prevGY);
    var curLabel  = cycleSeriesLabel(cal, cgy, ' (current)');
    var nextLabel = cycleSeriesLabel(cal, nextGY, ' (forecast)');
    var datasets = [
        // Invisible max line - exists only to anchor the range fill.
        { label: lookback + 'y max',     data: mx,        borderColor: 'rgba(0,0,0,0)', backgroundColor: 'rgba(0,0,0,0)', pointRadius: 0, fill: false, order: 20 },
        // Range = filled grey block, no line. Shown in legend as a colored box.
        { label: lookback + 'y range',   data: mn,        borderColor: 'rgba(0,0,0,0)', backgroundColor: 'rgba(147,149,162,0.28)', pointRadius: 0, fill: '-1', order: 19 },
        // Average = solid dark blue.
        { label: lookback + 'y average', data: avg,       borderColor: '#272962',       backgroundColor: 'rgba(0,0,0,0)', borderWidth: 1.8, pointRadius: 0, fill: false, tension: 0.25, order: 4 },
        // Previous GY = solid dark green.
        { label: prevLabel,              data: prevVals,  borderColor: '#0C5B19',       backgroundColor: 'rgba(0,0,0,0)', borderWidth: 1.8, pointRadius: 2, fill: false, tension: 0.25, order: 3 },
        // Current GY = solid red (most prominent).
        { label: curLabel,               data: currVals,  borderColor: '#C00000',       backgroundColor: 'rgba(0,0,0,0)', borderWidth: 2.6, pointRadius: 3, fill: false, tension: 0.25, order: 1 },
        // Next GY = dashed light green (forecast).
        { label: nextLabel,              data: nextVals,  borderColor: '#539648',       backgroundColor: 'rgba(0,0,0,0)', borderWidth: 1.8, borderDash: [6,4], pointRadius: 2, fill: false, tension: 0.25, order: 2 },
    ];

    if (seasonalityChart) {
        seasonalityChart.data.labels = labels;
        seasonalityChart.data.datasets = datasets;
        seasonalityChart.options.scales.y.title.text = getHeaderUnitLabel(unitKey);
        seasonalityChart.update();
    } else {
        var canvas = document.getElementById('seasonalityCanvas');
        if (!canvas) return;
        seasonalityChart = new Chart(canvas, {
            type: 'line',
            data: { labels: labels, datasets: datasets },
            options: chartOptionsLine(unitKey)
        });
    }

    // Render HTML legend - filter out the invisible "max" anchor dataset.
    renderHtmlLegend(seasonalityChart, 'seasonalityLegend', {
        filter: function(label) { return !(label || '').endsWith(' max'); }
    });
}

function chartOptionsLine(unitKey) {
    var fam = "'Gotham Book','Segoe UI',Calibri,sans-serif";
    return {
        responsive: true, maintainAspectRatio: false,
        interaction: { mode: 'index', intersect: false },
        plugins: {
            // Built-in legend disabled - we render our own HTML legend below
            // the canvas so we can fade hidden items without Chart.js drawing
            // the hardcoded strike-through.
            legend: { display: false },
            tooltip: {
                mode: 'index', intersect: false,
                filter: function(ctx) { return !(ctx.dataset.label || '').endsWith(' max'); },
                titleFont: scaledFont(11, fam),
                bodyFont: scaledFont(10.5, fam),
                callbacks: {
                    label: function(ctx) { return ctx.dataset.label + ': ' + (ctx.parsed.y === null || ctx.parsed.y === undefined ? '-' : ctx.parsed.y.toFixed(1)); }
                }
            }
        },
        scales: {
            x: { grid: { display: false }, ticks: { font: scaledFont(10, fam), color: '#272962' } },
            y: { title: { display: true, text: getHeaderUnitLabel(unitKey), font: scaledFont(10.5, fam), color: '#272962' },
                 grid: { color: '#E0E0E8' },
                 ticks: { font: scaledFont(10, fam), color: '#272962' } }
        }
    };
}

/* ---- Buildup chart ---- */
function getBuildupSourcesEffective(rawSelected) {
    if (rawSelected.length === 1 && rawSelected[0] === DATA.total_row) return getIndividualSourceLabels();
    return rawSelected;
}

function updateBuildupChart() {
    if (!DATA || !DATA.charts_enabled) return;
    var rawSelected = getMultiSelectState('msBuildup');
    if (rawSelected.length === 0) rawSelected = [DATA.total_row];
    var sources = getBuildupSourcesEffective(rawSelected);
    var unitKey = document.getElementById('unitSelector').value;
    var agg = document.getElementById('buildupAgg').value;
    var viewKey = agg;
    var view = DATA.views[viewKey];
    if (!view) return;

    var fromIdx = parseInt(document.getElementById('buildupFrom').value);
    var toIdx = parseInt(document.getElementById('buildupTo').value);
    if (isNaN(fromIdx) || isNaN(toIdx)) { fromIdx = 0; toIdx = view.col_meta.length - 1; }
    if (fromIdx > toIdx) { var t = fromIdx; fromIdx = toIdx; toIdx = t; }

    var labels = [];
    for (var i = fromIdx; i <= toIdx; i++) labels.push(view.short_columns[i] || view.columns[i]);

    var datasets = [];
    for (var s = 0; s < sources.length; s++) {
        var label = sources[s];
        var row = null;
        for (var r = 0; r < view.rows.length; r++) { if (view.rows[r].label === label) { row = view.rows[r]; break; } }
        if (!row) continue;
        var data = [];
        for (var i = fromIdx; i <= toIdx; i++) {
            // Clamp negatives to zero: small forecast quirks (e.g. negative
            // values in the Ember "Other Renewables" series) break stacked-area
            // rendering. Generation should not be negative for chart purposes.
            var v = applyUnitConversion(row.base[i], unitKey, view.days[i]);
            data.push(v < 0 ? 0 : v);
        }
        var color = getSourceColor(label);
        datasets.push({
            label: label,
            data: data,
            backgroundColor: color + 'CC',
            borderColor: color,
            borderWidth: 1,
            // For stacked area (line) charts, fill back to the previous
            // dataset's line so each layer renders ONLY as its incremental
            // band, not from 0 up to its cumulative top. Without this, fading
            // others would still leave the hovered dataset's fill spanning
            // 0 -> cumulative top.
            fill: s === 0 ? 'origin' : '-1',
            stack: 'gen',
            tension: 0.2,
            pointRadius: 0,
        });
    }

    var modeLabel = (sources.length === getIndividualSourceLabels().length && rawSelected.length === 1 && rawSelected[0] === DATA.total_row)
        ? 'Total (stacked breakdown)'
        : sources.join(' + ');
    document.getElementById('buildupTitle').textContent =
        'Generation — ' + modeLabel + ' (' + getHeaderUnitLabel(unitKey) + efficiencySuffix() + ')';

    var chartType = (agg === 'Monthly') ? 'line' : 'bar';

    if (buildupChart) { buildupChart.destroy(); buildupChart = null; }
    var canvas = document.getElementById('buildupCanvas');
    if (!canvas) return;
    buildupChart = new Chart(canvas, {
        type: chartType,
        data: { labels: labels, datasets: datasets },
        options: chartOptionsStacked(unitKey)
    });
    buildupStackHover = true;
    initBuildupHover(canvas);
    renderHtmlLegend(buildupChart, 'buildupLegend', {});
}

function chartOptionsStacked(unitKey) {
    var fam = "'Gotham Book','Segoe UI',Calibri,sans-serif";
    return {
        responsive: true, maintainAspectRatio: false,
        interaction: { mode: 'index', intersect: false },
        plugins: {
            // Custom HTML legend - see above for rationale.
            legend: { display: false },
            tooltip: {
                mode: 'index', intersect: false,
                titleFont: scaledFont(11, fam),
                bodyFont: scaledFont(10.5, fam),
                footerFont: scaledFont(10.5, fam),
                callbacks: {
                    label: function(ctx) { return ctx.dataset.label + ': ' + (ctx.parsed.y === null || ctx.parsed.y === undefined ? '-' : ctx.parsed.y.toFixed(1)); },
                    footer: function(items) {
                        var total = 0;
                        for (var i = 0; i < items.length; i++) {
                            var v = items[i].parsed.y;
                            if (typeof v === 'number' && !isNaN(v)) total += v;
                        }
                        return 'Total: ' + total.toFixed(1);
                    }
                }
            }
        },
        scales: {
            x: { stacked: true, grid: { display: false }, ticks: { font: scaledFont(10, fam), color: '#272962', maxRotation: 0, autoSkip: true, autoSkipPadding: 12 } },
            y: {
                stacked: true,
                beginAtZero: true,  // critical: keep all unit views on the same baseline
                title: { display: true, text: getHeaderUnitLabel(unitKey), font: scaledFont(10.5, fam), color: '#272962' },
                grid: { color: '#E0E0E8' },
                ticks: { font: scaledFont(10, fam), color: '#272962' }
            }
        }
    };
}

/* Custom HTML legend renderer.
   Chart.js's native legend has a hardcoded strike-through on hidden items
   that can't be suppressed via standard config. We render our own legend
   below the canvas so we can fade hidden items via CSS opacity. */
function renderHtmlLegend(chart, containerId, opts) {
    opts = opts || {};
    var container = document.getElementById(containerId);
    if (!container) return;
    container.innerHTML = '';

    var datasets = chart.data.datasets;
    for (var i = 0; i < datasets.length; i++) {
        var ds = datasets[i];
        var label = ds.label || '';
        if (opts.filter && !opts.filter(label)) continue;

        // Determine the marker color. For filled areas (range, stacked sources)
        // use the dataset's backgroundColor; for lines use borderColor.
        var bg = ds.backgroundColor;
        var border = ds.borderColor;
        var markerFill, markerBorder;
        // If a transparent background, use border color (line series).
        var bgIsTransparent = (bg === 'rgba(0,0,0,0)') || (typeof bg === 'string' && bg.indexOf('rgba(0,0,0,0)') === 0);
        if (bgIsTransparent || !bg) {
            markerFill = border || '#888';
            markerBorder = border || '#888';
        } else {
            markerFill = bg;
            markerBorder = border || bg;
        }

        var hidden = chart.getDatasetMeta(i).hidden === true;

        var item = document.createElement('div');
        item.className = 'cl-item' + (hidden ? ' cl-hidden' : '');
        item.setAttribute('data-idx', i);

        // Marker style depends on the dataset shape: filled box for areas/bands,
        // line-stripe for line series, dashed line for forecasts.
        var marker = document.createElement('span');
        marker.className = 'cl-marker';
        if (bgIsTransparent) {
            marker.classList.add('cl-marker-line');
            marker.style.setProperty('--ml', markerFill);
            if (ds.borderDash && ds.borderDash.length) marker.classList.add('cl-marker-dashed');
        } else {
            marker.classList.add('cl-marker-box');
            marker.style.backgroundColor = markerFill;
            marker.style.borderColor = markerBorder;
        }
        item.appendChild(marker);

        var text = document.createElement('span');
        text.className = 'cl-text';
        text.textContent = label;
        item.appendChild(text);

        item.addEventListener('click', function() {
            var idx = parseInt(this.getAttribute('data-idx'));
            var meta = chart.getDatasetMeta(idx);
            meta.hidden = meta.hidden === true ? false : true;
            chart.update();
            renderHtmlLegend(chart, containerId, opts);
        });

        container.appendChild(item);
    }
}

/* Hover-highlight on the buildup chart canvas.
   Chart.js's built-in hit detection on stacked AREA charts only fires on
   line points - not in the filled regions. So we compute the hovered
   dataset manually by mapping the cursor's Y position into the cumulative
   stack at the cursor's X column. */
var buildupHoverInited = false;
var buildupStackHover = true;  // false in gas line mode (no cumulative-stack hover)
function initBuildupHover(canvas) {
    if (buildupHoverInited) return;
    buildupHoverInited = true;
    canvas.addEventListener('mousemove', onBuildupMousemove);
    canvas.addEventListener('mouseleave', onBuildupMouseleave);
}

function onBuildupMousemove(e) {
    if (!buildupChart) return;
    if (!buildupStackHover) { applyBuildupHover(-1); return; }  // line mode: no stack-walk fade
    var canvas = buildupChart.canvas;
    var rect = canvas.getBoundingClientRect();
    var cx = e.clientX - rect.left;
    var cy = e.clientY - rect.top;

    var chartArea = buildupChart.chartArea;
    if (!chartArea || cx < chartArea.left || cx > chartArea.right || cy < chartArea.top || cy > chartArea.bottom) {
        applyBuildupHover(-1);
        return;
    }

    var xScale = buildupChart.scales.x;
    var yScale = buildupChart.scales.y;
    var labels = buildupChart.data.labels;
    var datasets = buildupChart.data.datasets;

    // Find nearest x category
    var dataIdx = 0, minDist = Infinity;
    for (var i = 0; i < labels.length; i++) {
        var px = xScale.getPixelForValue(i);
        var d = Math.abs(px - cx);
        if (d < minDist) { minDist = d; dataIdx = i; }
    }

    // Walk the cumulative stack at this x to find which dataset the cursor
    // is sitting within (in data-value space, not pixels).
    var cursorYVal = yScale.getValueForPixel(cy);
    var cum = 0;
    var hoveredIdx = -1;
    for (var i = 0; i < datasets.length; i++) {
        if (buildupChart.getDatasetMeta(i).hidden) continue;
        var v = datasets[i].data[dataIdx] || 0;
        var top = cum + v;
        if (cursorYVal >= cum && cursorYVal <= top) {
            hoveredIdx = i;
            break;
        }
        cum = top;
    }
    applyBuildupHover(hoveredIdx);
}

function onBuildupMouseleave() { applyBuildupHover(-1); }

function applyBuildupHover(hoveredIdx) {
    if (!buildupChart) return;
    var changed = false;
    var datasets = buildupChart.data.datasets;
    for (var i = 0; i < datasets.length; i++) {
        var ds = datasets[i];
        if (ds._origBg === undefined) ds._origBg = ds.backgroundColor;
        if (ds._origBorder === undefined) ds._origBorder = ds.borderColor;
        var targetBg, targetBorder;
        if (hoveredIdx === -1 || hoveredIdx === i) {
            targetBg = ds._origBg;
            targetBorder = ds._origBorder;
        } else {
            targetBg = fadeColor(ds._origBg, 0.10);
            targetBorder = fadeColor(ds._origBorder, 0.20);
        }
        if (ds.backgroundColor !== targetBg) { ds.backgroundColor = targetBg; changed = true; }
        if (ds.borderColor !== targetBorder) { ds.borderColor = targetBorder; changed = true; }
    }
    if (changed) buildupChart.update('none');
}

/* ---- Chart-control rebuilders ---- */
function rebuildChartControls() {
    if (!DATA.charts_enabled) return;
    var lab = document.getElementById('rangeSourceLabel');
    if (lab) lab.textContent = 'Sources';
    var individuals = getIndividualSourceLabels();
    var totalLabel = DATA.total_row;
    buildMultiSelect('msSeasonality', individuals, totalLabel, [totalLabel], function() { updateSeasonalityChart(); });
    buildMultiSelect('msBuildup',     individuals, totalLabel, [totalLabel], function() { updateBuildupChart(); });
    populateLookbackOptions();
    populateBuildupRangeSelectors();
}

function populateLookbackOptions() {
    var sel = document.getElementById('seasonalityLookback');
    if (!sel) return;
    var cgy = currentGY();
    var firstYear = DATA.views.Monthly.col_meta[0].year;
    var maxLookback = Math.max(3, cgy - firstYear);
    sel.innerHTML = '';
    for (var n = 3; n <= maxLookback; n++) {
        var s = (n === 5) ? ' selected' : '';
        sel.innerHTML += '<option value="' + n + '"' + s + '>' + n + ' years</option>';
    }
}

function populateBuildupRangeSelectors() {
    var aggEl = document.getElementById('buildupAgg');
    if (!aggEl) return;
    var agg = aggEl.value;
    var view = DATA.views[agg];
    if (!view) return;
    var fs = document.getElementById('buildupFrom');
    var ts = document.getElementById('buildupTo');
    fs.innerHTML = ''; ts.innerHTML = '';
    for (var i = 0; i < view.col_meta.length; i++) {
        var lbl = view.col_meta[i].label;
        fs.innerHTML += '<option value="' + i + '">' + lbl + '</option>';
        ts.innerHTML += '<option value="' + i + '">' + lbl + '</option>';
    }
    // Default: clip to admin display range
    var defaultFrom = 0, defaultTo = view.col_meta.length - 1;
    for (var i = 0; i < view.col_meta.length; i++) {
        if (view.col_meta[i].year >= DATA.selectable_start && view.col_meta[i].year <= DATA.selectable_end) { defaultFrom = i; break; }
    }
    for (var i = view.col_meta.length - 1; i >= 0; i--) {
        if (view.col_meta[i].year >= DATA.selectable_start && view.col_meta[i].year <= DATA.selectable_end) { defaultTo = i; break; }
    }
    fs.value = defaultFrom;
    ts.value = defaultTo;
}

function onBuildupAggChange() {
    populateBuildupRangeSelectors();
    updateBuildupChart();
}

function updateCharts() {
    if (!DATA.charts_enabled) return;
    updateSeasonalityChart();
    updateBuildupChart();
}

/* Dispatcher: power uses the flat seasonality chart, gas uses the hierarchical
   range chart. Both render into the same left-slot canvas. */
function updateRangeChart() {
    if (!DATA) return;
    if (DATA.chart_groups) updateGasRangeChart();
    else if (DATA.charts_enabled) updateSeasonalityChart();
}

/* ============================================================
   GAS RANGE CHART (left slot) — hierarchical source selector.
   Differs from the power seasonality chart in three ways:
     - sources are grouped (Imports / Consumption expand to members),
     - consumption & exports are shown as positive magnitudes (abs),
     - Storage percentage is a special %-axis series, mutually exclusive
       with everything else, and the default selection.
   The summed selection is drawn with the same prev/current/next gas-year
   lines + n-year average + min/max band styling as the power chart.
   ============================================================ */
var gasRangeSel = [];  // selected base-row labels (atoms)

function gasGroups() { return DATA.chart_groups || []; }
function gasEntryForKey(k) { var g=gasGroups(); for (var i=0;i<g.length;i++) if (g[i].key===k) return g[i]; return null; }
function gasStorageRow() { var e=null,g=gasGroups(); for (var i=0;i<g.length;i++) if (g[i].kind==='pct') e=g[i]; return e?e.rows[0]:null; }
// Map a base-row label -> its owning chart_groups entry.
function gasEntryForRow(rowLabel) {
    var g=gasGroups();
    for (var i=0;i<g.length;i++) for (var j=0;j<g[i].rows.length;j++) if (g[i].rows[j]===rowLabel) return g[i];
    return null;
}
function gasRowsSelectedIn(entry) {
    var out=[]; for (var i=0;i<entry.rows.length;i++) if (gasRangeSel.indexOf(entry.rows[i])>=0) out.push(entry.rows[i]); return out;
}
function gasEntryFull(entry)    { return gasRowsSelectedIn(entry).length === entry.rows.length; }
function gasEntryPartial(entry) { var n=gasRowsSelectedIn(entry).length; return n>0 && n<entry.rows.length; }
function gasAnyPartial() { var g=gasGroups(); for (var i=0;i<g.length;i++) if (g[i].kind==='group' && gasEntryPartial(g[i])) return g[i]; return null; }
function gasStorageSelected() { var sr=gasStorageRow(); return sr && gasRangeSel.indexOf(sr)>=0; }

// --- selection state machine (click-time auto-resolution) ---
function gasClickStorage() { var sr=gasStorageRow(); gasRangeSel = sr ? [sr] : []; }

function gasClickTotal(entry) {
    // Leaf or group header: an aggregate-combinable total. Exits any drill.
    var sr=gasStorageRow();
    var keep=[];
    for (var i=0;i<gasRangeSel.length;i++) {
        var lab=gasRangeSel[i];
        if (lab===sr) continue;                              // drop storage
        var owner=gasEntryForRow(lab);
        if (owner && owner.kind==='group' && gasEntryPartial(owner)) continue; // drop drilled (partial) groups
        keep.push(lab);
    }
    gasRangeSel = keep;
    // Is this total currently fully selected (against the cleaned-up selection)?
    var sel=0; for (var i=0;i<entry.rows.length;i++) if (gasRangeSel.indexOf(entry.rows[i])>=0) sel++;
    if (sel === entry.rows.length) {
        // currently full -> turn the whole total off
        gasRangeSel = gasRangeSel.filter(function(l){ return entry.rows.indexOf(l)<0; });
    } else {
        // turn the whole total on (add any missing members)
        for (var i=0;i<entry.rows.length;i++) if (gasRangeSel.indexOf(entry.rows[i])<0) gasRangeSel.push(entry.rows[i]);
    }
}

function gasClickChild(entry, rowLabel) {
    // Drilling into one group: selection becomes only this group's members.
    var start = gasRowsSelectedIn(entry);
    var pos = start.indexOf(rowLabel);
    if (pos>=0) start.splice(pos,1); else start.push(rowLabel);
    gasRangeSel = start;  // exclusive to this group
}

function gasNormalizeSelection() {
    if (gasRangeSel.length===0) { var sr=gasStorageRow(); gasRangeSel = sr?[sr]:[]; }
}

// Resolve current selection to the rows to sum + pct flag + a readable label.
function gasResolveSelection() {
    if (gasStorageSelected()) {
        return { rows:[{label:gasStorageRow(), abs:false}], isPct:true, label:'Storage %' };
    }
    var rows=[], parts=[], g=gasGroups();
    for (var i=0;i<g.length;i++) {
        var e=g[i];
        if (e.kind==='pct') continue;
        var selRows = gasRowsSelectedIn(e);
        if (selRows.length===0) continue;
        for (var j=0;j<selRows.length;j++) rows.push({label:selRows[j], abs:e.abs});
        if (e.kind==='group') {
            if (gasEntryFull(e)) parts.push(e.label);
            else parts.push(e.label + ' (' + selRows.join(', ') + ')');
        } else {
            parts.push(e.label);
        }
    }
    return { rows:rows, isPct:false, label: parts.join(' + ') || '—' };
}

function gasSumAtIndex(rows, idx) {
    var sum=0;
    for (var i=0;i<rows.length;i++) {
        var row=getRowByLabel(rows[i].label);
        if (row && idx>=0 && idx<row.base.length) {
            var v=row.base[idx];
            sum += rows[i].abs ? Math.abs(v) : v;
        }
    }
    return sum;
}

// --- selector widget ---
function buildGasRangeSelector() {
    var container = document.getElementById('msSeasonality');
    if (!container) return;
    container.innerHTML='';
    container.classList.add('multi-select');

    var button = document.createElement('button');
    button.type='button'; button.className='ms-button';
    container.appendChild(button);
    var panel = document.createElement('div');
    panel.className='ms-panel';
    container.appendChild(panel);

    var g = gasGroups();
    for (var i=0;i<g.length;i++) {
        var e = g[i];
        if (e.kind==='pct') {
            panel.appendChild(gasMakeItem(e.label, '', function(){ gasClickStorage(); }));
            var div=document.createElement('div'); div.className='ms-divider'; panel.appendChild(div);
        } else if (e.kind==='leaf') {
            (function(entry){ panel.appendChild(gasMakeItem(entry.label, '', function(){ gasClickTotal(entry); })); })(e);
        } else { // group
            (function(entry){
                panel.appendChild(gasMakeItem(entry.label, 'ms-group-header', function(){ gasClickTotal(entry); }, entry.key));
                for (var j=0;j<entry.rows.length;j++) {
                    (function(rowLabel){
                        panel.appendChild(gasMakeItem(rowLabel, 'ms-child', function(){ gasClickChild(entry, rowLabel); }, entry.key+'::'+rowLabel));
                    })(entry.rows[j]);
                }
            })(e);
        }
    }

    button.addEventListener('click', function(ev){ ev.stopPropagation(); container.classList.toggle('open'); });
    gasRefreshSelector();
}

function gasMakeItem(text, extraCls, onClick, cbKey) {
    var lbl=document.createElement('label');
    lbl.className='ms-item' + (extraCls ? ' '+extraCls : '');
    var cb=document.createElement('input');
    cb.type='checkbox';
    cb.setAttribute('data-key', cbKey || text);
    var sp=document.createElement('span');
    sp.textContent=text;
    lbl.appendChild(cb); lbl.appendChild(sp);
    cb.addEventListener('change', function(){
        onClick();
        gasNormalizeSelection();
        gasRefreshSelector();
        gasRangeButtonLabel();
        updateGasRangeChart();
    });
    return lbl;
}

// Sync every checkbox (checked / indeterminate) to gasRangeSel.
function gasRefreshSelector() {
    var g=gasGroups();
    for (var i=0;i<g.length;i++) {
        var e=g[i];
        if (e.kind==='pct') {
            gasSetCb(e.label, gasStorageSelected(), false);
        } else if (e.kind==='leaf') {
            gasSetCb(e.label, gasRangeSel.indexOf(e.rows[0])>=0, false);
        } else {
            gasSetCb(e.key, gasEntryFull(e), gasEntryPartial(e));
            for (var j=0;j<e.rows.length;j++) gasSetCb(e.key+'::'+e.rows[j], gasRangeSel.indexOf(e.rows[j])>=0, false);
        }
    }
    gasRangeButtonLabel();
}
function gasSetCb(key, checked, indeterminate) {
    var cb=document.querySelector('#msSeasonality input[data-key="'+CSS.escape(key)+'"]');
    if (!cb) return;
    cb.checked=!!checked;
    cb.indeterminate=!!indeterminate;
}

function gasRangeButtonLabel() {
    var container=document.getElementById('msSeasonality');
    if (!container) return;
    var btn=container.querySelector('.ms-button');
    if (!btn) return;
    var res=gasResolveSelection();
    var txt=res.label;
    if (res.isPct) txt='Storage %';
    else {
        var parts=res.label.split(' + ');
        if (parts.length>2) txt=parts[0]+' + '+(parts.length-1)+' more';
    }
    btn.textContent=txt;
    btn.title=res.label;
}

function updateGasRangeChart() {
    if (!DATA || !DATA.chart_groups) return;
    var res=gasResolveSelection();
    var unitKey=document.getElementById('unitSelector').value;
    var lookbackEl=document.getElementById('seasonalityLookback');
    var lookback=lookbackEl ? (parseInt(lookbackEl.value)||5) : 5;
    var cal=isCalCycle(document.getElementById('periodSelector').value);
    var cgy=cycleBase(cal), prevGY=cgy-1, nextGY=cgy+1;
    var labels=cycleLabels(cal);

    function valAt(idx) {
        if (idx<0) return null;
        var b=gasSumAtIndex(res.rows, idx);
        if (res.isPct) return b*100;
        var d=DATA.views.Monthly.days[idx];
        return applyUnitConversion(b, unitKey, d);
    }
    function gyValues(gy) {
        var out=[];
        for (var i=0;i<12;i++) {
            var amt=cycleToActual(cal,gy,i);
            out.push(valAt(findMonthlyIdx(amt.year, amt.month)));
        }
        return out;
    }
    var prevVals=gyValues(prevGY), currVals=gyValues(cgy), nextVals=gyValues(nextGY);

    var lbStart=cgy-lookback, lbEnd=cgy-1;
    var avg=[], mn=[], mx=[];
    for (var mo=0;mo<12;mo++) {
        var samples=[];
        for (var gy=lbStart;gy<=lbEnd;gy++) {
            var amt=cycleToActual(cal,gy,mo);
            var v=valAt(findMonthlyIdx(amt.year, amt.month));
            if (v!==null && v!==undefined) samples.push(v);
        }
        if (samples.length===0) { avg.push(null); mn.push(null); mx.push(null); continue; }
        var s=0; for (var k=0;k<samples.length;k++) s+=samples[k];
        avg.push(s/samples.length); mn.push(Math.min.apply(null,samples)); mx.push(Math.max.apply(null,samples));
    }

    var unitLabel = res.isPct ? '%' : getHeaderUnitLabel(unitKey);
    document.getElementById('seasonalityTitle').textContent = 'Seasonality';

    var prevLabel=cycleSeriesLabel(cal, prevGY);
    var curLabel =cycleSeriesLabel(cal, cgy, ' (current)');
    var nextLabel=cycleSeriesLabel(cal, nextGY, ' (forecast)');
    var datasets=[
        { label: lookback+'y max',     data: mx,       borderColor:'rgba(0,0,0,0)', backgroundColor:'rgba(0,0,0,0)', pointRadius:0, fill:false, order:20 },
        { label: lookback+'y range',   data: mn,       borderColor:'rgba(0,0,0,0)', backgroundColor:'rgba(147,149,162,0.28)', pointRadius:0, fill:'-1', order:19 },
        { label: lookback+'y average', data: avg,      borderColor:'#272962', backgroundColor:'rgba(0,0,0,0)', borderWidth:1.8, pointRadius:0, fill:false, tension:0.25, order:4 },
        { label: prevLabel,            data: prevVals, borderColor:'#0C5B19', backgroundColor:'rgba(0,0,0,0)', borderWidth:1.8, pointRadius:2, fill:false, tension:0.25, order:3 },
        { label: curLabel,             data: currVals, borderColor:'#C00000', backgroundColor:'rgba(0,0,0,0)', borderWidth:2.6, pointRadius:3, fill:false, tension:0.25, order:1 },
        { label: nextLabel,            data: nextVals, borderColor:'#539648', backgroundColor:'rgba(0,0,0,0)', borderWidth:1.8, borderDash:[6,4], pointRadius:2, fill:false, tension:0.25, order:2 },
    ];

    var opts=chartOptionsLine(unitKey);
    opts.scales.y.title.text=unitLabel;
    if (res.isPct) opts.scales.y.ticks.callback = function(v){ return v + '%'; };

    if (seasonalityChart) {
        seasonalityChart.data.labels=labels;
        seasonalityChart.data.datasets=datasets;
        seasonalityChart.options=opts;
        seasonalityChart.update();
    } else {
        var canvas=document.getElementById('seasonalityCanvas');
        if (!canvas) return;
        seasonalityChart=new Chart(canvas, { type:'line', data:{labels:labels, datasets:datasets}, options:opts });
    }
    renderHtmlLegend(seasonalityChart, 'seasonalityLegend', { filter: function(label){ return !(label||'').endsWith(' max'); } });
}

function rebuildGasChartControls() {
    if (!DATA.chart_groups) return;
    var lab=document.getElementById('rangeSourceLabel');
    if (lab) lab.textContent='Series';        // gas: don't call it "Sources"
    var sr=gasStorageRow();
    gasRangeSel = sr ? [sr] : [];             // range chart default = Storage %
    buildGasRangeSelector();
    populateLookbackOptions();
    // Build-up (right) chart: default = stacked bar, all volume series selected.
    gasBuildupType='bar';
    gasBuildupSelectAllVolumes();
    buildGasBuildupSelector();
    var btns=['bar','area','line'];
    for (var i=0;i<btns.length;i++){ var b=document.getElementById('btnGasBuild-'+btns[i]); if (b) b.className=(btns[i]==='bar')?'active':''; }
}

/* ============================================================
   GAS BUILD-UP CHART (right slot) — stacked bar / stacked area / line.
   Series = gas balance items (Domestic Production .. Exports), with Imports
   and Consumption expanding to members; consumption & exports are shown as
   positive magnitudes. Storage % is a line-only, mutually-exclusive %-series.
   Period / range / unit all follow the top table (getVisibleIndices()).
   ============================================================ */
var gasBuildupType = 'bar';   // 'bar' | 'area' | 'line'
var gasBuildupSel = [];        // selected base-row labels (leaves)

// Stackable leaves (everything except Storage %), in display/stack order.
function gasBuildupLeaves() {
    var out=[], g=gasGroups();
    for (var i=0;i<g.length;i++) {
        var e=g[i];
        if (e.kind==='pct') continue;
        for (var j=0;j<e.rows.length;j++)
            out.push({row:e.rows[j], abs:e.abs, isGroup:e.kind==='group'});
    }
    return out;
}
function gasBuildupSelectAllVolumes() { gasBuildupSel = gasBuildupLeaves().map(function(l){return l.row;}); }
function gasToggleInArr(arr,v){ var p=arr.indexOf(v); if(p>=0) arr.splice(p,1); else arr.push(v); }

function setGasBuildupType(t) {
    gasBuildupType=t;
    var btns=['bar','area','line'];
    for (var i=0;i<btns.length;i++){ var b=document.getElementById('btnGasBuild-'+btns[i]); if (b) b.className=(btns[i]===t)?'active':''; }
    // Storage % is line-only; leaving line with it selected reverts to all volumes.
    if (t!=='line') {
        var sr=gasStorageRow();
        if (gasBuildupSel.indexOf(sr)>=0) gasBuildupSelectAllVolumes();
    }
    buildGasBuildupSelector();   // re-render (adds/removes the Storage % option)
    updateGasBuildupChart();
}

function buildGasBuildupSelector() {
    var container=document.getElementById('msGasBuildup');
    if (!container) return;
    container.innerHTML='';
    container.classList.add('multi-select');
    var button=document.createElement('button'); button.type='button'; button.className='ms-button'; container.appendChild(button);
    var panel=document.createElement('div'); panel.className='ms-panel'; container.appendChild(panel);

    // "All" select-all at the very top (selects every series except Storage %).
    panel.appendChild(gasBuildItem('All','ms-group-header','__ALL__','all',null));
    var topDiv=document.createElement('div'); topDiv.className='ms-divider'; panel.appendChild(topDiv);

    var g=gasGroups(), pctEntry=null;
    for (var i=0;i<g.length;i++) {
        var e=g[i];
        if (e.kind==='pct') { pctEntry=e; continue; }   // Storage % rendered at the bottom
        else if (e.kind==='leaf') {
            panel.appendChild(gasBuildItem(e.label,'',e.rows[0],'leaf',null));
        } else {
            (function(entry){
                panel.appendChild(gasBuildItem(entry.label,'ms-group-header',entry.key,'group',entry));
                for (var j=0;j<entry.rows.length;j++)
                    panel.appendChild(gasBuildItem(entry.rows[j],'ms-child',entry.rows[j],'child',entry));
            })(e);
        }
    }
    // Storage % at the very bottom, line mode only (it uses a separate % axis).
    if (gasBuildupType==='line' && pctEntry) {
        var botDiv=document.createElement('div'); botDiv.className='ms-divider'; panel.appendChild(botDiv);
        panel.appendChild(gasBuildItem(pctEntry.label,'',pctEntry.rows[0],'storage',null));
    }
    button.addEventListener('click', function(ev){ ev.stopPropagation(); container.classList.toggle('open'); });
    gasBuildupRefresh();
}

function gasBuildItem(text,extraCls,key,kind,entry) {
    var lbl=document.createElement('label'); lbl.className='ms-item'+(extraCls?' '+extraCls:'');
    var cb=document.createElement('input'); cb.type='checkbox'; cb.setAttribute('data-key',key);
    var sp=document.createElement('span'); sp.textContent=text;
    lbl.appendChild(cb); lbl.appendChild(sp);
    cb.addEventListener('change', function(){ gasBuildupClick(kind,key,entry); gasBuildupRefresh(); updateGasBuildupChart(); });
    return lbl;
}

function gasBuildupClick(kind,key,entry) {
    var sr=gasStorageRow();
    if (kind==='storage') {
        if (gasBuildupSel.length===1 && gasBuildupSel[0]===sr) gasBuildupSelectAllVolumes(); // toggle off -> volumes
        else gasBuildupSel=[sr];                                                             // mutex: storage alone
        return;
    }
    if (kind==='all') {
        var allLeaves=gasBuildupLeaves().map(function(l){ return l.row; });
        var allIn=allLeaves.every(function(r){ return gasBuildupSel.indexOf(r)>=0; });
        gasBuildupSel = allIn ? [] : allLeaves.slice();  // toggle every series (storage cleared either way)
        return;
    }
    gasBuildupSel = gasBuildupSel.filter(function(l){ return l!==sr; });  // any volume click clears storage
    if (kind==='group') {
        var rows=(entry||gasEntryForKey(key)).rows;
        var allIn=rows.every(function(r){ return gasBuildupSel.indexOf(r)>=0; });
        if (allIn) gasBuildupSel=gasBuildupSel.filter(function(r){ return rows.indexOf(r)<0; });
        else rows.forEach(function(r){ if (gasBuildupSel.indexOf(r)<0) gasBuildupSel.push(r); });
    } else {
        gasToggleInArr(gasBuildupSel,key);  // leaf or child
    }
}

function gasBuildupRefresh() {
    var sr=gasStorageRow(), g=gasGroups();
    var allLeaves=gasBuildupLeaves().map(function(l){ return l.row; });
    var nAll=allLeaves.filter(function(r){ return gasBuildupSel.indexOf(r)>=0; }).length;
    gasBuildSetCb('__ALL__', nAll===allLeaves.length, nAll>0 && nAll<allLeaves.length);
    for (var i=0;i<g.length;i++) {
        var e=g[i];
        if (e.kind==='pct')      { gasBuildSetCb(e.rows[0], gasBuildupSel.indexOf(sr)>=0, false); }
        else if (e.kind==='leaf'){ gasBuildSetCb(e.rows[0], gasBuildupSel.indexOf(e.rows[0])>=0, false); }
        else {
            var n=e.rows.filter(function(r){ return gasBuildupSel.indexOf(r)>=0; }).length;
            gasBuildSetCb(e.key, n===e.rows.length, n>0 && n<e.rows.length);
            for (var j=0;j<e.rows.length;j++) gasBuildSetCb(e.rows[j], gasBuildupSel.indexOf(e.rows[j])>=0, false);
        }
    }
    gasBuildupButtonLabel();
}
function gasBuildSetCb(key,checked,indet) {
    var cb=document.querySelector('#msGasBuildup input[data-key="'+CSS.escape(key)+'"]');
    if (!cb) return; cb.checked=!!checked; cb.indeterminate=!!indet;
}
function gasBuildupButtonLabel() {
    var c=document.getElementById('msGasBuildup'); if (!c) return;
    var btn=c.querySelector('.ms-button'); if (!btn) return;
    var sr=gasStorageRow();
    if (gasBuildupSel.length===1 && gasBuildupSel[0]===sr) { btn.textContent='Storage %'; btn.title='Storage percentage'; return; }
    var allLeaves=gasBuildupLeaves().map(function(l){ return l.row; });
    var allIn=allLeaves.every(function(r){ return gasBuildupSel.indexOf(r)>=0; });
    if (allIn) { btn.textContent='All'; btn.title=allLeaves.join(', '); return; }
    var n=gasBuildupSel.length;
    if (n===0) { btn.textContent='None'; btn.title=''; return; }
    var clean=function(s){ return s.replace(/^[+-]\s*/,''); };
    btn.textContent = (n<=2) ? gasBuildupSel.map(clean).join(' + ') : clean(gasBuildupSel[0])+' + '+(n-1)+' more';
    btn.title = gasBuildupSel.join(', ');
}

function updateGasBuildupChart() {
    if (!DATA || !DATA.chart_groups) return;
    var unitKey=document.getElementById('unitSelector').value;
    var period=document.getElementById('periodSelector').value;
    var view=DATA.views[period];
    if (!view) return;
    var vis=getVisibleIndices();
    var labels=[]; for (var k=0;k<vis.length;k++) labels.push(view.short_columns[vis[k]]||view.columns[vis[k]]);

    var sr=gasStorageRow();
    var isStoragePct = (gasBuildupSel.length===1 && gasBuildupSel[0]===sr);
    var stacked = (gasBuildupType==='bar' || gasBuildupType==='area') && !isStoragePct;

    var datasets=[];
    if (isStoragePct) {
        var row=getRowByLabel(sr), data=[];
        for (var k=0;k<vis.length;k++) data.push(row ? row.base[vis[k]]*100 : null);
        datasets.push({ label:'Storage %', data:data, borderColor:'#272962', backgroundColor:'rgba(0,0,0,0)', borderWidth:2, pointRadius:0, fill:false, tension:0.2 });
    } else {
        var leaves=gasBuildupLeaves().filter(function(l){ return gasBuildupSel.indexOf(l.row)>=0; });
        for (var s=0;s<leaves.length;s++) {
            var lf=leaves[s], row=getRowByLabel(lf.row), data=[];
            for (var k=0;k<vis.length;k++) {
                var ci=vis[k];
                var raw = row ? (lf.abs ? Math.abs(row.base[ci]) : row.base[ci]) : 0;
                var v = applyUnitConversion(raw, unitKey, view.days[ci]);
                if (stacked && v<0) v=0;   // keep stacks clean
                data.push(v);
            }
            var color=getSourceColor(lf.row);
            var ds={ label:lf.row, data:data, borderColor:color, pointRadius:0, tension:0.2 };
            if (gasBuildupType==='bar') {
                ds.backgroundColor=color+'CC'; ds.borderWidth=1; ds.stack='gasb';
            } else if (gasBuildupType==='area') {
                ds.backgroundColor=color+'CC'; ds.borderWidth=1; ds.stack='gasb'; ds.fill = (s===0)?'origin':'-1';
            } else {  // line
                ds.backgroundColor='rgba(0,0,0,0)'; ds.borderWidth=2; ds.fill=false;
            }
            datasets.push(ds);
        }
    }

    var unitLabel = isStoragePct ? '%' : getHeaderUnitLabel(unitKey);
    document.getElementById('buildupTitle').textContent = 'Gas balance';

    var chartType = (gasBuildupType==='bar') ? 'bar' : 'line';
    var opts = stacked ? chartOptionsStacked(unitKey) : chartOptionsLine(unitKey);
    opts.scales.y.title.text = unitLabel;
    if (isStoragePct) opts.scales.y.ticks.callback = function(v){ return v + '%'; };

    if (buildupChart) { buildupChart.destroy(); buildupChart=null; }
    var canvas=document.getElementById('buildupCanvas'); if (!canvas) return;
    buildupChart=new Chart(canvas, { type:chartType, data:{labels:labels, datasets:datasets}, options:opts });
    buildupStackHover = stacked;
    if (stacked) initBuildupHover(canvas);
    renderHtmlLegend(buildupChart, 'buildupLegend', {});
}

function updateGasCharts() {
    if (!DATA.chart_groups) return;
    updateGasRangeChart();
    updateGasBuildupChart();
}

/* ============================================================
   GLOBAL LNG — composite tab (4 stacked tables + 2x2 charts).
   Self-contained: never touches the single-DATA gas/power path. Reuses the
   pure helpers (formatNum, formatGrowth, computeGrowthCell, computeDisplayValue,
   applyUnitConversion, chartOptionsLine/Stacked, renderHtmlLegend, gyMonth*).
   Tables/change tables follow the top Period/Unit/From/To bar; each chart has
   its own controls. Base unit MMT.
   ============================================================ */
var LNG = null;
var lngCharts = { impRange: null, expRange: null, impStack: null, expStack: null };
var lngExpanded = { imports: {}, exports: {} };
var lngChangeMode = { imports: 'pct', exports: 'pct' };
var lngChangeType = { imports: 'yoy', exports: 'yoy' };
var lngStackType  = { imports: 'bar', exports: 'bar' };
var lngStackSel   = { imports: [], exports: [] };
// Exports use a hierarchical selector (regions expand to member countries; mixed
// partial selections allowed). State = selected LEAF / standalone labels, kept
// separately for the range chart and the build-up chart.
var lngExpSelState = { range: [], stack: [] };
var lngExpExpanded = { range: {}, stack: {} };   // region label -> expanded?

function lngSubBlob(k) { return LNG.sub[k]; }
function lngGetRow(sub, label) { var rs = sub.views.Monthly.rows; for (var i=0;i<rs.length;i++) if (rs[i].label===label) return rs[i]; return null; }
function lngSumAt(sub, labels, idx) { var s=0; for (var i=0;i<labels.length;i++){ var r=lngGetRow(sub,labels[i]); if (r && idx>=0 && idx<r.base.length) s+=r.base[idx]; } return s; }
function lngFindMonthlyIdx(sub, y, m) { var meta=sub.views.Monthly.col_meta; for (var i=0;i<meta.length;i++) if (meta[i].year===y && meta[i].month===m) return i; return -1; }
function lngColor(sub, label) { return sub.source_colors[label] || '#888'; }
function lngToggleArr(arr, v) { var p=arr.indexOf(v); if (p>=0) arr.splice(p,1); else arr.push(v); }
function lngIds(subKey) {
    return subKey === 'imports'
        ? { th:'lngImpHead', tb:'lngImpBody', cth:'lngImpChHead', ctb:'lngImpChBody',
            rangeSel:'msLngImpRange', lookback:'lngImpRangeLookback', rangeCanvas:'lngImpRangeCanvas',
            rangeTitle:'lngImpRangeTitle', rangeLegend:'lngImpRangeLegend', rangeChart:'impRange',
            stackSel:'msLngImpStack', from:'lngImpStackFrom', to:'lngImpStackTo', stackCanvas:'lngImpStackCanvas',
            stackTitle:'lngImpStackTitle', stackLegend:'lngImpStackLegend', stackChart:'impStack', stackBtn:'lngImpStack',
            changeType:'lngImpChangeType', modePct:'lngImpModePct', modeAbs:'lngImpModeAbs' }
        : { th:'lngExpHead', tb:'lngExpBody', cth:'lngExpChHead', ctb:'lngExpChBody',
            rangeSel:'msLngExpRange', lookback:'lngExpRangeLookback', rangeCanvas:'lngExpRangeCanvas',
            rangeTitle:'lngExpRangeTitle', rangeLegend:'lngExpRangeLegend', rangeChart:'expRange',
            stackSel:'msLngExpStack', from:'lngExpStackFrom', to:'lngExpStackTo', stackCanvas:'lngExpStackCanvas',
            stackTitle:'lngExpStackTitle', stackLegend:'lngExpStackLegend', stackChart:'expStack', stackBtn:'lngExpStack',
            changeType:'lngExpChangeType', modePct:'lngExpModePct', modeAbs:'lngExpModeAbs' };
}

function lngInit() {
    LNG = DATA;
    var k;
    for (k in lngCharts) { if (lngCharts[k]) { lngCharts[k].destroy(); lngCharts[k] = null; } }
    lngExpanded = { imports: {}, exports: {} };
    lngChangeMode = { imports: 'pct', exports: 'pct' };
    lngChangeType = { imports: 'yoy', exports: 'yoy' };
    lngStackType  = { imports: 'bar', exports: 'bar' };

    var imp = lngSubBlob('imports'), exp = lngSubBlob('exports');
    // Imports: flat selectors. Range = countries + a mutex Total (default Total);
    // build-up = All-checklist of the 12 countries.
    buildMultiSelect('msLngImpRange', imp.chart_series, imp.total_row, [imp.total_row], function(){ lngUpdateRange('imports'); });
    lngStackSel.imports = imp.chart_series.slice();
    lngBuildStackSelector('imports');

    // Exports: hierarchical selectors for BOTH charts. Default = every leaf
    // selected (collapsed) -> range = Total, build-up = the 7 region totals.
    var expLeaves = lngExpLeaves();
    lngExpSelState = { range: expLeaves.slice(), stack: expLeaves.slice() };
    lngExpExpanded = { range: {}, stack: {} };
    lngBuildExpSelector('range');
    lngBuildExpSelector('stack');

    lngPopulateLookback('lngImpRangeLookback', imp);
    lngPopulateLookback('lngExpRangeLookback', exp);
    lngPopulateStackRange('imports');
    lngPopulateStackRange('exports');
    lngSetStackTypeButtons('imports');
    lngSetStackTypeButtons('exports');

    lngSyncChangeTypeSelectors();
    lngSetModeButtons('imports');
    lngSetModeButtons('exports');
    lngUpdateAll();
}

function lngUpdateAll() {
    if (!LNG) return;
    lngSyncChangeTypeSelectors();
    lngRenderTable('imports');
    lngRenderTable('exports');
    lngRenderChange('imports');
    lngRenderChange('exports');
    lngUpdateRange('imports');
    lngUpdateRange('exports');
    lngUpdateStack('imports');
    lngUpdateStack('exports');
}

/* ---- Tables ---- */
function lngRenderTable(subKey) {
    var sub = lngSubBlob(subKey), ids = lngIds(subKey);
    var period = document.getElementById('periodSelector').value;
    var unitKey = document.getElementById('unitSelector').value;
    var view = sub.views[period]; if (!view) return;
    var vis = getVisibleIndices(), days = view.days, rows = view.rows, hierarchy = sub.hierarchy;
    var unitLabel = getHeaderUnitLabel(unitKey);

    var hHtml = '<tr><th>'+unitLabel+'<span class="unit-label">('+period+')</span></th>';
    for (var i=0;i<vis.length;i++) hHtml += '<th>'+(view.short_columns[vis[i]]||view.columns[vis[i]])+'</th>';
    document.getElementById(ids.th).innerHTML = hHtml+'</tr>';

    var bHtml = '';
    for (var h=0;h<hierarchy.length;h++) {
        var item = hierarchy[h], rowData = rows[item.index];
        var isExp = lngExpanded[subKey][item.label] || false;
        var hasCh = item.children && item.children.length>0;
        var rc = item.type==='standalone' ? 'standalone-row' : 'parent-row';
        bHtml += '<tr class="'+rc+'">';
        var lbl = '';
        if (hasCh) lbl += '<span class="toggle-arrow'+(isExp?' expanded':'')+'">&#9654;</span> ';
        lbl += item.label.replace(/&/g,'&amp;').replace(/</g,'&lt;');
        bHtml += hasCh ? '<td data-lng-toggle="'+subKey+':'+h+'">'+lbl+'</td>' : '<td>'+lbl+'</td>';
        for (var i=0;i<vis.length;i++) {
            var ci = vis[i];
            bHtml += '<td>'+formatNum(computeDisplayValue(rowData.base[ci], unitKey, false, false, days[ci]), false)+'</td>';
        }
        bHtml += '</tr>';
        if (hasCh) {
            for (var c=0;c<item.children.length;c++) {
                var ch = item.children[c], cd = rows[ch.index];
                var hid = !isExp ? ' hidden' : '';
                bHtml += '<tr class="child-row'+hid+'"><td>'+ch.label.replace(/&/g,'&amp;').replace(/</g,'&lt;')+'</td>';
                for (var i=0;i<vis.length;i++) {
                    var ci = vis[i];
                    bHtml += '<td>'+formatNum(computeDisplayValue(cd.base[ci], unitKey, false, false, days[ci]), false)+'</td>';
                }
                bHtml += '</tr>';
            }
        }
    }
    document.getElementById(ids.tb).innerHTML = bHtml;
}

/* ---- Change tables (reuse computeGrowthCell/formatGrowth via the growthMode global) ---- */
function lngRenderChange(subKey) {
    var sub = lngSubBlob(subKey), ids = lngIds(subKey);
    var period = document.getElementById('periodSelector').value;
    var view = sub.views[period]; if (!view) return;
    var vis = getVisibleIndices(), days = view.days, rows = view.rows, meta = view.col_meta, hierarchy = sub.hierarchy;
    var gt = lngChangeType[subKey];
    var prevMode = growthMode; growthMode = lngChangeMode[subKey];   // computeGrowthCell + formatGrowth read this

    var hHtml = '<tr><th>Change</th>';
    for (var i=0;i<vis.length;i++) hHtml += '<th>'+(view.short_columns[vis[i]]||view.columns[vis[i]])+'</th>';
    document.getElementById(ids.cth).innerHTML = hHtml+'</tr>';

    var bHtml = '';
    for (var h=0;h<hierarchy.length;h++) {
        var item = hierarchy[h], rowData = rows[item.index];
        var isExp = lngExpanded[subKey][item.label] || false;
        var hasCh = item.children && item.children.length>0;
        var rc = item.type==='standalone' ? 'standalone-row' : 'parent-row';
        bHtml += '<tr class="'+rc+'">';
        var dl = '';
        if (hasCh) dl += '<span class="toggle-arrow'+(isExp?' expanded':'')+'">&#9654;</span> ';
        dl += item.label.replace(/&/g,'&amp;').replace(/</g,'&lt;');
        bHtml += hasCh ? '<td data-lng-toggle-g="'+subKey+':'+h+'">'+dl+'</td>' : '<td>'+dl+'</td>';
        for (var i=0;i<vis.length;i++) {
            var ci = vis[i];
            var gv = computeGrowthCell(rowData.base, days, meta, ci, period, gt, false);
            if (gv===null) { bHtml += '<td></td>'; continue; }
            var cls = gv>0.0001 ? 'g-pos' : (gv<-0.0001 ? 'g-neg' : '');
            bHtml += '<td'+(cls?' class="'+cls+'"':'')+'>'+formatGrowth(gv, false)+'</td>';
        }
        bHtml += '</tr>';
        if (hasCh) {
            for (var c=0;c<item.children.length;c++) {
                var ch = item.children[c], cd = rows[ch.index];
                var hid = !isExp ? ' hidden' : '';
                bHtml += '<tr class="child-row'+hid+'"><td>'+ch.label.replace(/&/g,'&amp;').replace(/</g,'&lt;')+'</td>';
                for (var i=0;i<vis.length;i++) {
                    var ci = vis[i];
                    var gv = computeGrowthCell(cd.base, days, meta, ci, period, gt, false);
                    if (gv===null) { bHtml += '<td></td>'; continue; }
                    var cls = gv>0.0001 ? 'g-pos' : (gv<-0.0001 ? 'g-neg' : '');
                    bHtml += '<td'+(cls?' class="'+cls+'"':'')+'>'+formatGrowth(gv, false)+'</td>';
                }
                bHtml += '</tr>';
            }
        }
    }
    document.getElementById(ids.ctb).innerHTML = bHtml;
    growthMode = prevMode;
}

function lngSyncChangeTypeSelectors() {
    var period = document.getElementById('periodSelector').value;
    var opts = GROWTH_OPTS[period] || [{k:'yoy',l:'Year-on-Year'}];
    ['imports','exports'].forEach(function(sk) {
        var sel = document.getElementById(lngIds(sk).changeType); if (!sel) return;
        var cur = lngChangeType[sk];
        if (!opts.some(function(o){ return o.k===cur; })) { cur = opts[0].k; lngChangeType[sk] = cur; }
        var html = '';
        for (var i=0;i<opts.length;i++) html += '<option value="'+opts[i].k+'"'+(opts[i].k===cur?' selected':'')+'>'+opts[i].l+'</option>';
        sel.innerHTML = html;
    });
}
function lngOnChangeType(subKey) { lngChangeType[subKey] = document.getElementById(lngIds(subKey).changeType).value; lngRenderChange(subKey); }
function lngSetMode(subKey, mode) { lngChangeMode[subKey] = mode; lngSetModeButtons(subKey); lngRenderChange(subKey); }
function lngSetModeButtons(subKey) {
    var ids = lngIds(subKey);
    var p = document.getElementById(ids.modePct), a = document.getElementById(ids.modeAbs);
    if (p) p.className = lngChangeMode[subKey]==='pct' ? 'active' : '';
    if (a) a.className = lngChangeMode[subKey]==='abs' ? 'active' : '';
}

/* ---- Range charts (gas-year cycle: prev / current / next + n-year avg + band).
        Current line is capped at the latest actual month. ---- */
function lngPopulateLookback(id, sub) {
    var sel = document.getElementById(id); if (!sel) return;
    var cgy = currentGY(), firstYear = sub.views.Monthly.col_meta[0].year;
    var maxLb = Math.max(3, cgy - firstYear);
    var html = '';
    for (var n=3;n<=maxLb;n++) html += '<option value="'+n+'"'+(n===5?' selected':'')+'>'+n+' years</option>';
    sel.innerHTML = html;
}

function lngUpdateRange(subKey) {
    var sub = lngSubBlob(subKey), ids = lngIds(subKey);
    var selected = lngRangeSelected(subKey);   // leaf labels (or [total_row])
    var unitKey = document.getElementById('unitSelector').value;
    var lookback = parseInt(document.getElementById(ids.lookback).value) || 5;
    var cal = isCalCycle(document.getElementById('periodSelector').value);
    var cgy = cycleBase(cal), prevGY = cgy-1, nextGY = cgy+1;
    var labels = cycleLabels(cal);
    var latestIdx = LNG.latest_actual ? LNG.latest_actual.index : 1e9;

    function valAt(idx) {
        if (idx<0) return null;
        var b = lngSumAt(sub, selected, idx);
        return applyUnitConversion(b, unitKey, sub.views.Monthly.days[idx]);
    }
    function gyValues(gy, cap) {
        var out = [];
        for (var i=0;i<12;i++) {
            var amt = cycleToActual(cal, gy, i);
            var idx = lngFindMonthlyIdx(sub, amt.year, amt.month);
            if (cap && idx > latestIdx) { out.push(null); continue; }   // cap current line at latest actual
            out.push(valAt(idx));
        }
        return out;
    }
    var prevVals = gyValues(prevGY, false), currVals = gyValues(cgy, true), nextVals = gyValues(nextGY, false);

    var lbStart = cgy-lookback, lbEnd = cgy-1, avg=[], mn=[], mx=[];
    for (var mo=0;mo<12;mo++) {
        var samples=[];
        for (var gy=lbStart;gy<=lbEnd;gy++) {
            var amt = cycleToActual(cal, gy, mo);
            var v = valAt(lngFindMonthlyIdx(sub, amt.year, amt.month));
            if (v!==null && v!==undefined) samples.push(v);
        }
        if (samples.length===0) { avg.push(null); mn.push(null); mx.push(null); continue; }
        var s=0; for (var k=0;k<samples.length;k++) s+=samples[k];
        avg.push(s/samples.length); mn.push(Math.min.apply(null,samples)); mx.push(Math.max.apply(null,samples));
    }

    var unitLabel = getHeaderUnitLabel(unitKey);
    var selLabel = lngRangeLabel(subKey, selected);
    document.getElementById(ids.rangeTitle).textContent = (subKey==='imports'?'Import Range':'Export Range') + ' — ' + selLabel + ' (' + unitLabel + ')';

    var prevLabel = cycleSeriesLabel(cal, prevGY);
    var curLabel  = cycleSeriesLabel(cal, cgy, ' (current)');
    var nextLabel = cycleSeriesLabel(cal, nextGY, ' (forecast)');
    var datasets = [
        { label: lookback+'y max',     data: mx,       borderColor:'rgba(0,0,0,0)', backgroundColor:'rgba(0,0,0,0)', pointRadius:0, fill:false, order:20 },
        { label: lookback+'y range',   data: mn,       borderColor:'rgba(0,0,0,0)', backgroundColor:'rgba(147,149,162,0.28)', pointRadius:0, fill:'-1', order:19 },
        { label: lookback+'y average', data: avg,      borderColor:'#272962', backgroundColor:'rgba(0,0,0,0)', borderWidth:1.8, pointRadius:0, fill:false, tension:0.25, order:4 },
        { label: prevLabel,            data: prevVals, borderColor:'#0C5B19', backgroundColor:'rgba(0,0,0,0)', borderWidth:1.8, pointRadius:2, fill:false, tension:0.25, order:3 },
        { label: curLabel,             data: currVals, borderColor:'#C00000', backgroundColor:'rgba(0,0,0,0)', borderWidth:2.6, pointRadius:3, fill:false, tension:0.25, order:1 },
        { label: nextLabel,            data: nextVals, borderColor:'#539648', backgroundColor:'rgba(0,0,0,0)', borderWidth:1.8, borderDash:[6,4], pointRadius:2, fill:false, tension:0.25, order:2 },
    ];
    var opts = chartOptionsLine(unitKey);
    opts.scales.y.title.text = unitLabel;

    if (lngCharts[ids.rangeChart]) { lngCharts[ids.rangeChart].destroy(); lngCharts[ids.rangeChart] = null; }
    var canvas = document.getElementById(ids.rangeCanvas); if (!canvas) return;
    lngCharts[ids.rangeChart] = new Chart(canvas, { type:'line', data:{labels:labels, datasets:datasets}, options:opts });
    renderHtmlLegend(lngCharts[ids.rangeChart], ids.rangeLegend, { filter: function(label){ return !(label||'').endsWith(' max'); } });
}

/* ---- Stacked build-up charts (own from/to + chart-type toggle) ---- */
function lngBuildStackSelector(subKey) {
    var sub = lngSubBlob(subKey), ids = lngIds(subKey);
    var container = document.getElementById(ids.stackSel); if (!container) return;
    container.innerHTML = ''; container.classList.add('multi-select');
    var button = document.createElement('button'); button.type='button'; button.className='ms-button'; container.appendChild(button);
    var panel = document.createElement('div'); panel.className='ms-panel'; container.appendChild(panel);
    panel.appendChild(lngStackItem(subKey, 'All', 'ms-group-header', '__ALL__'));
    var div = document.createElement('div'); div.className='ms-divider'; panel.appendChild(div);
    for (var i=0;i<sub.chart_series.length;i++) panel.appendChild(lngStackItem(subKey, sub.chart_series[i], '', sub.chart_series[i]));
    button.addEventListener('click', function(ev){ ev.stopPropagation(); container.classList.toggle('open'); });
    lngStackRefresh(subKey);
}
function lngStackItem(subKey, text, extraCls, key) {
    var lbl = document.createElement('label'); lbl.className = 'ms-item'+(extraCls?' '+extraCls:'');
    var cb = document.createElement('input'); cb.type='checkbox'; cb.setAttribute('data-key', key);
    var sp = document.createElement('span'); sp.textContent = text;
    lbl.appendChild(cb); lbl.appendChild(sp);
    cb.addEventListener('change', function(){ lngStackClick(subKey, key); lngStackRefresh(subKey); lngUpdateStack(subKey); });
    return lbl;
}
function lngStackClick(subKey, key) {
    var sub = lngSubBlob(subKey), all = sub.chart_series.slice();
    if (key==='__ALL__') {
        var allIn = all.every(function(r){ return lngStackSel[subKey].indexOf(r)>=0; });
        lngStackSel[subKey] = allIn ? [] : all.slice();
        return;
    }
    lngToggleArr(lngStackSel[subKey], key);
}
function lngStackRefresh(subKey) {
    var sub = lngSubBlob(subKey), ids = lngIds(subKey), all = sub.chart_series.slice();
    var nAll = all.filter(function(r){ return lngStackSel[subKey].indexOf(r)>=0; }).length;
    lngStackSetCb(ids.stackSel, '__ALL__', nAll===all.length, nAll>0 && nAll<all.length);
    for (var i=0;i<all.length;i++) lngStackSetCb(ids.stackSel, all[i], lngStackSel[subKey].indexOf(all[i])>=0, false);
    lngStackBtnLabel(subKey);
}
function lngStackSetCb(containerId, key, checked, indet) {
    var cb = document.querySelector('#'+containerId+' input[data-key="'+CSS.escape(key)+'"]');
    if (!cb) return; cb.checked = !!checked; cb.indeterminate = !!indet;
}
function lngStackBtnLabel(subKey) {
    var sub = lngSubBlob(subKey), ids = lngIds(subKey);
    var c = document.getElementById(ids.stackSel); if (!c) return;
    var btn = c.querySelector('.ms-button'); var all = sub.chart_series.slice(), sel = lngStackSel[subKey];
    if (all.every(function(r){ return sel.indexOf(r)>=0; })) { btn.textContent = 'All'; btn.title = all.join(', '); return; }
    if (sel.length===0) { btn.textContent = 'None'; btn.title = ''; return; }
    btn.textContent = (sel.length<=2) ? sel.join(' + ') : sel[0]+' + '+(sel.length-1)+' more';
    btn.title = sel.join(', ');
}
function lngSetStackType(subKey, t) { lngStackType[subKey] = t; lngSetStackTypeButtons(subKey); lngUpdateStack(subKey); }
function lngSetStackTypeButtons(subKey) {
    var pre = lngIds(subKey).stackBtn, btns = ['bar','area','line'];
    for (var i=0;i<btns.length;i++) { var b = document.getElementById('btn'+pre+'-'+btns[i]); if (b) b.className = (btns[i]===lngStackType[subKey]) ? 'active' : ''; }
}
function lngPopulateStackRange(subKey) {
    var sub = lngSubBlob(subKey), ids = lngIds(subKey), view = sub.views.Monthly;
    var fs = document.getElementById(ids.from), ts = document.getElementById(ids.to); if (!fs || !ts) return;
    var html = '';
    for (var i=0;i<view.col_meta.length;i++) html += '<option value="'+i+'">'+view.col_meta[i].label+'</option>';
    fs.innerHTML = html; ts.innerHTML = html;
    var fromIdx = 0;
    for (var i=0;i<view.col_meta.length;i++) { if (view.col_meta[i].year===2025 && view.col_meta[i].month===1) { fromIdx = i; break; } }
    var toIdx = view.col_meta.length-1;
    if (LNG.latest_actual) toIdx = Math.min(view.col_meta.length-1, LNG.latest_actual.index + 8);
    fs.value = fromIdx; ts.value = toIdx;
}
function lngUpdateStack(subKey) {
    var sub = lngSubBlob(subKey), ids = lngIds(subKey);
    var type = lngStackType[subKey];
    var unitKey = document.getElementById('unitSelector').value;
    var view = sub.views.Monthly;
    var fromIdx = parseInt(document.getElementById(ids.from).value), toIdx = parseInt(document.getElementById(ids.to).value);
    if (isNaN(fromIdx) || isNaN(toIdx)) { fromIdx = 0; toIdx = view.col_meta.length-1; }
    if (fromIdx > toIdx) { var t = fromIdx; fromIdx = toIdx; toIdx = t; }
    var labels = [];
    for (var i=fromIdx;i<=toIdx;i++) labels.push(view.short_columns[i] || view.columns[i]);

    var stacked = (type==='bar' || type==='area');
    // Series descriptors: imports = selected countries; exports = region totals
    // for fully-selected regions, else the individually-selected member countries.
    var series = lngStackDescriptors(subKey);
    var datasets = [];
    for (var s=0;s<series.length;s++) {
        var d = series[s], row = lngGetRow(sub, d.row), data = [];
        for (var i=fromIdx;i<=toIdx;i++) {
            var v = row ? applyUnitConversion(row.base[i], unitKey, view.days[i]) : 0;
            if (stacked && v<0) v = 0;
            data.push(v);
        }
        var ds = { label: d.label, data: data, borderColor: d.color, pointRadius: 0, tension: 0.2 };
        if (type==='bar') { ds.backgroundColor = fadeColor(d.color, 0.8); ds.borderWidth = 1; ds.stack = 'lng'; }
        else if (type==='area') { ds.backgroundColor = fadeColor(d.color, 0.8); ds.borderWidth = 1; ds.stack = 'lng'; ds.fill = (s===0)?'origin':'-1'; }
        else { ds.backgroundColor = 'rgba(0,0,0,0)'; ds.borderWidth = 2; ds.fill = false; }
        datasets.push(ds);
    }

    var unitLabel = getHeaderUnitLabel(unitKey);
    document.getElementById(ids.stackTitle).textContent = (subKey==='imports'?'Import Build-up':'Export Build-up') + ' (' + unitLabel + ')';
    var chartType = (type==='bar') ? 'bar' : 'line';
    var opts = stacked ? chartOptionsStacked(unitKey) : chartOptionsLine(unitKey);
    opts.scales.y.title.text = unitLabel;

    if (lngCharts[ids.stackChart]) { lngCharts[ids.stackChart].destroy(); lngCharts[ids.stackChart] = null; }
    var canvas = document.getElementById(ids.stackCanvas); if (!canvas) return;
    lngCharts[ids.stackChart] = new Chart(canvas, { type: chartType, data: {labels:labels, datasets:datasets}, options: opts });
    renderHtmlLegend(lngCharts[ids.stackChart], ids.stackLegend, {});
}

/* ---- Exports hierarchical selector (range + build-up share the widget code,
        with separate selection/expand state per chart). Regions expand to their
        member countries; selecting a region selects all members; partial
        selection shows the chosen countries instead of the region total. ---- */
function lngExpLeaves() {
    var sub = lngSubBlob('exports'), out = [];
    sub.hierarchy.forEach(function(h) {
        if (h.label === sub.total_row) return;            // skip Grand total
        if (h.children && h.children.length) h.children.forEach(function(c){ out.push(c.label); });
        else out.push(h.label);                            // standalone (Australia, Russia)
    });
    return out;
}
function lngExpContainer(which) { return which === 'range' ? 'msLngExpRange' : 'msLngExpStack'; }
function lngExpOnChange(which) { if (which === 'range') lngUpdateRange('exports'); else lngUpdateStack('exports'); }

function lngBuildExpSelector(which) {
    var sub = lngSubBlob('exports');
    var container = document.getElementById(lngExpContainer(which)); if (!container) return;
    container.innerHTML = ''; container.classList.add('multi-select');
    var button = document.createElement('button'); button.type='button'; button.className='ms-button'; container.appendChild(button);
    var panel = document.createElement('div'); panel.className='ms-panel'; container.appendChild(panel);
    panel.appendChild(lngExpItem(which, 'All', 'ms-group-header', '__ALL__', null));
    var div = document.createElement('div'); div.className='ms-divider'; panel.appendChild(div);
    sub.hierarchy.forEach(function(h) {
        if (h.label === sub.total_row) return;
        if (h.children && h.children.length) {
            panel.appendChild(lngExpGroupHeader(which, h));
            h.children.forEach(function(c){ panel.appendChild(lngExpItem(which, c.label, 'ms-child ms-exp-member', c.label, h.label)); });
        } else {
            panel.appendChild(lngExpItem(which, h.label, '', h.label, null));
        }
    });
    button.addEventListener('click', function(ev){ ev.stopPropagation(); container.classList.toggle('open'); });
    lngExpRefresh(which);
}
function lngExpItem(which, text, extraCls, key, region) {
    var lbl = document.createElement('label'); lbl.className = 'ms-item'+(extraCls?' '+extraCls:'');
    if (region != null) lbl.setAttribute('data-region', region);
    var cb = document.createElement('input'); cb.type='checkbox'; cb.setAttribute('data-key', key);
    var sp = document.createElement('span'); sp.textContent = text;
    lbl.appendChild(cb); lbl.appendChild(sp);
    cb.addEventListener('change', function(){ lngExpItemClick(which, key, region); });
    return lbl;
}
function lngExpGroupHeader(which, h) {
    var lbl = document.createElement('label'); lbl.className = 'ms-item ms-group-header';
    var arrow = document.createElement('span'); arrow.className='ms-exp-arrow'; arrow.textContent='▸';
    arrow.setAttribute('data-region-arrow', h.label);
    var cb = document.createElement('input'); cb.type='checkbox'; cb.setAttribute('data-key', '__grp__'+h.label);
    var sp = document.createElement('span'); sp.textContent = h.label;
    lbl.appendChild(arrow); lbl.appendChild(cb); lbl.appendChild(sp);
    arrow.addEventListener('click', function(ev){ ev.preventDefault(); ev.stopPropagation(); lngExpToggleExpand(which, h.label); });
    cb.addEventListener('change', function(){ lngExpClickGroup(which, h); });
    return lbl;
}
function lngExpItemClick(which, key, region) {
    if (key === '__ALL__') { lngExpClickAll(which); }
    else { lngToggleArr(lngExpSelState[which], key); if (region != null) lngExpExpanded[which][region] = true; }
    lngExpRefresh(which); lngExpOnChange(which);
}
function lngExpClickGroup(which, h) {
    var members = h.children.map(function(c){ return c.label; });
    var sel = lngExpSelState[which];
    var allIn = members.every(function(m){ return sel.indexOf(m)>=0; });
    if (allIn) {
        lngExpSelState[which] = sel.filter(function(m){ return members.indexOf(m)<0; });
        lngExpExpanded[which][h.label] = false;          // collapse on full deselect
    } else {
        members.forEach(function(m){ if (sel.indexOf(m)<0) sel.push(m); });
        lngExpExpanded[which][h.label] = true;           // auto-expand to show members
    }
    lngExpRefresh(which); lngExpOnChange(which);
}
function lngExpClickAll(which) {
    var leaves = lngExpLeaves(), sel = lngExpSelState[which];
    var allIn = leaves.every(function(l){ return sel.indexOf(l)>=0; });
    if (allIn) { lngExpSelState[which] = []; lngExpExpanded[which] = {}; }
    else { lngExpSelState[which] = leaves.slice(); }
}
function lngExpToggleExpand(which, region) { lngExpExpanded[which][region] = !lngExpExpanded[which][region]; lngExpRefresh(which); }
function lngExpSetCb(containerId, key, checked, indet) {
    var cb = document.querySelector('#'+containerId+' input[data-key="'+CSS.escape(key)+'"]');
    if (!cb) return; cb.checked = !!checked; cb.indeterminate = !!indet;
}
function lngExpRefresh(which) {
    var sub = lngSubBlob('exports'), containerId = lngExpContainer(which);
    var sel = lngExpSelState[which], leaves = lngExpLeaves();
    var nAll = leaves.filter(function(l){ return sel.indexOf(l)>=0; }).length;
    lngExpSetCb(containerId, '__ALL__', nAll===leaves.length, nAll>0 && nAll<leaves.length);
    sub.hierarchy.forEach(function(h) {
        if (h.label === sub.total_row) return;
        if (h.children && h.children.length) {
            var members = h.children.map(function(c){ return c.label; });
            var n = members.filter(function(m){ return sel.indexOf(m)>=0; }).length;
            lngExpSetCb(containerId, '__grp__'+h.label, n===members.length, n>0 && n<members.length);
            for (var j=0;j<members.length;j++) lngExpSetCb(containerId, members[j], sel.indexOf(members[j])>=0, false);
            var expanded = !!lngExpExpanded[which][h.label];
            var rows = document.querySelectorAll('#'+containerId+' [data-region="'+CSS.escape(h.label)+'"]');
            for (var r=0;r<rows.length;r++) rows[r].style.display = expanded ? '' : 'none';
            var arrow = document.querySelector('#'+containerId+' [data-region-arrow="'+CSS.escape(h.label)+'"]');
            if (arrow) arrow.className = 'ms-exp-arrow'+(expanded?' expanded':'');
        } else {
            lngExpSetCb(containerId, h.label, sel.indexOf(h.label)>=0, false);
        }
    });
    lngExpBtnLabel(which);
}
function lngExpBtnLabel(which) {
    var c = document.getElementById(lngExpContainer(which)); if (!c) return;
    var btn = c.querySelector('.ms-button'); if (!btn) return;
    var leaves = lngExpLeaves(), sel = lngExpSelState[which];
    if (sel.length >= leaves.length) { btn.textContent = 'All'; btn.title = ''; return; }
    if (sel.length === 0) { btn.textContent = 'None'; btn.title = ''; return; }
    btn.textContent = (sel.length<=2) ? sel.join(' + ') : sel.length+' selected';
    btn.title = sel.join(', ');
}

/* Resolve export selections to chart series.
   Build-up: a fully-selected region -> one region-total series; a partially
   selected region -> one series per selected member country; standalones as-is. */
function lngExpStackSeries() {
    var sub = lngSubBlob('exports'), sel = lngExpSelState.stack, out = [];
    sub.hierarchy.forEach(function(h) {
        if (h.label === sub.total_row) return;
        if (h.children && h.children.length) {
            var selM = h.children.filter(function(c){ return sel.indexOf(c.label)>=0; }).map(function(c){ return c.label; });
            if (selM.length === 0) return;
            if (selM.length === h.children.length) out.push({ label: h.label, row: h.label, color: lngColor(sub, h.label) });
            else selM.forEach(function(m){ out.push({ label: m, row: m, color: lngExpMemberColor(h.label, m) }); });
        } else if (sel.indexOf(h.label) >= 0) {
            out.push({ label: h.label, row: h.label, color: lngColor(sub, h.label) });
        }
    });
    return out;
}
function lngStackDescriptors(subKey) {
    if (subKey === 'exports') return lngExpStackSeries();
    var sub = lngSubBlob('imports'), sel = lngStackSel.imports;
    return sub.chart_series.filter(function(s){ return sel.indexOf(s)>=0; })
        .map(function(s){ return { label: s, row: s, color: lngColor(sub, s) }; });
}
function lngRangeSelected(subKey) {
    if (subKey === 'exports') {
        var s = lngExpSelState.range;
        return (s && s.length) ? s.slice() : [lngSubBlob('exports').total_row];
    }
    var st = getMultiSelectState('msLngImpRange');
    return (st && st.length) ? st : [lngSubBlob('imports').total_row];
}
function lngRangeLabel(subKey, selected) {
    if (subKey === 'exports') {
        var all = lngExpLeaves();
        if (selected.length >= all.length || selected.length === 0) return 'Total';
        return selected.length<=2 ? selected.join(' + ') : selected.length+' selected';
    }
    var sub = lngSubBlob('imports');
    if (selected.length===1 && selected[0]===sub.total_row) return 'Total';
    return selected.length<=2 ? selected.join(' + ') : selected[0]+' + '+(selected.length-1)+' more';
}
/* Member-country color: a lightness-shifted shade of the parent region color,
   so a drilled region's countries stay visually related in the stack. */
function lngExpMemberColor(region, member) {
    var sub = lngSubBlob('exports');
    var h = null; for (var i=0;i<sub.hierarchy.length;i++) if (sub.hierarchy[i].label===region) { h = sub.hierarchy[i]; break; }
    if (!h) return lngColor(sub, region);
    var idx = 0, n = h.children.length;
    for (var j=0;j<n;j++) if (h.children[j].label===member) { idx = j; break; }
    var t = (n>1) ? (idx/(n-1)) : 0.5;
    return lngShade(lngColor(sub, region), -0.28 + t*0.62);   // darken -> lighten across members
}
function lngShade(color, amt) {   // amt in [-1,1]: <0 darken, >0 lighten
    var hex = String(color).replace('#','');
    if (hex.length === 3) hex = hex.split('').map(function(c){ return c+c; }).join('');
    if (hex.length < 6) return color;
    var r = parseInt(hex.slice(0,2),16), g = parseInt(hex.slice(2,4),16), b = parseInt(hex.slice(4,6),16);
    function adj(v) { return amt>=0 ? Math.round(v+(255-v)*amt) : Math.round(v*(1+amt)); }
    return 'rgb('+adj(r)+','+adj(g)+','+adj(b)+')';
}

/* LNG table expand/collapse toggles (value + change tables share expand state). */
document.addEventListener('click', function(e) {
    var td = e.target.closest('td[data-lng-toggle]');
    if (td) {
        var p = td.getAttribute('data-lng-toggle').split(':'), sk = p[0], idx = parseInt(p[1]);
        var label = lngSubBlob(sk).hierarchy[idx].label;
        lngExpanded[sk][label] = !lngExpanded[sk][label];
        lngRenderTable(sk); lngRenderChange(sk);
        return;
    }
    var td2 = e.target.closest('td[data-lng-toggle-g]');
    if (td2) {
        var p = td2.getAttribute('data-lng-toggle-g').split(':'), sk = p[0], idx = parseInt(p[1]);
        var label = lngSubBlob(sk).hierarchy[idx].label;
        lngExpanded[sk][label] = !lngExpanded[sk][label];
        lngRenderTable(sk); lngRenderChange(sk);
    }
});

/* ============================================================
   LNG PROJECTS tab — project database with a progressive filter bar and a
   Region -> Country -> Project tree (Sub-tab 1 "Projects"). Self-contained;
   no unit/period machinery. Sub-tab 2 "Supply Outlook" is stubbed for now.
   ============================================================ */
var PRJ = null;
var prjFilters = { table:{status:[],region:[],country:[],company:[],project:[]},
                   chart:{status:[],region:[],country:[],company:[],project:[]} };
var prjChartLinked = true;            // charts follow the table filters by default
var prjExpanded = {};      // assumptions tree: 'r|Region' | 'c|Region|Country' | 'p|Name'
var prjOutExpanded = {};   // supply-outlook tree expand state (shared by production + change)
var prjSubtab = 'outlook'; // default sub-tab
var prjChangeMode = 'pct'; // change table: 'pct' | 'abs'
var prjRangeChart = null, prjStackChart = null;   // Supply Outlook charts
var PRJ_DIMS = [['prjFilterStatus','status'],['prjFilterRegion','region'],
                ['prjFilterCountry','country'],['prjFilterCompany','company'],
                ['prjFilterProject','project']];

/* Unit conversion (base = mmt). Capacity is an ANNUAL figure -> rate units use a
   365.25-day year; production is a per-period flow -> rate units use period days. */
function prjUnit(){ var s=document.getElementById('prjUnit'); return (s&&s.value)?s.value:PRJ.default_unit; }
function prjUnitCfg(){ return PRJ.unit_config[prjUnit()] || PRJ.unit_config[PRJ.default_unit]; }
function prjUnitLabel(){ var c=prjUnitCfg(); return c.isRate?c.rateLabel:c.volLabel; }
function prjConvCap(mmt){ if(mmt==null||mmt===''||isNaN(mmt)) return null; var c=prjUnitCfg(); return c.isRate?(mmt/365.25)*c.rateFactor:mmt*c.volFactor; }
function prjConvFlow(mmt,days){ if(mmt==null||isNaN(mmt)) return 0; var c=prjUnitCfg(); return c.isRate?(mmt/days)*c.rateFactor:mmt*c.volFactor; }
function prjOnUnitChange(){ prjRenderSummary(); prjRenderTree(); prjOutlookRender(); }

function prjEsc(s){ return String(s==null?'':s).replace(/&/g,'&amp;').replace(/</g,'&lt;'); }
function prjNum(v,dec){ if(v==null||v===''||isNaN(v)) return '—'; return Number(v).toFixed(dec).replace(/\B(?=(\d{3})+(?!\d))/g,','); }
function prjPct(v){ if(v==null||v===''||isNaN(v)) return '—'; return Math.round(Number(v)*100)+'%'; }
function prjStake(v){ if(v==null||v===''||isNaN(v)) return '—'; return (Number(v)*100).toFixed(1)+'%'; }
var PRJ_MONTHS = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'];
function prjStart(s){
    if(!s) return '—';
    var m=String(s).match(/^(\d{4})-(\d{2})/);
    if(!m) return String(s);
    return PRJ_MONTHS[parseInt(m[2],10)-1] + ' ' + m[1];   // "Jan 2029"
}
// Owner colors: operator gets Palissy dark blue; partners cycle a palette.
var PRJ_OWNER_COLORS = ['#272962','#539648','#C00000','#258EEB','#E5B83A','#0C5B19','#8E44AD','#92591C','#2A9D8F','#708B5A'];
function prjSum(arr,key){ var s=0; for(var i=0;i<arr.length;i++){ var v=arr[i][key]; if(typeof v==='number'&&!isNaN(v)) s+=v; } return s; }
function prjFieldVal(p,dim){ return dim==='project'?p.name : dim==='status'?p.status : dim==='region'?p.region : dim==='country'?p.country : null; }
function prjCap(s){ return s.charAt(0).toUpperCase()+s.slice(1); }
function prjContainerId(scope,dim){ return (scope==='chart'?'prjcFilter':'prjFilter')+prjCap(dim); }
function prjMatchDim(p,dim,scope){
    var sel=prjFilters[scope][dim]; if(!sel.length) return true;
    if(dim==='company') return p.companies.some(function(c){return sel.indexOf(c)>=0;});
    return sel.indexOf(prjFieldVal(p,dim))>=0;
}
function prjMatchExcept(p,exceptDim,scope){ return PRJ_DIMS.every(function(d){ return d[1]===exceptDim || prjMatchDim(p,d[1],scope); }); }
function prjFilteredFor(scope){ return PRJ.projects.filter(function(p){ return PRJ_DIMS.every(function(d){ return prjMatchDim(p,d[1],scope); }); }); }
function prjFiltered(){ return prjFilteredFor('table'); }   // assumptions + summary always use table filters
function prjAvailableValues(dim,scope){
    var out=[], seen={};
    PRJ.projects.forEach(function(p){
        if(!prjMatchExcept(p,dim,scope)) return;
        var vals = dim==='company' ? p.companies : [prjFieldVal(p,dim)];
        vals.forEach(function(v){ if(v!=null && !seen[v]){ seen[v]=1; out.push(v); } });
    });
    return out;
}

function prjInit(){
    PRJ = DATA;
    prjFilters = { table:{status:[],region:[],country:[],company:[],project:[]},
                   chart:{status:[],region:[],country:[],company:[],project:[]} };
    prjChartLinked = true;
    prjExpanded = {}; prjOutExpanded = {}; prjSubtab = 'outlook';
    prjChangeMode = 'pct';
    var us = document.getElementById('prjUnit');
    if (us) { us.innerHTML=''; PRJ.units.forEach(function(u){ us.innerHTML += '<option value="'+u+'"'+(u===PRJ.default_unit?' selected':'')+'>'+u+'</option>'; }); }
    if (prjRangeChart) { prjRangeChart.destroy(); prjRangeChart = null; }
    if (prjStackChart) { prjStackChart.destroy(); prjStackChart = null; }
    prjBuildFilters('table');
    prjBuildFilters('chart');
    prjCloneOutlookControls('prjcViewBy','prjcPeriod');   // fill chart view/period option lists
    prjOutlookPopulateRange();
    prjOutlookPopulateLookback();
    prjSetLink(true);                                     // start linked (chart controls hidden)
    prjSetSubtab('outlook');
    prjApplyFilters();
}
// table filters changed -> re-render assumptions + production + change + (charts if linked)
function prjApplyFilters(){ prjUpdateFilterUI('table'); prjRenderSummary(); prjRenderTree(); prjOutlookRender(); }
// chart filters changed -> re-render only the charts
function prjChartApply(){ prjUpdateFilterUI('chart'); prjOutlookCharts(); }

function prjBuildFilters(scope){
    prjMakeFilter(scope, prjContainerId(scope,'status'),  'status',  PRJ.status_order);
    prjMakeFilter(scope, prjContainerId(scope,'region'),  'region',  PRJ.region_order);
    prjMakeFilter(scope, prjContainerId(scope,'country'), 'country', PRJ.countries);
    prjMakeFilter(scope, prjContainerId(scope,'company'), 'company', PRJ.companies);
    prjMakeFilter(scope, prjContainerId(scope,'project'), 'project', PRJ.projects.map(function(p){return p.name;}).slice().sort());
}
function prjMakeFilter(scope, containerId, dimKey, items){
    var container=document.getElementById(containerId); if(!container) return;
    container.innerHTML=''; container.classList.add('multi-select');
    var button=document.createElement('button'); button.type='button'; button.className='ms-button'; container.appendChild(button);
    var panel=document.createElement('div'); panel.className='ms-panel'; container.appendChild(panel);
    var search=document.createElement('input'); search.type='text'; search.className='ms-search'; search.placeholder='Search…';
    search.addEventListener('input', function(){ prjUpdateFilterUI(scope); });
    search.addEventListener('click', function(e){ e.stopPropagation(); });
    panel.appendChild(search);
    var allLbl=document.createElement('label'); allLbl.className='ms-item ms-all'; allLbl.setAttribute('data-val','__ALL__');
    var allCb=document.createElement('input'); allCb.type='checkbox';
    var allSp=document.createElement('span'); allSp.textContent='All';
    allLbl.appendChild(allCb); allLbl.appendChild(allSp);
    allCb.addEventListener('change', function(){ prjToggle(scope, dimKey,'__ALL__'); });
    panel.appendChild(allLbl);
    var allDiv=document.createElement('div'); allDiv.className='ms-divider'; panel.appendChild(allDiv);
    items.forEach(function(it){
        var lbl=document.createElement('label'); lbl.className='ms-item'; lbl.setAttribute('data-val', it);
        var cb=document.createElement('input'); cb.type='checkbox';
        var sp=document.createElement('span'); sp.textContent=it;
        lbl.appendChild(cb); lbl.appendChild(sp);
        cb.addEventListener('change', function(){ prjToggle(scope, dimKey, it); });
        panel.appendChild(lbl);
    });
    button.addEventListener('click', function(ev){ ev.stopPropagation(); container.classList.toggle('open'); });
}
// Most-granular selected dimension -> the view the charts should snap to.
function prjSuggestView(scope){
    var f=prjFilters[scope];
    if(f.company.length || f.project.length) return 'project';
    if(f.country.length) return 'country';
    return 'region';
}
function prjToggle(scope, dimKey, val){
    if(val==='__ALL__'){ prjFilters[scope][dimKey]=[]; }
    else { var arr=prjFilters[scope][dimKey], i=arr.indexOf(val); if(i>=0) arr.splice(i,1); else arr.push(val); }
    // Selecting a company -> project view; a country -> country view; region -> region.
    // (A status toggle doesn't change the view.) Applies to the matching scope's
    // view dropdown — table when linked drives both table + charts.
    if(dimKey!=='status'){
        var vbEl=document.getElementById(scope==='chart' ? 'prjcViewBy' : 'prjViewBy');
        if(vbEl) vbEl.value=prjSuggestView(scope);
    }
    if(scope==='table') prjApplyFilters(); else prjChartApply();
}
function prjResetFilters(scope){
    prjFilters[scope]={status:[],region:[],country:[],company:[],project:[]};
    if(scope==='table') prjApplyFilters(); else prjChartApply();
}
function prjUpdateFilterUI(scope){
    PRJ_DIMS.forEach(function(d){
        var dim=d[1];
        var container=document.getElementById(prjContainerId(scope,dim)); if(!container) return;
        var avail=prjAvailableValues(dim,scope);
        var search=container.querySelector('.ms-search'); var q=((search&&search.value)||'').toLowerCase();
        var items=container.querySelectorAll('label.ms-item');
        for(var i=0;i<items.length;i++){
            var val=items[i].getAttribute('data-val');
            var cb=items[i].querySelector('input');
            if(val==='__ALL__'){ if(cb) cb.checked=(prjFilters[scope][dim].length===0); items[i].classList.remove('ms-unavail'); continue; }
            var checked=prjFilters[scope][dim].indexOf(val)>=0;
            if(cb) cb.checked=checked;
            var ok=(avail.indexOf(val)>=0 || checked) && (!q || val.toLowerCase().indexOf(q)>=0);
            items[i].classList.toggle('ms-unavail', !ok);
        }
        var btn=container.querySelector('.ms-button'); var sel=prjFilters[scope][dim];
        btn.textContent = sel.length===0 ? 'All' : (sel.length<=2 ? sel.join(', ') : sel.length+' selected');
        btn.title=sel.join(', ');
    });
}

function prjRenderSummary(){ /* summary pills removed from the Projects sub-tab per spec */ }

function prjBadge(status){
    var cls=String(status||'').toLowerCase().replace(/[^a-z]+/g,'-').replace(/^-|-$/g,'');
    return '<span class="prj-badge s-'+cls+'">'+prjEsc(status)+'</span>';
}
function prjRowAgg(cls, key, indent, label, n, unr, rk, exp){
    return '<tr class="'+cls+'" data-prj-x="'+prjEsc(key)+'">'
        + '<td><span class="prj-arrow'+(exp?' expanded':'')+'">&#9654;</span> '+prjEsc(label)
        + ' <span style="color:#9395A2;font-weight:normal;">('+n+')</span></td>'
        + '<td></td><td>'+prjNum(prjConvCap(unr),1)+'</td><td>'+prjNum(prjConvCap(rk),1)+'</td><td></td><td></td><td></td></tr>';
}
function prjRowProject(p, exp){
    return '<tr class="prj-row-project" data-prj-x="p|'+prjEsc(p.name)+'">'
        + '<td><span class="prj-arrow'+(exp?' expanded':'')+'">&#9654;</span> '+prjEsc(p.name)+'</td>'
        + '<td style="text-align:left;">'+prjBadge(p.status)+'</td>'
        + '<td>'+prjNum(prjConvCap(p.unrisked),1)+'</td><td>'+prjNum(prjConvCap(p.risked),1)+'</td>'
        + '<td>'+prjPct(p.cos)+'</td><td>'+prjStart(p.start)+'</td>'
        + '<td>'+prjEsc(p.operator||'—')+'</td></tr>';
}
function prjRowDetail(p){
    var owners = p.owners || [];
    // Stake bar (only for owners with a numeric stake; grey out any unallocated remainder).
    var bar='', allocated=0, anyStake=false;
    owners.forEach(function(o,i){
        if(typeof o.stake==='number' && !isNaN(o.stake)){
            anyStake=true; allocated+=o.stake;
            bar+='<span style="width:'+(o.stake*100).toFixed(2)+'%;background:'+PRJ_OWNER_COLORS[i%PRJ_OWNER_COLORS.length]+'" title="'+prjEsc(o.name)+' '+prjStake(o.stake)+'"></span>';
        }
    });
    if(anyStake && allocated<0.999) bar+='<span style="width:'+((1-allocated)*100).toFixed(2)+'%;background:#cdd0db" title="Unallocated"></span>';
    var barHtml = anyStake ? '<div class="prj-stakebar">'+bar+'</div>' : '';
    var ownerList = owners.length
        ? owners.map(function(o,i){
            var sw='<span class="sw" style="background:'+PRJ_OWNER_COLORS[i%PRJ_OWNER_COLORS.length]+'"></span>';
            var tag = i===0 ? '<span class="tag">Operator</span>' : '';
            return '<div class="prj-owner">'+sw+'<span class="nm">'+prjEsc(o.name)+tag+'</span><span class="pct">'+prjStake(o.stake)+'</span></div>';
          }).join('')
        : '<div class="prj-owner"><span class="nm" style="color:#9395A2;">Not available</span></div>';

    function stat(k,v){ return '<div class="prj-stat"><span class="k">'+k+'</span><span class="v">'+v+'</span></div>'; }
    var stats = stat('Start', prjStart(p.start))
        + stat('Chance of success', prjPct(p.cos))
        + stat('Utilisation forecast', prjPct(p.util_forecast))
        + stat('Utilisation decline', prjPct(p.util_decline))
        + stat('Unrisked capacity', prjNum(prjConvCap(p.unrisked),2)+' '+prjUnitLabel())
        + stat('Risked capacity', prjNum(prjConvCap(p.risked),2)+' '+prjUnitLabel());

    return '<tr class="prj-detail"><td colspan="7"><div class="prj-detail-inner">'
        + '<div class="prj-detail-col owners"><h4>Ownership</h4>'+barHtml+'<div class="prj-owners">'+ownerList+'</div></div>'
        + '<div class="prj-detail-col"><h4>Key assumptions</h4><div class="prj-stats">'+stats+'</div></div>'
        + '</div></td></tr>';
}
function prjRenderTree(){
    var f=prjFiltered(), ul=prjUnitLabel();
    document.getElementById('prjTableHead').innerHTML =
        '<tr><th>Project</th><th style="text-align:left;">Status</th><th>Unrisked ('+ul+')</th>'
        + '<th>Risked ('+ul+')</th><th>CoS</th><th>Start</th><th>Operator</th></tr>';
    var byRegion={};
    f.forEach(function(p){ (byRegion[p.region]=byRegion[p.region]||[]).push(p); });
    var html='';
    PRJ.region_order.forEach(function(region){
        var rp=byRegion[region]; if(!rp||!rp.length) return;
        var rExp=!!prjExpanded['r|'+region];
        html+=prjRowAgg('prj-row-region','r|'+region,0,region,rp.length,prjSum(rp,'unrisked'),prjSum(rp,'risked'),rExp);
        if(!rExp) return;
        var byC={}; rp.forEach(function(p){ (byC[p.country]=byC[p.country]||[]).push(p); });
        Object.keys(byC).sort(function(a,b){ return prjSum(byC[b],'risked')-prjSum(byC[a],'risked'); }).forEach(function(country){
            var cp=byC[country]; var ckey='c|'+region+'|'+country; var cExp=!!prjExpanded[ckey];
            html+=prjRowAgg('prj-row-country',ckey,1,country,cp.length,prjSum(cp,'unrisked'),prjSum(cp,'risked'),cExp);
            if(!cExp) return;
            cp.slice().sort(function(a,b){ return (b.risked||0)-(a.risked||0); }).forEach(function(p){
                var pExp=!!prjExpanded['p|'+p.name];
                html+=prjRowProject(p,pExp);
                if(pExp) html+=prjRowDetail(p);
            });
        });
    });
    if(!html){
        html='<tr><td colspan="7" style="text-align:center;color:#9395A2;padding:20px;">No projects match the current filters.</td></tr>';
    } else {
        // Grand total row (always shown): summed unrisked + risked of the filtered set.
        html+='<tr class="prj-row-total"><td>Grand total ('+f.length+')</td><td></td>'
            + '<td>'+prjNum(prjConvCap(prjSum(f,'unrisked')),1)+'</td>'
            + '<td>'+prjNum(prjConvCap(prjSum(f,'risked')),1)+'</td><td></td><td></td><td></td></tr>';
    }
    document.getElementById('prjTableBody').innerHTML=html;
}

function prjSetSubtab(t){
    prjSubtab=t;
    document.getElementById('prjSubtabProjects').classList.toggle('active', t==='projects');
    document.getElementById('prjSubtabOutlook').classList.toggle('active', t==='outlook');
    document.getElementById('prjPaneProjects').classList.toggle('active', t==='projects');
    document.getElementById('prjPaneOutlook').classList.toggle('active', t==='outlook');
}

// Region/Country/Project expand toggles (event delegation on the tree).
document.addEventListener('click', function(e){
    var tr=e.target.closest('tr[data-prj-x]'); if(!tr) return;
    var key=tr.getAttribute('data-prj-x');
    prjExpanded[key]=!prjExpanded[key];
    prjRenderTree();
});

/* ---- Supply Outlook: project-level production table + change table + charts.
        View by Region/Country/Project; unit/period/from-to controls. ---- */
function prjDefaultRange(period, view, fs, ts){
    var from=0, to=view.col_meta.length-1;
    if(period==='Monthly'){
        for(var i=0;i<view.col_meta.length;i++){ if(view.col_meta[i].year===2025 && view.col_meta[i].month===1){ from=i; break; } }
        to=Math.min(view.col_meta.length-1, from+35);   // Jan 2025 -> +35 months
    } else {
        for(var i=0;i<view.col_meta.length;i++){ if(view.col_meta[i].year>=2020){ from=i; break; } }
        for(var i=view.col_meta.length-1;i>=0;i--){ if(view.col_meta[i].year<=2035){ to=i; break; } }
    }
    fs.value=from; ts.value=to;
}
function prjFillRangeOptions(view, fs, ts){
    var html=''; for(var i=0;i<view.col_meta.length;i++) html+='<option value="'+i+'">'+view.col_meta[i].label+'</option>';
    fs.innerHTML=html; ts.innerHTML=html;
}
function prjOutlookPopulateRange(){
    var sel=document.getElementById('prjPeriod'); if(!sel) return;
    var view=PRJ.production.views[sel.value]; if(!view) return;
    var fs=document.getElementById('prjFrom'), ts=document.getElementById('prjTo');
    prjFillRangeOptions(view, fs, ts); prjDefaultRange(sel.value, view, fs, ts);
}
function prjOnOutlookPeriod(){ prjOutlookPopulateRange(); prjOutlookRender(); }
function prjVisIndices(view, fromEl, toEl){
    var fi=parseInt(fromEl.value), ti=parseInt(toEl.value);
    if(isNaN(fi)||isNaN(ti)){ fi=0; ti=view.col_meta.length-1; }
    if(fi>ti){ var t=fi; fi=ti; ti=t; }
    var out=[]; for(var i=fi;i<=ti;i++) out.push(i); return out;
}

// Shared Region/Country/Project tree builder. cellsFn(names) -> the value <td>s.
function prjBuildTree(viewBy, leaves, prod, vis, cellsFn){
    function nm(p){ return p.name; }
    function groupBy(list,key){ var g={}; list.forEach(function(p){ (g[p[key]]=g[p[key]]||[]).push(p); }); return g; }
    function tot(names){ var s=0; for(var n=0;n<names.length;n++){ var b=prod[names[n]]; if(b) for(var k=0;k<vis.length;k++) s+=b[vis[k]]; } return s; }
    function byTotDesc(a,b){ return tot([b.name])-tot([a.name]); }
    function aggRow(cls,key,label,n,names,exp){
        return '<tr class="'+cls+'" data-prj-ox="'+prjEsc(key)+'"><td><span class="prj-arrow'+(exp?' expanded':'')+'">&#9654;</span> '
            + prjEsc(label)+' <span style="color:#9395A2;font-weight:normal;">('+n+')</span></td>'+cellsFn(names)+'</tr>';
    }
    function projRow(p){ return '<tr class="prj-row-project"><td>'+prjEsc(p.name)+'</td>'+cellsFn([p.name])+'</tr>'; }
    var body='';
    if(viewBy==='project'){
        leaves.slice().sort(byTotDesc).forEach(function(p){ body+=projRow(p); });
    } else if(viewBy==='country'){
        var byC=groupBy(leaves,'country');
        Object.keys(byC).sort(function(a,b){ return tot(byC[b].map(nm))-tot(byC[a].map(nm)); }).forEach(function(c){
            var cExp=!!prjOutExpanded['c|'+c];
            body+=aggRow('prj-row-country','c|'+c,c,byC[c].length,byC[c].map(nm),cExp);
            if(cExp) byC[c].slice().sort(byTotDesc).forEach(function(p){ body+=projRow(p); });
        });
    } else {
        var byR=groupBy(leaves,'region');
        PRJ.region_order.forEach(function(region){
            var rp=byR[region]; if(!rp||!rp.length) return;
            var rExp=!!prjOutExpanded['r|'+region];
            body+=aggRow('prj-row-region','r|'+region,region,rp.length,rp.map(nm),rExp);
            if(!rExp) return;
            var byC=groupBy(rp,'country');
            Object.keys(byC).sort(function(a,b){ return tot(byC[b].map(nm))-tot(byC[a].map(nm)); }).forEach(function(c){
                var ckey='r|'+region+'|'+c; var cExp=!!prjOutExpanded[ckey];
                body+=aggRow('prj-row-country',ckey,c,byC[c].length,byC[c].map(nm),cExp);
                if(cExp) byC[c].slice().sort(byTotDesc).forEach(function(p){ body+=projRow(p); });
            });
        });
    }
    var allNames=leaves.map(nm);
    if(allNames.length) body+='<tr class="prj-row-total"><td>Total ('+allNames.length+')</td>'+cellsFn(allNames)+'</tr>';
    return body;
}

function prjOutlookRender(){
    if(!PRJ || !PRJ.production) return;
    var period=document.getElementById('prjPeriod').value;
    var view=PRJ.production.views[period]; if(!view) return;
    var vis=prjVisIndices(view, document.getElementById('prjFrom'), document.getElementById('prjTo'));
    var viewBy=document.getElementById('prjViewBy').value, ul=prjUnitLabel();
    var prod={}; view.rows.forEach(function(r){ prod[r.label]=r.base; });
    var leaves=prjFiltered().filter(function(p){ return prod[p.name]; });

    var hHtml='<tr><th>Production ('+ul+')<span class="unit-label"> ('+period+')</span></th>';
    for(var i=0;i<vis.length;i++) hHtml+='<th>'+(view.short_columns[vis[i]]||view.columns[vis[i]])+'</th>';
    document.getElementById('prjOutlookHead').innerHTML=hHtml+'</tr>';

    function cellsFn(names){
        var s='';
        for(var k=0;k<vis.length;k++){ var ci=vis[k], sum=0; for(var n=0;n<names.length;n++){ var b=prod[names[n]]; if(b) sum+=b[ci]; } s+='<td>'+formatNum(prjConvFlow(sum, view.days[ci]), false)+'</td>'; }
        return s;
    }
    var body=prjBuildTree(viewBy, leaves, prod, vis, cellsFn);
    if(!body) body='<tr><td colspan="'+(vis.length+1)+'" style="text-align:center;color:#9395A2;padding:20px;">No projects match the current filters.</td></tr>';
    document.getElementById('prjOutlookBody').innerHTML=body;
    prjChangeRender();
    prjOutlookCharts();
}

/* Change table (under the production table): YoY/MoM/YTD/LTM/YTD-GY + %/abs,
   same engine as the gas balance growth table, on the production series. */
function prjSyncChangeType(){
    var period=document.getElementById('prjPeriod').value;
    var opts=GROWTH_OPTS[period]||[{k:'yoy',l:'Year-on-Year'}];
    var sel=document.getElementById('prjChangeType'); if(!sel) return;
    var cur=sel.value, valid=opts.some(function(o){return o.k===cur;});
    if(!valid) cur=opts[0].k;
    sel.innerHTML=opts.map(function(o){return '<option value="'+o.k+'"'+(o.k===cur?' selected':'')+'>'+o.l+'</option>';}).join('');
}
function prjSetChangeMode(m){
    prjChangeMode=m;
    document.getElementById('prjChangePct').className=m==='pct'?'active':'';
    document.getElementById('prjChangeAbs').className=m==='abs'?'active':'';
    prjChangeRender();
}
function prjChangeRender(){
    if(!PRJ || !PRJ.production) return;
    var period=document.getElementById('prjPeriod').value;
    var view=PRJ.production.views[period]; if(!view) return;
    prjSyncChangeType();
    var gt=document.getElementById('prjChangeType').value;
    var vis=prjVisIndices(view, document.getElementById('prjFrom'), document.getElementById('prjTo'));
    var viewBy=document.getElementById('prjViewBy').value;
    var prod={}; view.rows.forEach(function(r){ prod[r.label]=r.base; });
    var leaves=prjFiltered().filter(function(p){ return prod[p.name]; });

    var hHtml='<tr><th>Change<span class="unit-label"> ('+period+')</span></th>';
    for(var i=0;i<vis.length;i++) hHtml+='<th>'+(view.short_columns[vis[i]]||view.columns[vis[i]])+'</th>';
    document.getElementById('prjChangeHead').innerHTML=hHtml+'</tr>';

    var prevMode=growthMode; growthMode=prjChangeMode;   // computeGrowthCell + formatGrowth read this
    var ncol=view.col_meta.length;
    function aggBase(names){ var a=new Array(ncol); for(var i=0;i<ncol;i++) a[i]=0; for(var n=0;n<names.length;n++){ var b=prod[names[n]]; if(b) for(var i=0;i<ncol;i++) a[i]+=b[i]; } return a; }
    function cellsFn(names){
        var base=aggBase(names), s='';
        for(var k=0;k<vis.length;k++){
            var gv=computeGrowthCell(base, view.days, view.col_meta, vis[k], period, gt, false);
            if(gv===null){ s+='<td></td>'; continue; }
            var cls=gv>0.0001?'g-pos':(gv<-0.0001?'g-neg':'');
            s+='<td'+(cls?' class="'+cls+'"':'')+'>'+formatGrowth(gv,false)+'</td>';
        }
        return s;
    }
    var body=prjBuildTree(viewBy, leaves, prod, vis, cellsFn);
    if(!body) body='<tr><td colspan="'+(vis.length+1)+'" style="text-align:center;color:#9395A2;padding:20px;">No projects match the current filters.</td></tr>';
    document.getElementById('prjChangeBody').innerHTML=body;
    growthMode=prevMode;
}

/* ---- Chart-filters link/unlink: charts follow the table filters by default;
        when unlinked the chart section drives the charts with its own copy. ---- */
function prjCloneOutlookControls(viewId, periodId){
    var vb=document.getElementById(viewId);
    if(vb) vb.innerHTML='<option value="region">Region</option><option value="country">Country</option><option value="project">Project</option>';
    var pd=document.getElementById(periodId);
    if(pd){ var L={'Monthly':'Monthly','Quarterly':'Quarterly','Annual CY':'Annual (Calendar Year)','Gas Year':'Annual (Gas Year)','Winter':'Winter (Oct-Mar)','Summer':'Summer (Apr-Sep)'};
        pd.innerHTML=['Monthly','Quarterly','Annual CY','Gas Year','Winter','Summer'].map(function(p){return '<option value="'+p+'"'+(p==='Annual CY'?' selected':'')+'>'+L[p]+'</option>';}).join(''); }
}
function prjChartScope(){ return prjChartLinked ? 'table' : 'chart'; }
function prjChartProjects(){ return prjFilteredFor(prjChartScope()); }
function prjChartViewBy(){ return document.getElementById(prjChartLinked?'prjViewBy':'prjcViewBy').value; }
function prjChartPeriodVal(){ return document.getElementById(prjChartLinked?'prjPeriod':'prjcPeriod').value; }
function prjChartFromEl(){ return document.getElementById(prjChartLinked?'prjFrom':'prjcFrom'); }
function prjChartToEl(){ return document.getElementById(prjChartLinked?'prjTo':'prjcTo'); }
function prjcPopulateRange(copyFromTable){
    var period=document.getElementById('prjcPeriod').value; var view=PRJ.production.views[period]; if(!view) return;
    var fs=document.getElementById('prjcFrom'), ts=document.getElementById('prjcTo'); if(!fs||!ts) return;
    prjFillRangeOptions(view, fs, ts);
    if(copyFromTable && document.getElementById('prjPeriod').value===period){
        fs.value=document.getElementById('prjFrom').value; ts.value=document.getElementById('prjTo').value;
    } else { prjDefaultRange(period, view, fs, ts); }
}
function prjcOnPeriod(){ prjcPopulateRange(false); prjOutlookCharts(); }
function prjSetLink(linked){
    prjChartLinked=linked;
    var btn=document.getElementById('prjLinkBtn');
    if(btn){ btn.classList.toggle('active', linked); btn.textContent = linked ? '✓ Linked to table filters' : 'Unlinked — independent'; }
    var ctrls=document.getElementById('prjChartFilterControls'); if(ctrls) ctrls.style.display = linked ? 'none' : 'flex';
    if(!linked){
        ['status','region','country','company','project'].forEach(function(d){ prjFilters.chart[d]=prjFilters.table[d].slice(); });
        var vb=document.getElementById('prjcViewBy'); if(vb) vb.value=document.getElementById('prjViewBy').value;
        var pd=document.getElementById('prjcPeriod'); if(pd) pd.value=document.getElementById('prjPeriod').value;
        prjcPopulateRange(true);
        prjUpdateFilterUI('chart');
    }
    prjOutlookCharts();
}
function prjToggleLink(){ prjSetLink(!prjChartLinked); }

/* ---- Supply Outlook charts: range (gas-year cycle of total filtered production)
        + stacked area (production over time, grouped by the View-by dimension). ---- */
function prjOutlookPopulateLookback(){
    var sel=document.getElementById('prjRangeLookback'); if(!sel) return;
    var firstYear=PRJ.production.views.Monthly.col_meta[0].year;
    var maxLb=Math.max(3, currentGY()-firstYear);
    var html=''; for(var n=3;n<=maxLb;n++) html+='<option value="'+n+'"'+(n===5?' selected':'')+'>'+n+' years</option>';
    sel.innerHTML=html;
}
function prjOutlookCharts(){ if(!PRJ||!PRJ.production) return; prjRangeChartRender(); prjStackChartRender(); }

function prjRangeChartRender(){
    var unit=prjUnit(), ul=prjUnitLabel(), mv=PRJ.production.views.Monthly;
    var prod={}; mv.rows.forEach(function(r){ prod[r.label]=r.base; });
    var leaves=prjChartProjects().filter(function(p){ return prod[p.name]; });
    var n=mv.col_meta.length, tot=new Array(n); for(var i=0;i<n;i++) tot[i]=0;
    leaves.forEach(function(p){ var b=prod[p.name]; for(var i=0;i<n;i++) tot[i]+=b[i]; });
    var idxMap={}; for(var i=0;i<mv.col_meta.length;i++) idxMap[mv.col_meta[i].year+'-'+mv.col_meta[i].month]=i;
    var lbEl=document.getElementById('prjRangeLookback'); var lookback=lbEl?(parseInt(lbEl.value)||5):5;
    var cal=isCalCycle(prjChartPeriodVal());
    var cgy=cycleBase(cal), prevGY=cgy-1, nextGY=cgy+1, labels=cycleLabels(cal);
    var latestIdx=PRJ.latest_actual?PRJ.latest_actual.index:1e9;
    function findIdx(y,m){ var k=y+'-'+m; return (k in idxMap)?idxMap[k]:-1; }
    function valAt(idx){ if(idx==null||idx<0) return null; return applyUnitConversion(tot[idx], unit, mv.days[idx]); }
    // Show the FULL current cycle, including the forecast tail — production is a
    // reported+forecast series, so no actual-only cap here.
    function gyVals(gy){ var out=[]; for(var i=0;i<12;i++){ var a=cycleToActual(cal,gy,i); out.push(valAt(findIdx(a.year,a.month))); } return out; }
    var prevVals=gyVals(prevGY), curVals=gyVals(cgy), nextVals=gyVals(nextGY);
    var lbStart=cgy-lookback, lbEnd=cgy-1, avg=[],mn=[],mx=[];
    for(var mo=0;mo<12;mo++){ var s=[]; for(var gy=lbStart;gy<=lbEnd;gy++){ var a=cycleToActual(cal,gy,mo); var v=valAt(findIdx(a.year,a.month)); if(v!=null) s.push(v);} if(!s.length){avg.push(null);mn.push(null);mx.push(null);continue;} var su=0;for(var k=0;k<s.length;k++)su+=s[k]; avg.push(su/s.length); mn.push(Math.min.apply(null,s)); mx.push(Math.max.apply(null,s)); }
    document.getElementById('prjRangeTitle').textContent='Production Range — '+leaves.length+' projects ('+ul+')';
    var prevLabel=cycleSeriesLabel(cal, prevGY);
    var curLabel =cycleSeriesLabel(cal, cgy, ' (current)');
    var nextLabel=cycleSeriesLabel(cal, nextGY, ' (forecast)');
    var datasets=[
        { label: lookback+'y max',     data: mx,       borderColor:'rgba(0,0,0,0)', backgroundColor:'rgba(0,0,0,0)', pointRadius:0, fill:false, order:20 },
        { label: lookback+'y range',   data: mn,       borderColor:'rgba(0,0,0,0)', backgroundColor:'rgba(147,149,162,0.28)', pointRadius:0, fill:'-1', order:19 },
        { label: lookback+'y average', data: avg,      borderColor:'#272962', backgroundColor:'rgba(0,0,0,0)', borderWidth:1.8, pointRadius:0, fill:false, tension:0.25, order:4 },
        { label: prevLabel,            data: prevVals, borderColor:'#0C5B19', backgroundColor:'rgba(0,0,0,0)', borderWidth:1.8, pointRadius:2, fill:false, tension:0.25, order:3 },
        { label: curLabel,             data: curVals,  borderColor:'#C00000', backgroundColor:'rgba(0,0,0,0)', borderWidth:2.6, pointRadius:3, fill:false, tension:0.25, order:1 },
        { label: nextLabel,            data: nextVals, borderColor:'#539648', backgroundColor:'rgba(0,0,0,0)', borderWidth:1.8, borderDash:[6,4], pointRadius:2, fill:false, tension:0.25, order:2 },
    ];
    var opts=chartOptionsLine(unit); opts.scales.y.title.text=ul;
    if(prjRangeChart){ prjRangeChart.destroy(); prjRangeChart=null; }
    var c=document.getElementById('prjRangeCanvas'); if(!c) return;
    prjRangeChart=new Chart(c,{type:'line',data:{labels:labels,datasets:datasets},options:opts});
    renderHtmlLegend(prjRangeChart,'prjRangeLegend',{filter:function(l){return !(l||'').endsWith(' max');}});
}

function prjStackChartRender(){
    var period=prjChartPeriodVal(), view=PRJ.production.views[period]; if(!view) return;
    var vis=prjVisIndices(view, prjChartFromEl(), prjChartToEl());
    var viewBy=prjChartViewBy(), unit=prjUnit(), ul=prjUnitLabel();
    var prod={}; view.rows.forEach(function(r){ prod[r.label]=r.base; });
    var leaves=prjChartProjects().filter(function(p){ return prod[p.name]; });
    var groups=[];
    if(viewBy==='project'){ leaves.forEach(function(p){ groups.push({label:p.name,names:[p.name]}); }); }
    else if(viewBy==='country'){ var byC={}; leaves.forEach(function(p){ (byC[p.country]=byC[p.country]||[]).push(p.name); }); Object.keys(byC).forEach(function(c){ groups.push({label:c,names:byC[c]}); }); }
    else { var byR={}; leaves.forEach(function(p){ (byR[p.region]=byR[p.region]||[]).push(p.name); }); PRJ.region_order.forEach(function(r){ if(byR[r]) groups.push({label:r,names:byR[r]}); }); }
    var labels=[]; for(var k=0;k<vis.length;k++) labels.push(view.short_columns[vis[k]]||view.columns[vis[k]]);
    var datasets=[];
    for(var g=0;g<groups.length;g++){
        var grp=groups[g], data=[];
        for(var k=0;k<vis.length;k++){ var ci=vis[k], s=0; for(var nn=0;nn<grp.names.length;nn++){ var b=prod[grp.names[nn]]; if(b) s+=b[ci]; } var v=applyUnitConversion(s,unit,view.days[ci]); data.push(v<0?0:v); }
        var color = (viewBy==='region') ? (PRJ.region_colors[grp.label]||PRJ_OWNER_COLORS[g%PRJ_OWNER_COLORS.length]) : PRJ_OWNER_COLORS[g%PRJ_OWNER_COLORS.length];
        datasets.push({ label:grp.label, data:data, backgroundColor:fadeColor(color,0.8), borderColor:color, borderWidth:1, fill:(g===0)?'origin':'-1', stack:'prj', tension:0.2, pointRadius:0 });
    }
    document.getElementById('prjStackTitle').textContent='Production Build-up ('+ul+')';
    var opts=chartOptionsStacked(unit); opts.scales.y.title.text=ul;
    if(prjStackChart){ prjStackChart.destroy(); prjStackChart=null; }
    var c=document.getElementById('prjStackCanvas'); if(!c) return;
    prjStackChart=new Chart(c,{type:'line',data:{labels:labels,datasets:datasets},options:opts});
    renderHtmlLegend(prjStackChart,'prjStackLegend',{});
}

// Supply Outlook tree expand toggles.
document.addEventListener('click', function(e){
    var tr=e.target.closest('tr[data-prj-ox]'); if(!tr) return;
    var key=tr.getAttribute('data-prj-ox');
    prjOutExpanded[key]=!prjOutExpanded[key];
    prjOutlookRender();
});

/* === EVENTS === */
function onPeriodChange() {
    updateRangeSelectors();
    updateGrowthTypeSelector();
    updatePeriodNote();
    updateAll();
}
function onRangeChange() { updateAll(); }
function onGrowthTypeChange() { growthType=document.getElementById('growthTypeSelector').value; updateGrowthTable(); }
function updateAll() {
    if (DATA.projects_tab) return;   // projects tab renders via prjRender(), not here
    if (DATA.composite) { lngUpdateAll(); return; }
    updateTableValueModeLabels();
    updateEfficiencyNote();
    updateTable();
    updateGrowthTable();
    if (DATA.charts_enabled) updateCharts();
    else if (DATA.chart_groups) updateGasCharts();
}

document.addEventListener('click', function(e) {
    var td = e.target.closest('td[data-toggle]');
    if (td) {
        var idx=parseInt(td.getAttribute('data-toggle'));
        var label=DATA.hierarchy[idx].label;
        expandedMain[label]=!expandedMain[label];
        updateTable();
        return;
    }
    var td2 = e.target.closest('td[data-toggle-g]');
    if (td2) {
        var idx=parseInt(td2.getAttribute('data-toggle-g'));
        var label=DATA.hierarchy[idx].label;
        expandedGrowth[label]=!expandedGrowth[label];
        updateGrowthTable();
    }
});

function setupHover(tbodyId) {
    var container = document.getElementById(tbodyId);
    if (!container) return;
    container.addEventListener('mouseover', function(e) {
        var td = e.target.closest('td');
        if (!td) return;
        var tr = td.closest('tr');
        if (!tr) return;
        var colIdx = Array.from(tr.children).indexOf(td);
        var rowIdx = Array.from(container.children).indexOf(tr);
        if (colIdx===highlightedCol && rowIdx===highlightedRow) return;
        clearHighlight();
        highlightedCol=colIdx; highlightedRow=rowIdx;
        var allRows = container.querySelectorAll('tr');
        for (var r=0;r<allRows.length;r++) {
            var cells=allRows[r].children;
            if (colIdx<cells.length) {
                cells[colIdx].classList.add(r===rowIdx?'cell-highlight':'col-highlight');
            }
        }
        for (var c=0;c<tr.children.length;c++) {
            if (c!==colIdx) tr.children[c].classList.add('row-highlight');
        }
    });
    container.addEventListener('mouseleave', function() { clearHighlight(); });
}
function clearHighlight() {
    var els = document.querySelectorAll('.col-highlight,.row-highlight,.cell-highlight');
    for (var i=0;i<els.length;i++) els[i].classList.remove('col-highlight','row-highlight','cell-highlight');
    highlightedCol=-1; highlightedRow=-1;
}

document.addEventListener('DOMContentLoaded', function() {
    updateDatasetUI();
    rebuildUnitSelector();
    updateRangeSelectors();
    updateGrowthTypeSelector();
    updatePeriodNote();
    if (DATA.charts_enabled) rebuildChartControls();
    else if (DATA.chart_groups) rebuildGasChartControls();
    updateAll();
    setupHover('tableBody');
    setupHover('growthBody');
    setupHover('lngImpBody');
    setupHover('lngImpChBody');
    setupHover('lngExpBody');
    setupHover('lngExpChBody');
});
"""

    js = js.replace("__DATA_PLACEHOLDER__", data_json)
    js = js.replace("__EMBED_TABS__", json.dumps(EMBED_TABS))

    # Build tab toggle buttons: data datasets first, then embed tabs (Storage/LNG).
    toggle_buttons = []
    for k in ordered_keys:
        ds = datasets_by_key[k]
        toggle_buttons.append(
            f'<button class="dataset-toggle-btn{" active" if k == ordered_keys[0] else ""}" '
            f'id="btnDataset-{k}" onclick="switchDataset(\'{k}\')">{ds["tab_label"]}</button>'
        )
    for t in EMBED_TABS:
        toggle_buttons.append(
            f'<button class="dataset-toggle-btn" '
            f'id="btnDataset-{t["key"]}" onclick="showEmbedTab(\'{t["key"]}\')">{t["label"]}</button>'
        )

    html = '<!DOCTYPE html>\n<html lang="en">\n<head>\n'
    html += '<meta charset="UTF-8">\n'
    html += '<meta name="viewport" content="width=device-width, initial-scale=1.0">\n'
    html += '<title>Palissy Advisors</title>\n'
    html += '<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.7/dist/chart.umd.min.js"></script>\n'
    html += '<style>\n' + css + '\n</style>\n'
    html += '</head>\n<body>\n\n'

    html += '<div class="header">\n'
    html += '    <img src="data:image/png;base64,' + logo_b64 + '" alt="Palissy Advisors">\n'
    html += '    <h1 id="pageTitle"></h1>\n'
    html += '</div>\n\n'

    html += '<div class="dataset-toggle-bar">\n    '
    html += '\n    '.join(toggle_buttons)
    html += '\n</div>\n\n'

    html += '<div class="controls">\n'
    html += '    <div class="control-group">\n'
    html += '        <label>Period</label>\n'
    html += '        <select id="periodSelector" onchange="onPeriodChange()">\n'
    html += '            <option value="Monthly" selected>Monthly</option>\n'
    html += '            <option value="Quarterly">Quarterly</option>\n'
    html += '            <option value="Annual CY">Annual (Calendar Year)</option>\n'
    html += '            <option value="Gas Year">Annual (Gas Year)</option>\n'
    html += '            <option value="Winter">Winter (Oct-Mar)</option>\n'
    html += '            <option value="Summer">Summer (Apr-Sep)</option>\n'
    html += '        </select>\n'
    html += '    </div>\n'
    html += '    <div class="control-group">\n'
    html += '        <label>Unit</label>\n'
    html += '        <select id="unitSelector" onchange="updateAll()"></select>\n'
    html += '        <span class="unit-efficiency-note" id="unitEfficiencyNote"></span>\n'
    html += '    </div>\n'
    html += '    <div class="control-group">\n'
    html += '        <label>From</label>\n'
    html += '        <select id="rangeFrom" onchange="onRangeChange()"></select>\n'
    html += '    </div>\n'
    html += '    <div class="control-group">\n'
    html += '        <label>To</label>\n'
    html += '        <select id="rangeTo" onchange="onRangeChange()"></select>\n'
    html += '    </div>\n'
    html += '    <button class="reset-btn" onclick="resetRange()">Reset</button>\n'
    html += '</div>\n'
    html += '<div class="period-note" id="periodNote"></div>\n\n'

    html += '<div class="main-grid">\n'

    # === Top-left: table ===
    html += '<div class="grid-quad quad-table">\n'
    html += '    <div class="table-value-toggle-bar">\n'
    html += '        <div class="value-toggle">\n'
    html += '            <button id="btnTblGen" class="active" onclick="setTableValueMode(\'gen\')">Generation</button>\n'
    html += '            <button id="btnTblPct" onclick="setTableValueMode(\'pct\')">% of Total Fuel</button>\n'
    html += '        </div>\n'
    html += '    </div>\n'
    html += '    <div class="table-container" id="tableContainer">\n'
    html += '        <table id="dataTable">\n'
    html += '            <thead id="tableHead"></thead>\n'
    html += '            <tbody id="tableBody"></tbody>\n'
    html += '        </table>\n'
    html += '    </div>\n'
    html += '</div>\n\n'

    # === Bottom-left: growth table ===
    html += '<div class="grid-quad quad-growth">\n'
    html += '    <div class="growth-section">\n'
    html += '        <div class="growth-controls">\n'
    html += '            <div class="control-group">\n'
    html += '                <label>Change</label>\n'
    html += '                <select id="growthTypeSelector" onchange="onGrowthTypeChange()"></select>\n'
    html += '            </div>\n'
    html += '            <div class="growth-toggle">\n'
    html += '                <button id="btnPct" class="active" onclick="setGrowthMode(\'pct\')">% Change</button>\n'
    html += '                <button id="btnAbs" onclick="setGrowthMode(\'abs\')">Absolute</button>\n'
    html += '            </div>\n'
    html += '        </div>\n'
    html += '        <div class="growth-table-container">\n'
    html += '            <table>\n'
    html += '                <thead id="growthHead"></thead>\n'
    html += '                <tbody id="growthBody"></tbody>\n'
    html += '            </table>\n'
    html += '        </div>\n'
    html += '    </div>\n'
    html += '</div>\n\n'

    # === Top-right: seasonality chart (only shown for datasets with charts_enabled) ===
    html += '<div class="grid-quad quad-seasonality chart-quadrant">\n'
    html += '    <div class="chart-controls">\n'
    html += '        <div class="control-group">\n'
    html += '            <label id="rangeSourceLabel">Sources</label>\n'
    html += '            <div class="multi-select" id="msSeasonality"></div>\n'
    html += '        </div>\n'
    html += '        <div class="control-group">\n'
    html += '            <label>Average range</label>\n'
    html += '            <select id="seasonalityLookback" onchange="updateRangeChart()"></select>\n'
    html += '        </div>\n'
    html += '    </div>\n'
    html += '    <div class="chart-title" id="seasonalityTitle">Seasonality</div>\n'
    html += '    <div class="chart-canvas-wrap">\n'
    html += '        <canvas id="seasonalityCanvas"></canvas>\n'
    html += '    </div>\n'
    html += '    <div class="custom-legend" id="seasonalityLegend"></div>\n'
    html += '</div>\n\n'

    # === Bottom-right: buildup chart. Power uses Sources/Period/From/To;
    #     gas uses a chart-type toggle + Series picker (range/unit follow the table). ===
    html += '<div class="grid-quad quad-buildup chart-quadrant">\n'
    html += '    <div class="chart-controls buildup-controls-power">\n'
    html += '        <div class="control-group">\n'
    html += '            <label>Sources</label>\n'
    html += '            <div class="multi-select" id="msBuildup"></div>\n'
    html += '        </div>\n'
    html += '        <div class="control-group">\n'
    html += '            <label>Period</label>\n'
    html += '            <select id="buildupAgg" onchange="onBuildupAggChange()">\n'
    html += '                <option value="Monthly">Monthly</option>\n'
    html += '                <option value="Annual CY">Annual (Calendar Year)</option>\n'
    html += '                <option value="Gas Year">Annual (Gas Year)</option>\n'
    html += '            </select>\n'
    html += '        </div>\n'
    html += '        <div class="control-group">\n'
    html += '            <label>From</label>\n'
    html += '            <select id="buildupFrom" onchange="updateBuildupChart()"></select>\n'
    html += '        </div>\n'
    html += '        <div class="control-group">\n'
    html += '            <label>To</label>\n'
    html += '            <select id="buildupTo" onchange="updateBuildupChart()"></select>\n'
    html += '        </div>\n'
    html += '    </div>\n'
    html += '    <div class="chart-controls buildup-controls-gas">\n'
    html += '        <div class="control-group">\n'
    html += '            <label>Chart</label>\n'
    html += '            <div class="value-toggle">\n'
    html += '                <button id="btnGasBuild-bar" class="active" onclick="setGasBuildupType(\'bar\')">Stacked Bar</button>\n'
    html += '                <button id="btnGasBuild-area" onclick="setGasBuildupType(\'area\')">Stacked Area</button>\n'
    html += '                <button id="btnGasBuild-line" onclick="setGasBuildupType(\'line\')">Line</button>\n'
    html += '            </div>\n'
    html += '        </div>\n'
    html += '        <div class="control-group">\n'
    html += '            <label>Series</label>\n'
    html += '            <div class="multi-select" id="msGasBuildup"></div>\n'
    html += '        </div>\n'
    html += '    </div>\n'
    html += '    <div class="chart-title" id="buildupTitle">Generation</div>\n'
    html += '    <div class="chart-canvas-wrap">\n'
    html += '        <canvas id="buildupCanvas"></canvas>\n'
    html += '    </div>\n'
    html += '    <div class="custom-legend" id="buildupLegend"></div>\n'
    html += '</div>\n\n'

    # === Placeholder chart boxes (shown for chart_area datasets without live
    #     charts yet, e.g. gas). Occupy the same grid areas as the real charts. ===
    html += '<div class="chart-placeholder-quad left">\n'
    html += '    <div class="chart-title">Chart 1</div>\n'
    html += '    <div class="chart-placeholder-box">Chart to be configured</div>\n'
    html += '</div>\n\n'
    html += '<div class="chart-placeholder-quad right">\n'
    html += '    <div class="chart-title">Chart 2</div>\n'
    html += '    <div class="chart-placeholder-box">Chart to be configured</div>\n'
    html += '</div>\n\n'

    html += '</div>\n\n'  # end main-grid

    # === Global LNG composite grid: 4 stacked tables + 2 chart rows ===
    html += '<div class="lng-grid" id="lngGrid">\n'

    # Imports table
    html += '  <div class="lng-block">\n'
    html += '    <div class="lng-block-title">Imports</div>\n'
    html += '    <div class="table-container"><table><thead id="lngImpHead"></thead><tbody id="lngImpBody"></tbody></table></div>\n'
    html += '  </div>\n'

    # Imports change table
    html += '  <div class="lng-block">\n'
    html += '    <div class="growth-section">\n'
    html += '      <div class="growth-controls">\n'
    html += '        <span class="lng-block-title">Change in Imports</span>\n'
    html += '        <div class="control-group"><label>Change</label><select id="lngImpChangeType" onchange="lngOnChangeType(\'imports\')"></select></div>\n'
    html += '        <div class="growth-toggle">\n'
    html += '          <button id="lngImpModePct" class="active" onclick="lngSetMode(\'imports\',\'pct\')">% Change</button>\n'
    html += '          <button id="lngImpModeAbs" onclick="lngSetMode(\'imports\',\'abs\')">Absolute</button>\n'
    html += '        </div>\n'
    html += '      </div>\n'
    html += '      <div class="growth-table-container"><table><thead id="lngImpChHead"></thead><tbody id="lngImpChBody"></tbody></table></div>\n'
    html += '    </div>\n'
    html += '  </div>\n'

    # Exports table
    html += '  <div class="lng-block">\n'
    html += '    <div class="lng-block-title">Exports</div>\n'
    html += '    <div class="table-container"><table><thead id="lngExpHead"></thead><tbody id="lngExpBody"></tbody></table></div>\n'
    html += '  </div>\n'

    # Exports change table
    html += '  <div class="lng-block">\n'
    html += '    <div class="growth-section">\n'
    html += '      <div class="growth-controls">\n'
    html += '        <span class="lng-block-title">Change in Exports</span>\n'
    html += '        <div class="control-group"><label>Change</label><select id="lngExpChangeType" onchange="lngOnChangeType(\'exports\')"></select></div>\n'
    html += '        <div class="growth-toggle">\n'
    html += '          <button id="lngExpModePct" class="active" onclick="lngSetMode(\'exports\',\'pct\')">% Change</button>\n'
    html += '          <button id="lngExpModeAbs" onclick="lngSetMode(\'exports\',\'abs\')">Absolute</button>\n'
    html += '        </div>\n'
    html += '      </div>\n'
    html += '      <div class="growth-table-container"><table><thead id="lngExpChHead"></thead><tbody id="lngExpChBody"></tbody></table></div>\n'
    html += '    </div>\n'
    html += '  </div>\n'

    # Range charts row (Import range | Export range)
    html += '  <div class="lng-charts-row">\n'
    for side, pre, label in [("imports", "lngImp", "Import Range"), ("exports", "lngExp", "Export Range")]:
        html += '    <div class="grid-quad">\n'
        html += '      <div class="chart-controls">\n'
        html += f'        <div class="control-group"><label>Series</label><div class="multi-select" id="msLng{pre[3:]}Range"></div></div>\n'
        html += f'        <div class="control-group"><label>Average range</label><select id="{pre}RangeLookback" onchange="lngUpdateRange(\'{side}\')"></select></div>\n'
        html += '      </div>\n'
        html += f'      <div class="chart-title" id="{pre}RangeTitle">{label}</div>\n'
        html += f'      <div class="chart-canvas-wrap"><canvas id="{pre}RangeCanvas"></canvas></div>\n'
        html += f'      <div class="custom-legend" id="{pre}RangeLegend"></div>\n'
        html += '    </div>\n'
    html += '  </div>\n'

    # Stacked charts row (Import build-up | Export build-up)
    html += '  <div class="lng-charts-row">\n'
    for side, pre, label in [("imports", "lngImp", "Import Build-up"), ("exports", "lngExp", "Export Build-up")]:
        html += '    <div class="grid-quad">\n'
        html += '      <div class="chart-controls">\n'
        html += '        <div class="control-group"><label>Chart</label>\n'
        html += '          <div class="value-toggle">\n'
        html += f'            <button id="btn{pre}Stack-bar" class="active" onclick="lngSetStackType(\'{side}\',\'bar\')">Stacked Bar</button>\n'
        html += f'            <button id="btn{pre}Stack-area" onclick="lngSetStackType(\'{side}\',\'area\')">Stacked Area</button>\n'
        # Exports build-up: Bar/Area only (Line dropped per spec). Imports keeps Line.
        if side == "imports":
            html += f'            <button id="btn{pre}Stack-line" onclick="lngSetStackType(\'{side}\',\'line\')">Line</button>\n'
        html += '          </div>\n'
        html += '        </div>\n'
        html += f'        <div class="control-group"><label>Series</label><div class="multi-select" id="msLng{pre[3:]}Stack"></div></div>\n'
        html += f'        <div class="control-group"><label>From</label><select id="{pre}StackFrom" onchange="lngUpdateStack(\'{side}\')"></select></div>\n'
        html += f'        <div class="control-group"><label>To</label><select id="{pre}StackTo" onchange="lngUpdateStack(\'{side}\')"></select></div>\n'
        html += '      </div>\n'
        html += f'      <div class="chart-title" id="{pre}StackTitle">{label}</div>\n'
        html += f'      <div class="chart-canvas-wrap"><canvas id="{pre}StackCanvas"></canvas></div>\n'
        html += f'      <div class="custom-legend" id="{pre}StackLegend"></div>\n'
        html += '    </div>\n'
    html += '  </div>\n'

    html += '</div>\n\n'  # end lng-grid

    # Embed container for external dashboard tabs (Storage / LNG Sendout).
    # Shown only when an embed tab is active (CSS body.embed-active).
    html += '<div class="embed-container" id="embedContainer">\n'
    html += '    <iframe id="embedFrame" title="Palissy external dashboard" loading="lazy"></iframe>\n'
    html += '</div>\n\n'

    # === LNG Projects tab: sub-tabs + filter bar + Region/Country/Project tree ===
    html += '<div class="projects-tab-wrap" id="projectsTab">\n'
    html += '  <div class="prj-subtabs">\n'
    html += '    <button class="prj-subtab-btn active" id="prjSubtabOutlook" onclick="prjSetSubtab(\'outlook\')">Supply Outlook</button>\n'
    html += '    <button class="prj-subtab-btn" id="prjSubtabProjects" onclick="prjSetSubtab(\'projects\')">Projects</button>\n'
    html += '  </div>\n'
    html += '  <div class="prj-filter-bar">\n'
    html += '    <div class="prj-filter-group"><label>Unit</label><select id="prjUnit" onchange="prjOnUnitChange()"></select></div>\n'
    for lbl, fid in [("Status", "prjFilterStatus"), ("Region", "prjFilterRegion"),
                     ("Country", "prjFilterCountry"), ("Company", "prjFilterCompany"),
                     ("Project", "prjFilterProject")]:
        html += f'    <div class="prj-filter-group"><label>{lbl}</label><div class="multi-select" id="{fid}"></div></div>\n'
    html += '    <button class="prj-filter-reset" onclick="prjResetFilters(\'table\')">Reset filters</button>\n'
    html += '  </div>\n'
    # Supply Outlook pane (default) — production table by region/country/project
    html += '  <div class="prj-pane active" id="prjPaneOutlook">\n'
    html += '    <div class="prj-outlook-controls">\n'
    html += '      <div class="control-group"><label>View by</label>\n'
    html += '        <select id="prjViewBy" onchange="prjOutlookRender()">\n'
    html += '          <option value="region" selected>Region</option>\n'
    html += '          <option value="country">Country</option>\n'
    html += '          <option value="project">Project</option>\n'
    html += '        </select>\n'
    html += '      </div>\n'
    html += '      <div class="control-group"><label>Period</label>\n'
    html += '        <select id="prjPeriod" onchange="prjOnOutlookPeriod()">\n'
    html += '          <option value="Monthly">Monthly</option>\n'
    html += '          <option value="Quarterly">Quarterly</option>\n'
    html += '          <option value="Annual CY" selected>Annual (Calendar Year)</option>\n'
    html += '          <option value="Gas Year">Annual (Gas Year)</option>\n'
    html += '          <option value="Winter">Winter (Oct-Mar)</option>\n'
    html += '          <option value="Summer">Summer (Apr-Sep)</option>\n'
    html += '        </select>\n'
    html += '      </div>\n'
    html += '      <div class="control-group"><label>From</label><select id="prjFrom" onchange="prjOutlookRender()"></select></div>\n'
    html += '      <div class="control-group"><label>To</label><select id="prjTo" onchange="prjOutlookRender()"></select></div>\n'
    html += '    </div>\n'
    html += '    <div class="prj-table-container">\n'
    html += '      <table class="prj-table prj-outlook-table"><thead id="prjOutlookHead"></thead><tbody id="prjOutlookBody"></tbody></table>\n'
    html += '    </div>\n'
    # Change table (YoY/MoM/YTD/LTM/YTD-GY + %/abs) on the production series
    html += '    <div class="prj-change-controls">\n'
    html += '      <span class="lng-block-title">Change in production</span>\n'
    html += '      <div class="control-group"><label>Change</label><select id="prjChangeType" onchange="prjChangeRender()"></select></div>\n'
    html += '      <div class="growth-toggle">\n'
    html += '        <button id="prjChangePct" class="active" onclick="prjSetChangeMode(\'pct\')">% Change</button>\n'
    html += '        <button id="prjChangeAbs" onclick="prjSetChangeMode(\'abs\')">Absolute</button>\n'
    html += '      </div>\n'
    html += '    </div>\n'
    html += '    <div class="prj-table-container">\n'
    html += '      <table class="prj-table prj-outlook-table"><thead id="prjChangeHead"></thead><tbody id="prjChangeBody"></tbody></table>\n'
    html += '    </div>\n'
    # Chart filters section: link-to-table toggle + (when unlinked) an independent filter set
    html += '    <div class="prj-chart-filters">\n'
    html += '      <div class="prj-chart-filters-head">\n'
    html += '        <span class="lng-block-title">Chart filters</span>\n'
    html += '        <button class="prj-link-btn active" id="prjLinkBtn" onclick="prjToggleLink()">✓ Linked to table filters</button>\n'
    html += '      </div>\n'
    html += '      <div class="prj-chart-filter-controls" id="prjChartFilterControls">\n'
    html += '        <div class="prj-filter-group"><label>View by</label><select id="prjcViewBy" onchange="prjOutlookCharts()"></select></div>\n'
    html += '        <div class="prj-filter-group"><label>Period</label><select id="prjcPeriod" onchange="prjcOnPeriod()"></select></div>\n'
    html += '        <div class="prj-filter-group"><label>From</label><select id="prjcFrom" onchange="prjOutlookCharts()"></select></div>\n'
    html += '        <div class="prj-filter-group"><label>To</label><select id="prjcTo" onchange="prjOutlookCharts()"></select></div>\n'
    for lbl, fid in [("Status", "prjcFilterStatus"), ("Region", "prjcFilterRegion"),
                     ("Country", "prjcFilterCountry"), ("Company", "prjcFilterCompany"),
                     ("Project", "prjcFilterProject")]:
        html += f'        <div class="prj-filter-group"><label>{lbl}</label><div class="multi-select" id="{fid}"></div></div>\n'
    html += '        <button class="prj-filter-reset" onclick="prjResetFilters(\'chart\')">Reset</button>\n'
    html += '      </div>\n'
    html += '    </div>\n'
    # Charts row: range (left) + stacked area (right)
    html += '    <div class="prj-outlook-charts">\n'
    html += '      <div class="grid-quad">\n'
    html += '        <div class="chart-controls prj-chart-head">\n'
    html += '          <div class="control-group"><label>Average range</label><select id="prjRangeLookback" onchange="prjOutlookCharts()"></select></div>\n'
    html += '          <div class="chart-title-inline" id="prjRangeTitle">Production Range</div>\n'
    html += '        </div>\n'
    html += '        <div class="chart-canvas-wrap"><canvas id="prjRangeCanvas"></canvas></div>\n'
    html += '        <div class="custom-legend" id="prjRangeLegend"></div>\n'
    html += '      </div>\n'
    html += '      <div class="grid-quad">\n'
    html += '        <div class="chart-controls prj-chart-head">\n'
    html += '          <div class="chart-title-inline" id="prjStackTitle">Production Build-up</div>\n'
    html += '        </div>\n'
    html += '        <div class="chart-canvas-wrap"><canvas id="prjStackCanvas"></canvas></div>\n'
    html += '        <div class="custom-legend" id="prjStackLegend"></div>\n'
    html += '      </div>\n'
    html += '    </div>\n'
    html += '  </div>\n'
    # Projects pane (assumptions)
    html += '  <div class="prj-pane" id="prjPaneProjects">\n'
    html += '    <div class="prj-table-container">\n'
    html += '      <table class="prj-table"><thead id="prjTableHead"></thead><tbody id="prjTableBody"></tbody></table>\n'
    html += '    </div>\n'
    html += '  </div>\n'
    html += '</div>\n\n'

    html += '<div class="footer">\n'
    html += '    <span>Source: Palissy Advisors</span>\n'
    html += '    &nbsp;&bull;&nbsp;\n'
    html += '    <span>Last updated: ' + generated + '</span>\n'
    html += '</div>\n\n'

    html += '<script>\n' + js + '\n</script>\n\n'
    html += '</body>\n</html>'
    return html


def main():
    print("=" * 60)
    print("Palissy Multi-Dataset Dashboard Generator")
    print("=" * 60)

    datasets_by_key = {}
    ordered_keys = []
    for config in DATASETS:
        print(f"\n--- Building dataset: {config['key']} ({config['title']}) ---")
        if config.get("composite"):
            blob = build_composite_blob(config)
        elif config.get("projects_tab"):
            blob = build_projects_blob(config)
        else:
            blob = build_dataset_blob(config)
        datasets_by_key[config["key"]] = blob
        ordered_keys.append(config["key"])

    print("\nLoading assets...")
    assets = load_assets()

    print("\nGenerating HTML...")
    html = generate_html(datasets_by_key, ordered_keys, assets)

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        f.write(html)

    print(f"\nDashboard saved to: {OUTPUT_FILE}")
    print(f"  File size: {os.path.getsize(OUTPUT_FILE)/1024:.0f} KB")
    print(f"  Datasets: {', '.join(ordered_keys)}")
    print(f"  Display range: {DISPLAY_START_YEAR} - {DISPLAY_END_YEAR}")
    print("=" * 60)
    print("Done! Open output/index.html in a browser.")


if __name__ == "__main__":
    main()
