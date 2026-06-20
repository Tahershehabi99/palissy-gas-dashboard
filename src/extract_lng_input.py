"""
Extract LNG monthly data from the master model into a clean INPUT workbook.

Source : INPUT/AKAP Global LNG Model.xlsx  (the user drops the latest master here;
         READ-ONLY — opened but never modified/saved)
Output : WORKING/lng_model_input.xlsx  (auto-generated clean extract)

Copied faithfully (values only), preserving original row/column positions so the
master's row numbers still line up:
  - 'Monthly Imports' : source rows 4-21, cols A..col 290 (Jan 2017 - Dec 2040)
        row 4  = month date header
        rows 7-21 = EU+UK (potential), EU+UK, China, Japan, South Korea, Taiwan,
                    India, Other Asia, LatAm, Middle East, Egypt, Turkey, RoW,
                    Unaccounted demand, Total
  - 'Monthly Exports' : source rows 4-44, cols A..col 293 (Jan 2017 - Dec 2040)
        row 4  = month date header (data begins at col F; B-E are 'Working' cols)
        rows 8-44 = Production (mmt) block by country, with regional subtotals
                    (Asia, LatAm, MENA and Europe, North America,
                     Sub-Saharan Africa) and Grand total (row 44)

NOTE: This is a one-shot hardcode of current data. When the user finalises the
paste/refresh methodology, this is the script to adjust.
"""
import openpyxl
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "INPUT" / "AKAP Global LNG Model.xlsx"
OUT = ROOT / "WORKING" / "lng_model_input.xlsx"   # auto-generated clean extract

# (source sheet, first row, last row, first col, last col)
BLOCKS = [
    ("Monthly Imports", 4, 21, 1, 290),
    ("Monthly Exports", 4, 44, 1, 293),
]


def main():
    # NOTE: load WITHOUT read_only. Random .cell() access on a read_only sheet is
    # pathologically slow (re-scans the stream each call), which hung earlier runs.
    src = openpyxl.load_workbook(SRC, read_only=False, data_only=True)
    # A SECOND load keeping formulas (data_only=False) so we can tell a hardcoded
    # utilisation cell (typed constant) from an engine-driven one (formula).
    srcf = openpyxl.load_workbook(SRC, read_only=False, data_only=False)
    out = openpyxl.Workbook()
    out.remove(out.active)

    for sheet, r0, r1, c0, c1 in BLOCKS:
        ws_in = src[sheet]
        ws_out = out.create_sheet(sheet)
        copied = 0
        for r in range(r0, r1 + 1):
            for c in range(c0, c1 + 1):
                v = ws_in.cell(row=r, column=c).value
                if v is not None:
                    ws_out.cell(row=r, column=c, value=v)
                    copied += 1
        print(f"{sheet}: rows {r0}-{r1}, cols {c0}-{c1} -> {copied} non-empty cells")

    build_projects(src, srcf, out)

    OUT.parent.mkdir(exist_ok=True)
    out.save(OUT)
    print(f"Saved {OUT}")


# ============================================================
# LNG PROJECTS (for the LNG Projects tab)
# ------------------------------------------------------------
# Build a clean projects dataset from the master and write three sheets:
#   'LNG Projects'        - one row per MAIN terminal (trains excluded), joined
#                           to the Assumptions tab metadata + computed Risked Cap.
#   'LNG Proj Production' - mains x months (Reported mmt profile, reported+forecast)
#   'LNG Proj Utilisation'- mains x months (Utilisation %)
# Trains (col C='T' in Monthly Exports 'Reported' section) are dropped entirely;
# only mains (col C='M') are kept. Country comes from that section's aggregate
# rows; region from REGION_OF; metadata is joined from Assumptions by name.
# ============================================================
REGION_OF = {
    "Brunei": "Asia", "Indonesia": "Asia", "Malaysia": "Asia",
    "Papua New Guinea": "Asia", "Timor Leste": "Asia",
    "Australia": "Australia",
    "Argentina": "LatAm", "Mexico": "LatAm", "Peru": "LatAm",
    "Suriname": "LatAm", "Trinidad": "LatAm", "Venezuela": "LatAm",
    "Algeria": "MENA and Europe", "Egypt": "MENA and Europe", "Israel": "MENA and Europe",
    "Mauritania": "MENA and Europe", "Norway": "MENA and Europe", "Oman": "MENA and Europe",
    "Qatar": "MENA and Europe", "UAE": "MENA and Europe",
    "Canada": "North America", "United States": "North America",
    "Russia": "Russia",
    "Angola": "Sub-Saharan Africa", "Cameroon": "Sub-Saharan Africa",
    "Congo (Rep.)": "Sub-Saharan Africa", "Equatorial Guinea": "Sub-Saharan Africa",
    "Mozambique": "Sub-Saharan Africa", "Nigeria": "Sub-Saharan Africa",
    "Tanzania": "Sub-Saharan Africa",
}
COUNTRIES = set(REGION_OF)

# Monthly Exports layout is located by ANCHORS (header text), not fixed row
# numbers, so the extraction survives projects being inserted (every row below an
# insertion shifts). All four project blocks sit BELOW the 'Reported (mmt)'
# header; the regional capacity copies up top say '(mmt)' (not '(mmtpa)') and are
# skipped by both the position and the text. If an anchor can't be found we raise
# — the preflight surfaces it as a layout change needing review rather than
# silently extracting the wrong rows. Each block still has a MAIN-named summary row
# carrying the model's phased train total; train rows differ in name and are ignored.
EXP_DATE_ROW = 4
EXP_COL0 = 6                           # data begins at col F (B-E are working cols)


def _norm(v):
    return v.strip().lower() if isinstance(v, str) else ""


def _find_row(ws, pred, lo=1, hi=None):
    hi = hi or ws.max_row
    for r in range(lo, hi + 1):
        if pred(_norm(ws.cell(r, 1).value)):
            return r
    return None


def _block_end(ws, start):
    # A project block runs from header+1 to the row before its 'Grand total'.
    for r in range(start, ws.max_row + 1):
        if _norm(ws.cell(r, 1).value) == "grand total":
            return r - 1
    return ws.max_row


def locate_exports(me):
    """Locate the four Monthly-Exports project blocks + data column span by header
    anchors. Raises RuntimeError (surfaced by the preflight) if any is missing."""
    rep = _find_row(me, lambda s: s == "reported (mmt)")
    if not rep:
        raise RuntimeError("Monthly Exports: 'Reported (mmt)' anchor not found — layout changed")
    # Project capacity/util headers all sit BELOW Reported (regional '(mmt)' copies
    # are above it; excluded by both position and the required 'mmtpa' text).
    # NOTE: distinguish unrisked vs risked EXPLICITLY — "unrisked" CONTAINS the
    # substring "risked", so a naive `"risked" not in s` test breaks the moment the
    # master's old typo ("Unisked Capacity (mmtpa)") is corrected to "Unrisked".
    # Match either spelling for unrisked, and exclude it from the risked match.
    unr = _find_row(me, lambda s: "capacity (mmtpa)" in s and ("unrisked" in s or "unisked" in s), lo=rep)
    rsk = _find_row(me, lambda s: "risked capacity (mmtpa)" in s and "unrisked" not in s and "unisked" not in s, lo=rep)
    utl = _find_row(me, lambda s: s == "utilisation (%)", lo=rep)
    for label, hr in (("Unrisked Capacity (mmtpa)", unr), ("Risked Capacity (mmtpa)", rsk), ("Utilisation (%)", utl)):
        if not hr:
            raise RuntimeError(f"Monthly Exports: '{label}' anchor not found below Reported — layout changed")
    last_col = EXP_COL0
    for c in range(EXP_COL0, me.max_column + 1):
        if isinstance(me.cell(EXP_DATE_ROW, c).value, datetime):
            last_col = c
    return {
        "col0": EXP_COL0, "col1": last_col,
        "rep": (rep + 1, _block_end(me, rep)),
        "unr": (unr + 1, _block_end(me, unr)),
        "rsk": (rsk + 1, _block_end(me, rsk)),
        "utl": (utl + 1, _block_end(me, utl)),
    }


def _clean(v):
    return v.strip() if isinstance(v, str) else v


def detect_forecast_start(src, dates):
    """Return the index in `dates` (the exports month axis) of the first FORECAST
    month. The Monthly Imports 'Unaccounted demand' row is EMPTY for actual months
    and populated for forecast months, so the first non-empty cell marks the first
    forecast month. Falls back to None if the marker can't be located."""
    mi = src["Monthly Imports"]
    dcols = [(c, mi.cell(EXP_DATE_ROW, c).value) for c in range(1, mi.max_column + 1)
             if isinstance(mi.cell(EXP_DATE_ROW, c).value, datetime)]
    if not dcols:
        return None
    ud_row = _find_row(mi, lambda s: s == "unaccounted demand", lo=EXP_DATE_ROW + 1)
    if not ud_row:
        return None
    first_fc = next((d for c, d in dcols if mi.cell(ud_row, c).value is not None), None)
    if first_fc is None:
        return None
    for i, d in enumerate(dates):
        if isinstance(d, datetime) and (d.year, d.month) == (first_fc.year, first_fc.month):
            return i
    return None


def build_projects(src, srcf, out):
    me = src["Monthly Exports"]
    mef = srcf["Monthly Exports"]        # same layout, formulas preserved
    asm = src["Assumptions"]
    loc = locate_exports(me)                 # anchor-found block boundaries
    c0, c1 = loc["col0"], loc["col1"]
    print(f"  anchors: Reported {loc['rep']}  Unrisked {loc['unr']}  Risked {loc['rsk']}  "
          f"Util {loc['utl']}  cols {c0}-{c1}")

    # --- Assumptions metadata, keyed by project name ---
    # cols: A name, B country, C status, D unrisked mmt, E unrisked bcf/d, F start,
    #       G CoS, H util forecast, I util decline, J decline start,
    #       W(23) operator, X(24) partners, Y(25) primary stake, Z-AF(26-32) S1-S7
    meta = {}
    for r in range(3, asm.max_row + 1):
        name = _clean(asm.cell(r, 1).value)
        if not name or name == "Total":
            continue
        start = asm.cell(r, 6).value
        meta[name] = {
            "status": _clean(asm.cell(r, 3).value) or "",
            "unrisked_mmt": asm.cell(r, 4).value,
            "unrisked_bcfd": asm.cell(r, 5).value,
            "start": start.strftime("%Y-%m-%d") if isinstance(start, datetime) else "",
            "cos": asm.cell(r, 7).value,
            "util_forecast": asm.cell(r, 8).value,
            "util_decline": asm.cell(r, 9).value,
            "decline_start": (lambda d: d.strftime("%Y-%m-%d") if isinstance(d, datetime) else "")(asm.cell(r, 10).value),
            "operator": _clean(asm.cell(r, 23).value) or "",
            "partners": _clean(asm.cell(r, 24).value) or "",
            "primary_stake": asm.cell(r, 25).value,
            "stakes": [asm.cell(r, c).value for c in range(26, 33)],  # S1..S7
        }

    # --- Utilisation series, keyed by name ---
    util_by_name = {}
    for r in range(loc["utl"][0], loc["utl"][1] + 1):
        name = _clean(me.cell(r, 1).value)
        if not name or name in COUNTRIES or name in ("Grand total",):
            continue
        util_by_name[name] = [me.cell(r, c).value for c in range(c0, c1 + 1)]

    # --- Unrisked / Risked Capacity series (mmtpa), keyed by main name. The model
    #     already carries the phased train-by-train total on each main-named row,
    #     so no derivation is needed downstream — end-of-period sampling only. ---
    def cap_block(r0, r1):
        d = {}
        for r in range(r0, r1 + 1):
            name = _clean(me.cell(r, 1).value)
            if not name or name in COUNTRIES or name in ("Grand total",):
                continue
            d.setdefault(name, [me.cell(r, c).value for c in range(c0, c1 + 1)])
        return d
    unrcap_by_name = cap_block(*loc["unr"])
    rkcap_by_name = cap_block(*loc["rsk"])

    # --- Walk the Reported section: accumulate project rows, assign country on the
    #     country aggregate row; keep mains (M) only. ---
    dates = [me.cell(EXP_DATE_ROW, c).value for c in range(c0, c1 + 1)]
    projects = []           # ordered list of main dicts
    pending = []            # rows since last country marker
    for r in range(loc["rep"][0], loc["rep"][1] + 1):
        name = _clean(me.cell(r, 1).value)
        if not name:
            continue
        marker = _clean(me.cell(r, 3).value)
        if marker in ("M", "T"):
            pending.append((name, marker, [me.cell(r, c).value for c in range(c0, c1 + 1)]))
        elif name in COUNTRIES:
            country = name
            for pname, pmark, series in pending:
                if pmark != "M":
                    continue                      # drop trains
                m = meta.get(pname, {})
                cos = m.get("cos")
                unr = m.get("unrisked_mmt")
                risked = (unr * cos) if (isinstance(unr, (int, float)) and isinstance(cos, (int, float))) \
                    else (unr if isinstance(unr, (int, float)) else None)  # producing: blank CoS -> 1.0
                projects.append({
                    "name": pname, "country": country, "region": REGION_OF.get(country, country),
                    "prod": series, "util": util_by_name.get(pname),
                    "unrcap": unrcap_by_name.get(pname), "rkcap": rkcap_by_name.get(pname),
                    **m, "risked_mmt": risked,
                })
            pending = []
        # else: regional/grand-total rows -> ignore (and reset so they don't leak)
        elif name in ("Grand total",):
            pending = []

    print(f"LNG Projects: {len(projects)} mains across "
          f"{len(set(p['country'] for p in projects))} countries")
    missing = [p['name'] for p in projects if p['name'] not in meta]
    if missing:
        print(f"  (no Assumptions match for {len(missing)}: {missing})")

    # --- Hardcoded-utilisation month mask per MAIN ---------------------------------
    # A forecast utilisation cell is "hardcoded" (fixed by Palissy, not engine-
    # derived) when it's a TYPED CONSTANT rather than a formula in the Monthly
    # Exports Utilisation block. The override can live on the main's own util row OR
    # on any of its trains (trains PRECEDE their main in the sheet and their hardcode
    # rolls up into the main's derived util). We emit, per main, the list of monthly
    # indices (into `dates`) that are fixed, so the dashboard's tweak engine can keep
    # those months frozen when the user changes the utilisation-decline rate.
    # Leading pre-start zeros are ignored; a constant 0 that appears AFTER production
    # has begun (a deliberate cease/shutdown, e.g. Cameroon FLNG) is counted.
    off = loc["utl"][0] - loc["rep"][0]          # util block aligns 1:1 with reported
    fc_start = detect_forecast_start(src, dates)
    if fc_start is None:
        fc_start = 0
        print("  WARNING: forecast boundary not found — hardcode mask spans all months")

    # main name -> {own: reported_row, trains: [reported_rows]} (trains precede main)
    main_rows, pend = {}, []
    for r in range(loc["rep"][0], loc["rep"][1] + 1):
        nm = _clean(me.cell(r, 1).value)
        if not nm:
            continue
        mk = _clean(me.cell(r, 3).value)
        if mk == "T":
            pend.append(r)
        elif mk == "M":
            main_rows.setdefault(nm, {"own": r, "trains": pend})
            pend = []
        elif nm in COUNTRIES or nm == "Grand total":
            pend = []

    def util_hc_indices(rows):
        hc = set()
        for rr in rows:
            ur = rr + off
            producing = False
            for i in range(len(dates)):
                c = c0 + i
                num = me.cell(ur, c).value
                num = num if isinstance(num, (int, float)) else None
                if num is not None and num > 1e-9:
                    producing = True
                if i < fc_start:
                    continue                      # history is frozen anyway
                fcell = mef.cell(ur, c).value
                if isinstance(fcell, str) and fcell.startswith("="):
                    continue                      # engine-driven
                if not isinstance(fcell, (int, float)):
                    continue                      # blank / text
                if num is not None and abs(num) > 1e-9:
                    hc.add(i)
                elif producing and num is not None and abs(num) <= 1e-9:
                    hc.add(i)                     # deliberate drop to zero
        return sorted(hc)

    hc_by_name = {}
    for p in projects:
        mr = main_rows.get(p["name"])
        if not mr:
            continue
        idx = util_hc_indices([mr["own"]] + mr["trains"])
        if idx:
            hc_by_name[p["name"]] = idx
    print(f"  hardcoded-utilisation mains (typed-constant forecast months): {len(hc_by_name)}"
          + (f" -> {sorted(hc_by_name)}" if hc_by_name else ""))

    # --- Write 'LNG Projects' metadata sheet ---
    ws = out.create_sheet("LNG Projects")
    headers = ["Project", "Country", "Region", "Status", "Unrisked (mmt)", "Unrisked (bcf/d)",
               "CoS", "Risked (mmt)", "Start date", "Util forecast", "Util decline",
               "Operator", "Partners", "Primary stake", "S1", "S2", "S3", "S4", "S5", "S6", "S7",
               "Decline start", "HC months"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    for i, p in enumerate(projects, start=2):
        row = [p["name"], p["country"], p["region"], p.get("status", ""),
               p.get("unrisked_mmt"), p.get("unrisked_bcfd"), p.get("cos"), p.get("risked_mmt"),
               p.get("start", ""), p.get("util_forecast"), p.get("util_decline"),
               p.get("operator", ""), p.get("partners", ""), p.get("primary_stake")]
        row += (p.get("stakes") or [None] * 7)[:7]
        row.append(p.get("decline_start", ""))           # col 22
        hc = hc_by_name.get(p["name"])                   # col 23: hardcoded-util months
        row.append(",".join(str(x) for x in hc) if hc else "")
        for c, v in enumerate(row, start=1):
            ws.cell(row=i, column=c, value=v)

    # --- Write series sheets (row 1 = dates from col B; col A = project name) ---
    for sheet_name, key in [("LNG Proj Production", "prod"), ("LNG Proj Utilisation", "util"),
                            ("LNG Proj Unrisked Cap", "unrcap"), ("LNG Proj Risked Cap", "rkcap")]:
        sh = out.create_sheet(sheet_name)
        for j, d in enumerate(dates, start=2):
            sh.cell(row=1, column=j, value=d)
        for i, p in enumerate(projects, start=2):
            sh.cell(row=i, column=1, value=p["name"])
            series = p.get(key)
            if series:
                for j, v in enumerate(series, start=2):
                    if v is not None:
                        sh.cell(row=i, column=j, value=v)
    print("LNG Projects sheets written: 'LNG Projects', 'LNG Proj Production', 'LNG Proj Utilisation'")


if __name__ == "__main__":
    main()
