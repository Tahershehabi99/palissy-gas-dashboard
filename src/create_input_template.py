"""
One-time script to extract monthly bcf data from the AKAP Global Gas Model
and create the clean INPUT template Excel file.

The user will paste updated data into this template going forward.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from copy import copy
import os

# Paths
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(SCRIPT_DIR)
SOURCE_FILE = os.path.join(PROJECT_DIR, "Context", "AKAP Global Gas Model.xlsx")
OUTPUT_FILE = os.path.join(PROJECT_DIR, "INPUT", "gas_model_input.xlsx")

def create_input_template():
    print("Reading source model...")
    wb_src = openpyxl.load_workbook(SOURCE_FILE, read_only=True, data_only=True)
    ws_src = wb_src['MASTER']

    # Find last column with date data (row 6)
    last_col = 1
    for row in ws_src.iter_rows(min_row=6, max_row=6, max_col=500, values_only=False):
        for cell in row:
            if cell.value is not None:
                last_col = cell.column
    print(f"Data spans columns B to {openpyxl.utils.get_column_letter(last_col)} ({last_col - 1} months)")

    # Read all the data we need
    # Row 4: Days
    # Row 6: Dates
    # Rows 8-28: bcf section (header + data, excluding stock change at row 29)
    rows_to_read = [4, 6] + list(range(8, 29))  # Days, Dates, bcf header, data rows 9-28

    data = {}
    for target_row in rows_to_read:
        row_data = []
        for row in ws_src.iter_rows(min_row=target_row, max_row=target_row, min_col=1, max_col=last_col, values_only=True):
            row_data = list(row)
        data[target_row] = row_data

    wb_src.close()

    # Create the output workbook
    print("Creating input template...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Monthly Data"

    # Styles
    header_font = Font(name='Calibri', size=10, bold=True)
    data_font = Font(name='Calibri', size=10)
    date_font = Font(name='Calibri', size=9, bold=True)
    label_font = Font(name='Calibri', size=10, bold=False)
    parent_font = Font(name='Calibri', size=10, bold=True)
    header_fill = PatternFill(start_color='272962', end_color='272962', fill_type='solid')
    header_text = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    thin_border = Border(
        bottom=Side(style='thin', color='D0D0D0')
    )

    # Row 1: Dates
    ws.cell(row=1, column=1, value="Date").font = header_font
    date_row = data[6]
    for col_idx in range(1, last_col):
        val = date_row[col_idx] if col_idx < len(date_row) else None
        cell = ws.cell(row=1, column=col_idx + 1, value=val)
        cell.font = date_font
        if val is not None:
            cell.number_format = 'MMM-YY'

    # Row 2: Days
    ws.cell(row=2, column=1, value="Days").font = header_font
    days_row = data[4]
    for col_idx in range(1, last_col):
        val = days_row[col_idx] if col_idx < len(days_row) else None
        cell = ws.cell(row=2, column=col_idx + 1, value=val)
        cell.font = data_font

    # Row 3: Empty separator
    # Row 4: bcf header
    ws.cell(row=4, column=1, value="bcf").font = Font(name='Calibri', size=10, bold=True, italic=True)

    # Rows 5-24: Data rows (from source rows 9-28)
    output_row = 5
    for src_row in range(9, 29):
        src_data = data[src_row]
        label = src_data[0] if src_data[0] is not None else ""

        # Style based on whether it's a parent row
        is_parent = label.startswith('+') or label.startswith('-') or label in ['Opening Storage', 'Closing Storage', 'Storage percentage']

        cell = ws.cell(row=output_row, column=1, value=label)
        cell.font = parent_font if is_parent else label_font
        cell.border = thin_border

        for col_idx in range(1, last_col):
            val = src_data[col_idx] if col_idx < len(src_data) else None
            cell = ws.cell(row=output_row, column=col_idx + 1, value=val)
            cell.font = data_font
            cell.border = thin_border
            if label == 'Storage percentage' and val is not None:
                cell.number_format = '0.0%'
            elif val is not None:
                cell.number_format = '#,##0.0'

        output_row += 1

    # Set column widths
    ws.column_dimensions['A'].width = 30
    for col_idx in range(2, last_col + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 12

    # Freeze panes: freeze column A and row 1
    ws.freeze_panes = 'B2'

    # Add an Instructions sheet
    ws_inst = wb.create_sheet("Instructions")
    instructions = [
        "PALISSY GAS MODEL - INPUT FILE",
        "",
        "HOW TO UPDATE:",
        "1. Open your AKAP Global Gas Model Excel file",
        "2. Go to the MASTER tab",
        "3. Copy the bcf section data (rows with Opening Storage through Storage percentage)",
        "4. Also copy the Days row and Date row",
        "5. Paste into the 'Monthly Data' sheet, matching the existing structure",
        "6. Save this file",
        "7. Double-click update_dashboard.bat in the project root folder",
        "",
        "IMPORTANT NOTES:",
        "- Only paste bcf values. All other units are calculated automatically.",
        "- If you add a new supply source (e.g., new pipeline), add it as a new row",
        "  between the existing children and the parent total row.",
        "  Example: Add 'US Pipeline' between 'Libya' and 'Reverse flow into Ukraine',",
        "  or between 'Reverse flow into Ukraine' and '+ Imports'.",
        "- Keep the + and - prefixes on parent/total rows.",
        "- The date range can extend beyond the current 2015-2040 range.",
        "",
        "DISPLAY RANGE:",
        "- The dashboard currently shows 2020-2030.",
        "- To change this, edit the config in src/generate_dashboard.py",
        "  (DISPLAY_START_YEAR and DISPLAY_END_YEAR variables).",
    ]
    for i, line in enumerate(instructions):
        ws_inst.cell(row=i + 1, column=1, value=line).font = Font(name='Calibri', size=11)
    ws_inst.column_dimensions['A'].width = 80

    # Save
    wb.save(OUTPUT_FILE)
    print(f"Input template saved to: {OUTPUT_FILE}")
    print(f"  - Sheet 'Monthly Data': {output_row - 5} data rows x {last_col - 1} months")
    print(f"  - Sheet 'Instructions': Usage guide")

if __name__ == "__main__":
    create_input_template()
